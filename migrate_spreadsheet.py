"""
migrate_spreadsheet.py
======================
AI-Prowler Job Tracker spreadsheet migration engine.

PURPOSE
-------
When a user already has an AI-Prowler_Job_Tracker.xlsx from a previous version,
the installer's onlyifdoesntexist flag correctly skips overwriting it. This means
structural improvements (new sheets, new columns, freeze pane fixes, formula
updates, number format standardisation) never reach existing users automatically.

This module bridges that gap. It is called by rag_gui.py at startup — once per
schema version bump — and migrates the user's live spreadsheet in-place while
guaranteeing that their data is NEVER lost.

SAFETY CONTRACT
---------------
1. A timestamped backup is made BEFORE any changes. If the backup fails, the
   migration aborts immediately — nothing is touched.
2. If Excel (or any process) has the file open and locked, we detect it early
   and abort with a clear message.
3. If ANY exception occurs during migration, the backup is restored atomically
   to the original path. The corrupted/partial file is saved as a separate
   .failed copy for diagnostics.
4. The new template is always copied alongside the user's file as
   AI-Prowler_Job_Tracker_TEMPLATE_v{VERSION}.xlsx so they can inspect it and
   Claude can help migrate manually if needed.
5. On success, spreadsheet_schema_version is written to config.json.
6. Every event is written to ~/.ai-prowler/spreadsheet_migration.log.

SCHEMA VERSIONS
---------------
0  — Pre-v9.1.0 (no version tracking, assumed for all existing installs)
1  — v9.1.0: Settings sheet added, date format standardised to MM/DD/YYYY,
             freeze panes corrected, Services_Pricing restored,
             AI-Prowler_Commands restored, example rows added

Future versions append _migrate_N_to_N1() functions and bump CURRENT_SCHEMA_VERSION.

COLUMN SAFETY
-------------
New columns are APPENDED to the right of existing columns on a sheet — they are
NEVER inserted in the middle, which would shift existing column indices and break
MCP tool lookups that find columns by header name. The MCP tools already do
header-name lookup (not fixed column index), so appending is safe.
"""
from __future__ import annotations

import datetime
import json
import logging
import os
import shutil
import sys
from pathlib import Path
from typing import Optional

# ══════════════════════════════════════════════════════════════════════════════
# SCHEMA VERSION
# ══════════════════════════════════════════════════════════════════════════════
#
# CURRENT_SCHEMA_VERSION is the version this copy of migrate_spreadsheet.py
# targets. It must be bumped every release that changes the spreadsheet
# structure (new sheets, new columns, renamed headers, new formulas).
#
# HOW VERSIONING WORKS
# --------------------
# Each user's config.json stores "spreadsheet_schema_version" (0 if absent).
# On startup, check_and_migrate() reads this, compares it to
# CURRENT_SCHEMA_VERSION, and runs every migration step in _MIGRATIONS that
# falls between the user's version and the current one — in sequence.
#
# Example: user is on v1, current is v5:
#   Runs: _migrate_1_to_2 → _migrate_2_to_3 → _migrate_3_to_4 → _migrate_4_to_5
#   Each function only handles one step — no need to know the starting version.
#
# ADDING A NEW VERSION (e.g. v9.2.0 adds a QB_Export sheet)
# ----------------------------------------------------------
# 1. Bump CURRENT_SCHEMA_VERSION to 2
# 2. Write _migrate_1_to_2(wb, template_wb) -> list[str]
#    - Use _ensure_sheet(), _ensure_columns(), _apply_sheet_protection()
#    - Return a list of human-readable change descriptions
#    - Only add/fix — never remove or modify existing data
# 3. Append (1, 2, _migrate_1_to_2) to _MIGRATIONS
# 4. Update AI-Prowler_Job_Tracker.xlsx template with the new structure
# 5. Run release.bat with --remanifest so the template ships with the update
#
# RULE: every _migrate_N_to_N1 function must be additive and idempotent.
# If it runs twice on the same workbook, the result must be identical to
# running it once. _ensure_sheet() and _ensure_columns() already guarantee
# this — they check before acting.
#
CURRENT_SCHEMA_VERSION = 1   # bump this each release that changes sheet structure

# ── Paths ──────────────────────────────────────────────────────────────────────
_CONFIG_PATH    = Path.home() / ".ai-prowler" / "config.json"
_LOG_PATH       = Path.home() / ".ai-prowler" / "spreadsheet_migration.log"
_APP_DIR        = Path(sys.executable).parent if getattr(sys, 'frozen', False) \
                  else Path(__file__).parent
_TEMPLATE_NAME  = "AI-Prowler_Job_Tracker.xlsx"

# ── Logger ─────────────────────────────────────────────────────────────────────
_log = logging.getLogger("spreadsheet_migration")
_log.setLevel(logging.DEBUG)


def _setup_log():
    _LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not _log.handlers:
        fh = logging.FileHandler(_LOG_PATH, encoding="utf-8")
        fh.setFormatter(logging.Formatter(
            "%(asctime)s  %(levelname)-8s  %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S"
        ))
        _log.addHandler(fh)


# ── Config helpers ─────────────────────────────────────────────────────────────

def _read_config() -> dict:
    try:
        if _CONFIG_PATH.exists():
            return json.loads(_CONFIG_PATH.read_text(encoding="utf-8-sig")) or {}
    except Exception:
        pass
    return {}


def _write_config(data: dict):
    _CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    existing = _read_config()
    existing.update(data)
    _CONFIG_PATH.write_text(
        json.dumps(existing, indent=2), encoding="utf-8"
    )


def _get_schema_version() -> int:
    return int(_read_config().get("spreadsheet_schema_version", 0))


def _set_schema_version(v: int):
    _write_config({"spreadsheet_schema_version": v})


def _get_spreadsheet_path() -> Optional[Path]:
    cfg = _read_config()
    p = cfg.get("default_spreadsheet_path", "")
    if p:
        return Path(p)
    # Fallback: default install location
    default = Path.home() / "Documents" / "AI-Prowler" / _TEMPLATE_NAME
    return default if default.exists() else None


def _get_template_path() -> Optional[Path]:
    """Return the bundled template in Program Files (app directory)."""
    t = _APP_DIR / _TEMPLATE_NAME
    return t if t.exists() else None


# ── File-lock detection ────────────────────────────────────────────────────────

def _is_file_locked(path: Path) -> bool:
    """Return True if another process has the file exclusively locked."""
    try:
        with open(path, "r+b"):
            return False
    except (IOError, OSError, PermissionError):
        return True


# ── Backup helper ──────────────────────────────────────────────────────────────

def _make_backup(path: Path) -> Path:
    """Copy path to a timestamped .bak file. Returns the backup path."""
    ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = path.parent / "_backups" / f"{path.stem}_pre_migration_v{CURRENT_SCHEMA_VERSION}_{ts}.xlsx"
    bak.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path, bak)
    return bak


# ── Migration helpers ──────────────────────────────────────────────────────────

def _get_headers(ws, hdr_row: int = 2) -> dict[str, int]:
    """Return {normalized_header: col_index} for a worksheet."""
    hdrs = {}
    for cell in ws[hdr_row]:
        if cell.value:
            norm = str(cell.value).replace("\n", " ").strip()
            hdrs[norm] = cell.column
    return hdrs


def _ensure_sheet(wb, name: str, template_wb, position: Optional[int] = None):
    """Add sheet `name` from template_wb if it doesn't exist in wb.
    Returns (ws, was_added)."""
    import openpyxl
    from copy import copy

    if name in wb.sheetnames:
        return wb[name], False

    _log.info(f"  Adding missing sheet: {name}")
    # Copy template sheet structure
    src = template_wb[name]
    if position is not None:
        ws = wb.create_sheet(name, position)
    else:
        ws = wb.create_sheet(name)

    # Copy rows, styles, column widths, freeze panes, row heights
    for row in src.iter_rows():
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                # `copy` (the stdlib copy.copy) is already imported at the top
                # of this function. A stray second import used to sit here —
                # `from openpyxl.styles import copy as _style_copy` — which
                # openpyxl doesn't provide at all, so it raised ImportError on
                # every styled cell and crashed the whole migration before it
                # could reach the working calls below. Removed.
                new_cell.font      = copy(cell.font)
                new_cell.fill      = copy(cell.fill)
                new_cell.border    = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    for col_letter, cd in src.column_dimensions.items():
        ws.column_dimensions[col_letter].width = cd.width

    for row_num, rd in src.row_dimensions.items():
        ws.row_dimensions[row_num].height = rd.height

    ws.freeze_panes = src.freeze_panes

    return ws, True


def _header_similarity(a: str, b: str) -> float:
    """Return a rough similarity score 0.0–1.0 between two header strings.
    Uses character-level overlap (Jaccard on trigrams) — fast, no dependencies.
    Scores above 0.6 suggest the headers are probably the same concept under
    a different name (e.g. 'Service Date' vs 'Job Date' scores ~0.45,
    'Service Date' vs 'ServiceDate' scores ~0.85).
    """
    def trigrams(s):
        s = s.lower().replace(" ", "").replace("\n", "")
        return set(s[i:i+3] for i in range(max(1, len(s)-2)))
    ta, tb = trigrams(a), trigrams(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def _ensure_columns(ws, template_ws, hdr_row: int = 2):
    """Add any columns present in template_ws but missing from ws.
    Appended to the right — never inserted mid-sheet.

    Renamed-column guard: before appending a template column that is
    missing from the user's sheet, we check whether any existing user
    column has a similar name (trigram Jaccard > 0.55). If so, we log
    a WARNING and skip the append — the user probably renamed the column
    and we don't want to create a confusing duplicate. The warning is
    included in result.changes so it surfaces in the startup dialog.

    Returns (added_list, warned_list).
    """
    from copy import copy
    from openpyxl.utils import get_column_letter

    existing_hdrs = _get_headers(ws, hdr_row)
    template_hdrs = _get_headers(template_ws, hdr_row)
    added   = []
    warned  = []

    # Build a normalised lookup of existing headers for similarity check
    existing_norms = list(existing_hdrs.keys())   # list of normalised strings

    for hdr_norm, t_col in template_hdrs.items():
        if hdr_norm in existing_hdrs:
            continue    # exact match — already present, nothing to do

        # ── Renamed-column guard ──────────────────────────────────────────────
        # Check if any existing column looks like a renamed version of this one
        best_score  = 0.0
        best_match  = ""
        for existing_norm in existing_norms:
            score = _header_similarity(hdr_norm, existing_norm)
            if score > best_score:
                best_score = score
                best_match = existing_norm

        SIMILARITY_THRESHOLD = 0.55
        if best_score >= SIMILARITY_THRESHOLD:
            msg = (f"Skipped adding column '{hdr_norm}' — "
                   f"existing column '{best_match}' looks like a renamed "
                   f"version (similarity {best_score:.2f}). "
                   f"If this is wrong, manually add '{hdr_norm}' to the sheet.")
            _log.warning(f"  RENAME-GUARD: {msg}")
            warned.append(msg)
            continue

        # ── Append the missing column ─────────────────────────────────────────
        new_col = ws.max_column + 1
        t_cell  = template_ws.cell(row=hdr_row, column=t_col)
        h_cell  = ws.cell(row=hdr_row, column=new_col)
        h_cell.value = t_cell.value
        if t_cell.has_style:
            h_cell.font           = copy(t_cell.font)
            h_cell.fill           = copy(t_cell.fill)
            h_cell.border         = copy(t_cell.border)
            h_cell.alignment      = copy(t_cell.alignment)
            h_cell.number_format  = t_cell.number_format
        t_letter = get_column_letter(t_col)
        n_letter = get_column_letter(new_col)
        if t_letter in template_ws.column_dimensions:
            ws.column_dimensions[n_letter].width = \
                template_ws.column_dimensions[t_letter].width
        added.append(hdr_norm)
        _log.info(f"    + Column added: '{hdr_norm}' → col {new_col}")

    return added, warned


def _apply_sheet_protection(ws, sheet_name: str,
                             user_added_cols: "set[int] | None" = None):
    """Lock header rows 1-2 against editing AND formatting in Excel.

    Protection model:
      - Rows 1-2 (title banner + column headers):
          LOCKED for editing — cannot rename, delete or type into cells
          LOCKED for formatting — alignment, number format, font are preserved
      - Rows 3+ (data rows):
          UNLOCKED for editing — fully editable as normal
          UNLOCKED for formatting — users can format their own data cells
      - Insert/delete COLUMNS blocked across the whole sheet
      - Insert/delete ROWS allowed — users must be able to add/remove data
      - No password — openpyxl migration scripts bypass protection natively.

    user_added_cols: set of 1-based column indices added by the user (not
    in the template). These columns are NOT locked — user owns their own
    columns. None means all header columns are locked (fresh sheet from
    template with no user additions).
    """
    from openpyxl.styles import Protection as _Prot

    hdr_row = 3 if sheet_name == "Route_Planner" else 2

    # Step 1: Unlock ALL cells — data rows and user-added columns stay editable
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.protection = _Prot(locked=False)
            except AttributeError:
                pass  # skip merged-cell shadow cells

    # Step 2: Re-lock header rows — but skip user-added columns
    for r in range(1, hdr_row + 1):
        for c in range(1, ws.max_column + 1):
            if user_added_cols and c in user_added_cols:
                continue   # user's own column — don't lock it
            try:
                ws.cell(row=r, column=c).protection = _Prot(locked=True,
                                                             hidden=False)
            except AttributeError:
                pass

    # Step 3: Enable sheet protection with precise permissions
    ws.protection.sheet                = True
    ws.protection.selectLockedCells    = False
    ws.protection.selectUnlockedCells  = False
    ws.protection.insertColumns        = True   # BLOCK — no adding columns
    ws.protection.deleteColumns        = True   # BLOCK — no removing columns
    ws.protection.insertRows           = False  # allow — add data rows
    ws.protection.deleteRows           = False  # allow — remove data rows
    ws.protection.formatCells          = True   # BLOCK on locked cells only
    ws.protection.formatColumns        = False  # ALLOW — users resize columns freely
    ws.protection.formatRows           = False  # ALLOW — users resize rows freely
    ws.protection.sort                 = False  # allow — users can sort data
    ws.protection.autoFilter           = False  # allow — users can use filters

    _log.info(f"    Protection applied to '{sheet_name}' "
              f"(template headers locked; user-added cols unlocked: "
              f"{sorted(user_added_cols) if user_added_cols else 'none'})")


def _fix_freeze_panes(ws, template_ws):
    """Set freeze_panes to match the template if different."""
    if ws.freeze_panes != template_ws.freeze_panes:
        _log.info(f"    freeze_panes: {ws.freeze_panes!r} → {template_ws.freeze_panes!r}")
        ws.freeze_panes = template_ws.freeze_panes


# ══════════════════════════════════════════════════════════════════════════════
# MIGRATION FUNCTIONS — one per schema version transition
# ══════════════════════════════════════════════════════════════════════════════

def _migrate_0_to_1(wb, template_wb) -> list[str]:
    """
    v9.1.0 migration:
    - Add Settings sheet (first position) if missing
    - Replace AI-Prowler_Commands wholesale from template
    - Ensure all existing TEMPLATE sheets have correct freeze panes
    - Add any missing TEMPLATE columns to existing sheets
    - Apply header protection to template sheets only
    - Preserve user-added sheets and user-added columns untouched
    - User-added column headers are NOT locked (user owns them)
    """
    changes = []

    # ── Identify which sheets are user-added (not in template) ───────────────
    template_sheet_names = set(template_wb.sheetnames)
    user_added_sheets    = {s for s in wb.sheetnames
                            if s not in template_sheet_names}
    if user_added_sheets:
        _log.info(f"  User-added sheets (preserved, not migrated): "
                  f"{sorted(user_added_sheets)}")
        changes.append(
            f"Preserved user-added sheet(s): {', '.join(sorted(user_added_sheets))}"
        )

    # 1. Settings sheet — must be first tab
    _, added = _ensure_sheet(wb, "Settings", template_wb, position=0)
    if added:
        changes.append("Added Settings sheet (business constants & tax rate)")

    # 2. AI-Prowler_Commands — ALWAYS replace wholesale from template.
    _log.info("  Replacing AI-Prowler_Commands from template (reference sheet)")
    if "AI-Prowler_Commands" in wb.sheetnames:
        del wb["AI-Prowler_Commands"]
        changes.append("Updated AI-Prowler_Commands (command reference refreshed "
                       "from new template)")
    else:
        changes.append("Added AI-Prowler_Commands (quick command reference)")
    _, _ = _ensure_sheet(wb, "AI-Prowler_Commands", template_wb)

    # 3. For every TEMPLATE sheet: fix freeze panes, add missing columns,
    #    apply protection (preserving user-added columns as unlocked).
    sheet_map = {
        "Customers":           2,
        "Jobs_Schedule":       2,
        "Invoices":            2,
        "Quotes":              2,
        "TimeLog":             2,
        "Route_Planner":       3,
        "Services_Pricing":    2,
        "Settings":            2,
    }

    for sheet_name, hdr_row in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        if sheet_name not in template_wb.sheetnames:
            continue

        ws  = wb[sheet_name]
        tws = template_wb[sheet_name]

        # Detect user-added columns BEFORE we add template columns,
        # so we know which columns were already there vs newly added.
        template_hdrs  = set(_get_headers(tws, hdr_row).keys())
        existing_hdrs  = _get_headers(ws, hdr_row)
        user_added_col_indices: set[int] = {
            col_idx for hdr_norm, col_idx in existing_hdrs.items()
            if hdr_norm not in template_hdrs
        }
        if user_added_col_indices:
            user_col_names = [h for h, c in existing_hdrs.items()
                              if c in user_added_col_indices]
            _log.info(f"  {sheet_name}: user-added column(s) "
                      f"(preserved, not locked): {user_col_names}")
            changes.append(
                f"{sheet_name}: preserved user-added column(s): "
                f"{', '.join(user_col_names)}"
            )

        _fix_freeze_panes(ws, tws)

        added_cols, warned_cols = _ensure_columns(ws, tws, hdr_row=hdr_row)
        if added_cols:
            changes.append(
                f"{sheet_name}: added column(s): {', '.join(added_cols)}"
            )
        if warned_cols:
            for w in warned_cols:
                changes.append(f"⚠ {sheet_name}: {w}")

        # Apply protection — user-added column indices are NOT locked
        _apply_sheet_protection(ws, sheet_name,
                                user_added_cols=user_added_col_indices or None)

    return changes


def _verify_migration(original_wb, migrated_wb) -> list[str]:
    """
    Compare the migrated workbook against the original to detect data corruption.

    Checks performed for every sheet that existed in the original:
      1. Sheet still exists in migrated file
      2. Row count is >= original (rows never deleted)
      3. Every cell value in every original data row is identical in the
         migrated file at the same sheet/row/col position
      4. No existing column header was renamed, moved, or removed
      5. Formula cells still have formulas (not silently converted to values)

    Returns a list of problem strings — empty list means clean.
    The comparison is done in-memory against the loaded workbook objects,
    not by re-reading the saved file, so it catches corruption that occurred
    during the migration logic itself (not openpyxl I/O errors on re-read).
    """
    problems = []

    # AI-Prowler_Commands is intentionally replaced wholesale from the
    # template on every migration — exclude it from data integrity checks.
    SKIP_VERIFY = {"AI-Prowler_Commands"}

    for sheet_name in original_wb.sheetnames:
        if sheet_name in SKIP_VERIFY:
            continue
        if sheet_name not in migrated_wb.sheetnames:
            problems.append(
                f"CRITICAL: Sheet '{sheet_name}' is MISSING from migrated file"
            )
            continue

        ws_orig = original_wb[sheet_name]
        ws_migr = migrated_wb[sheet_name]

        # ── 1. Row count ──────────────────────────────────────────────────────
        orig_rows  = ws_orig.max_row
        migr_rows  = ws_migr.max_row
        if migr_rows < orig_rows:
            problems.append(
                f"CRITICAL: '{sheet_name}' has {migr_rows} rows after migration "
                f"but {orig_rows} before — {orig_rows - migr_rows} row(s) lost"
            )

        # ── 2. Column headers intact ───────────────────────────────────────────
        hdr_row = 3 if sheet_name == "Route_Planner" else 2
        orig_hdrs  = {}
        migr_hdrs  = {}
        for cell in ws_orig[hdr_row]:
            if cell.value is not None:
                orig_hdrs[cell.column] = str(cell.value).strip()
        for cell in ws_migr[hdr_row]:
            if cell.value is not None:
                migr_hdrs[cell.column] = str(cell.value).strip()

        for col_idx, orig_hdr in orig_hdrs.items():
            migr_hdr = migr_hdrs.get(col_idx)
            if migr_hdr is None:
                problems.append(
                    f"CRITICAL: '{sheet_name}' col {col_idx} header "
                    f"'{orig_hdr}' is GONE after migration"
                )
            elif migr_hdr.replace("\n", " ") != orig_hdr.replace("\n", " "):
                problems.append(
                    f"WARNING: '{sheet_name}' col {col_idx} header changed: "
                    f"'{orig_hdr}' → '{migr_hdr}'"
                )

        # ── 3. Data cell values identical ─────────────────────────────────────
        # Only check rows that existed in the original; skip header rows.
        data_start = hdr_row + 1
        orig_max_col = ws_orig.max_column
        row_errors = 0

        for row_num in range(data_start, ws_orig.max_row + 1):
            for col_num in range(1, orig_max_col + 1):
                orig_cell = ws_orig.cell(row=row_num, column=col_num)
                migr_cell = ws_migr.cell(row=row_num, column=col_num)

                orig_val = orig_cell.value
                migr_val = migr_cell.value

                # Normalise for comparison — None and "" are equivalent
                orig_norm = orig_val if orig_val is not None else ""
                migr_norm = migr_val if migr_val is not None else ""

                if orig_norm == migr_norm:
                    continue

                # Tolerate float precision drift (e.g. 125.0 vs 125.000000001)
                try:
                    if (isinstance(orig_val, (int, float)) and
                            isinstance(migr_val, (int, float)) and
                            abs(float(orig_val) - float(migr_val)) < 1e-9):
                        continue
                except (TypeError, ValueError):
                    pass

                # ── Formula check ─────────────────────────────────────────────
                orig_is_formula = (isinstance(orig_val, str) and
                                   orig_val.startswith("="))
                migr_is_formula = (isinstance(migr_val, str) and
                                   migr_val.startswith("=")) if migr_val else False
                orig_was_formula = orig_is_formula

                if orig_was_formula and not migr_is_formula:
                    problems.append(
                        f"WARNING: '{sheet_name}' R{row_num}C{col_num}: "
                        f"formula '{orig_val}' became value '{migr_val}'"
                    )
                elif not orig_was_formula:
                    row_errors += 1
                    if row_errors <= 5:   # cap per-sheet noise at 5 examples
                        hdr_name = orig_hdrs.get(col_num, f"col {col_num}")
                        problems.append(
                            f"CRITICAL: '{sheet_name}' R{row_num} "
                            f"[{hdr_name}]: "
                            f"value changed {orig_norm!r} → {migr_norm!r}"
                        )
                    elif row_errors == 6:
                        problems.append(
                            f"  ... and more value mismatches in '{sheet_name}' "
                            f"(showing first 5 only)"
                        )

    return problems


# ══════════════════════════════════════════════════════════════════════════════
# MIGRATION TABLE
# Each entry: (from_version, to_version, migration_function)
# The loop in check_and_migrate() walks this list in order and runs every
# step between the user's current version and CURRENT_SCHEMA_VERSION.
# A user jumping from v0 to v5 runs all four steps automatically.
# NEVER remove or reorder entries — only append new ones.
# ══════════════════════════════════════════════════════════════════════════════

_MIGRATIONS = [
    (0, 1, _migrate_0_to_1),
    # ── Add future migrations below this line ─────────────────────────────────
    # Pattern:
    #   (1, 2, _migrate_1_to_2),   # v9.2.0: QB_Export sheet, Actual Cost col
    #   (2, 3, _migrate_2_to_3),   # v9.3.0: ...
    # Always append — never remove or reorder existing entries.
]


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

class MigrationResult:
    """Result object returned to rag_gui.py for dialog display."""
    def __init__(self):
        self.needed               = False
        self.success              = False
        self.changes              = []
        self.integrity_problems   = []
        self.error                = ""
        self.backup_path          = ""
        self.template_path        = ""
        self.log_path             = str(_LOG_PATH)
        self.from_version         = 0
        self.to_version           = CURRENT_SCHEMA_VERSION
        self.user_spreadsheet_path = ""   # shown in consent dialog
        self.backup_name_preview   = ""   # shown in consent dialog before backup is made


class MigrationPlan:
    """Preview of what migration will do — computed without touching any files.
    Shown to user in the consent dialog so they know exactly what will happen.
    """
    def __init__(self):
        self.needed               = False
        self.from_version         = 0
        self.to_version           = CURRENT_SCHEMA_VERSION
        self.user_path            = ""      # full path to user spreadsheet
        self.backup_name          = ""      # what the backup will be called
        self.backup_dir           = ""      # where the backup will go
        self.sheets_to_add        = []      # sheet names that will be added
        self.sheets_to_replace    = []      # sheets that will be replaced (Commands)
        self.sheets_preserved     = []      # user-added sheets left untouched
        self.columns_to_add       = {}      # {sheet: [col names]} to be added
        self.columns_preserved    = {}      # {sheet: [col names]} user-added
        self.freeze_fixes         = []      # sheets whose freeze pane will be fixed
        self.error                = ""      # non-empty if plan computation failed


def get_migration_plan() -> MigrationPlan:
    """Compute what migration would do WITHOUT touching any files.
    Returns a MigrationPlan for display in the consent dialog.
    """
    _setup_log()
    plan = MigrationPlan()

    current_ver = _get_schema_version()
    plan.from_version = current_ver

    if current_ver >= CURRENT_SCHEMA_VERSION:
        plan.needed = False
        return plan

    plan.needed = True

    user_path = _get_spreadsheet_path()
    if not user_path or not user_path.exists():
        plan.needed = False
        return plan

    plan.user_path = str(user_path)

    # Preview backup name
    import datetime as _pdt
    ts = _pdt.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak_dir  = user_path.parent / "_backups"
    bak_name = f"{user_path.stem}_pre_migration_v{CURRENT_SCHEMA_VERSION}_{ts}.xlsx"
    plan.backup_name = bak_name
    plan.backup_dir  = str(bak_dir)

    template_path = _get_template_path()
    if not template_path:
        plan.error = f"Template not found in {_APP_DIR}"
        return plan

    try:
        import openpyxl as _opx
        # NOTE: read_only=True was removed here — openpyxl's read-only
        # worksheets don't support .freeze_panes (raises AttributeError),
        # which was silently killing plan computation below with no log
        # output (this function has no _log calls of its own). The file
        # is small, so full-mode loading costs nothing.
        wb      = _opx.load_workbook(user_path, data_only=False)
        tmpl_wb = _opx.load_workbook(template_path, data_only=False)

        template_sheets = set(tmpl_wb.sheetnames)
        user_sheets     = set(wb.sheetnames)

        # Sheets to add
        plan.sheets_to_add     = [s for s in template_sheets
                                   if s not in user_sheets
                                   and s != "AI-Prowler_Commands"]
        plan.sheets_to_replace = ["AI-Prowler_Commands"]
        plan.sheets_preserved  = sorted(user_sheets - template_sheets)

        # Per-sheet column analysis
        sheet_map = {
            "Customers": 2, "Jobs_Schedule": 2, "Invoices": 2,
            "Quotes": 2, "TimeLog": 2, "Route_Planner": 3,
            "Services_Pricing": 2, "Settings": 2,
        }
        for sheet_name, hdr_row in sheet_map.items():
            if sheet_name not in user_sheets or sheet_name not in template_sheets:
                continue
            ws  = wb[sheet_name]
            tws = tmpl_wb[sheet_name]
            t_hdrs = set(_get_headers(tws, hdr_row).keys())
            e_hdrs = _get_headers(ws, hdr_row)
            user_cols = [h for h, _ in e_hdrs.items() if h not in t_hdrs]
            missing   = [h for h in t_hdrs if h not in e_hdrs]
            if user_cols:
                plan.columns_preserved[sheet_name] = user_cols
            if missing:
                plan.columns_to_add[sheet_name] = missing
            if ws.freeze_panes != tws.freeze_panes:
                plan.freeze_fixes.append(sheet_name)

        wb.close()
        tmpl_wb.close()
    except Exception as e:
        plan.error = str(e)

    return plan


def check_and_migrate() -> MigrationResult:
    """
    Main entry point called by rag_gui.py at startup.

    Checks the schema version in config.json, runs any needed migrations
    on the user's spreadsheet, and returns a MigrationResult describing
    what happened (or that nothing was needed).

    Thread-safe: may be called from a background thread. Does NOT interact
    with tkinter directly — rag_gui.py handles all dialog/UI work.
    """
    _setup_log()
    result = MigrationResult()

    current_ver = _get_schema_version()
    result.from_version = current_ver

    if current_ver >= CURRENT_SCHEMA_VERSION:
        _log.debug(f"Schema version {current_ver} is current — no migration needed.")
        return result   # nothing to do

    result.needed = True
    _log.info("=" * 60)
    _log.info(f"Migration needed: schema v{current_ver} → v{CURRENT_SCHEMA_VERSION}")

    # ── Locate user spreadsheet ────────────────────────────────────────────────
    user_path = _get_spreadsheet_path()
    if not user_path or not user_path.exists():
        _log.warning("No user spreadsheet found — nothing to migrate.")
        # No file = treat as already migrated (they'll get the template on
        # first install via the installer's onlyifdoesntexist logic)
        _set_schema_version(CURRENT_SCHEMA_VERSION)
        result.success    = True
        result.needed     = False
        return result

    _log.info(f"User spreadsheet: {user_path}")
    result.user_spreadsheet_path = str(user_path)

    # ── Locate template ────────────────────────────────────────────────────────
    template_path = _get_template_path()
    if not template_path:
        msg = (f"Template not found in {_APP_DIR}. "
               "Reinstall AI-Prowler to restore the template file.")
        _log.error(msg)
        result.error = msg
        return result

    _log.info(f"Template:         {template_path}")

    # ── Check file lock ────────────────────────────────────────────────────────
    if _is_file_locked(user_path):
        msg = ("The Job Tracker spreadsheet is open in Excel. "
               "Please close it and restart AI-Prowler to apply the update.")
        _log.warning(msg)
        result.error = msg
        return result

    # ── Backup ────────────────────────────────────────────────────────────────
    try:
        backup = _make_backup(user_path)
        result.backup_path = str(backup)
        _log.info(f"Backup created:   {backup}")
    except Exception as e:
        msg = f"Could not create backup: {e}\nMigration aborted — your file is unchanged."
        _log.error(msg)
        result.error = msg
        return result

    # ── Copy template alongside user file ─────────────────────────────────────
    template_dest = user_path.parent / \
        f"AI-Prowler_Job_Tracker_TEMPLATE_v{CURRENT_SCHEMA_VERSION}.xlsx"
    try:
        shutil.copy2(template_path, template_dest)
        result.template_path = str(template_dest)
        _log.info(f"Template copy:    {template_dest}")
    except Exception as e:
        _log.warning(f"Could not copy template alongside user file: {e}")
        # Non-fatal — continue with migration

    # ── Load workbooks ────────────────────────────────────────────────────────
    try:
        import openpyxl
        wb        = openpyxl.load_workbook(user_path, data_only=False)
        wb_orig   = openpyxl.load_workbook(user_path, data_only=False)  # clean reference copy
        tmpl_wb   = openpyxl.load_workbook(template_path, data_only=False)
    except Exception as e:
        msg = f"Could not open spreadsheet: {e}"
        _log.error(msg)
        result.error = msg
        return result

    # ── Run migrations in sequence ─────────────────────────────────────────────
    all_changes = []
    try:
        ver = current_ver
        for from_v, to_v, fn in _MIGRATIONS:
            if ver >= to_v:
                continue
            if ver != from_v:
                continue
            _log.info(f"Running migration v{from_v} → v{to_v} ...")
            changes = fn(wb, tmpl_wb)
            all_changes.extend(changes)
            for c in changes:
                _log.info(f"  ✓ {c}")
            ver = to_v

        # ── Verify data integrity before saving ───────────────────────────────
        _log.info("Running post-migration data integrity check...")
        problems = _verify_migration(wb_orig, wb)

        # Separate CRITICALs from WARNINGs
        criticals = [p for p in problems if p.startswith("CRITICAL")]
        warnings  = [p for p in problems if p.startswith("WARNING")]

        for p in problems:
            level = _log.critical if p.startswith("CRITICAL") else _log.warning
            level(f"  {p}")

        if criticals:
            # Data corruption detected — abort, restore backup
            raise RuntimeError(
                f"Data integrity check failed with {len(criticals)} critical "
                f"issue(s):\n" + "\n".join(criticals)
            )

        if warnings:
            _log.warning(
                f"  {len(warnings)} warning(s) — non-critical, proceeding with save"
            )
            all_changes.append(
                f"Integrity warnings (non-critical): {len(warnings)} — "
                f"see migration log for details"
            )
        else:
            _log.info("  Integrity check passed — no issues found")

        result.integrity_problems = problems   # expose all for dialog display

        # ── Save ──────────────────────────────────────────────────────────────
        wb.save(user_path)
        _log.info(f"Saved: {user_path}")

        # ── Verify we can re-open ──────────────────────────────────────────────
        openpyxl.load_workbook(user_path, read_only=True).close()
        _log.info("Verification: file opens OK after migration.")

        # ── Write schema version ───────────────────────────────────────────────
        _set_schema_version(CURRENT_SCHEMA_VERSION)
        _log.info(f"Schema version set to {CURRENT_SCHEMA_VERSION} in config.json")

        result.success = True
        result.changes = all_changes
        _log.info(f"Migration complete. {len(all_changes)} change(s) applied.")

    except Exception as e:
        # ── ROLLBACK ──────────────────────────────────────────────────────────
        _log.error(f"Migration FAILED: {e}", exc_info=True)

        # Save the failed partial file for diagnostics
        failed_path = user_path.parent / "_backups" / \
            f"{user_path.stem}_migration_FAILED_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            shutil.copy2(user_path, failed_path)
            _log.info(f"Partial file saved for diagnostics: {failed_path}")
        except Exception:
            pass

        # Restore the backup
        try:
            shutil.copy2(backup, user_path)
            _log.info(f"Backup restored to {user_path} — your data is safe.")
        except Exception as re:
            _log.critical(
                f"CRITICAL: Could not restore backup! "
                f"Manual restore needed from {backup}. Error: {re}"
            )

        result.error = (
            f"Migration failed: {e}\n\n"
            f"Your original spreadsheet has been restored from backup.\n"
            f"Backup location: {backup}\n"
            f"Template copy:   {template_dest}\n\n"
            f"You can ask Claude to help migrate your data to the new template, "
            f"or copy your data manually. See the log for details:\n{_LOG_PATH}"
        )

    return result

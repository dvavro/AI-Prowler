"""
scheduler_jobs.py
=================
Proactive alert job functions for AI-Prowler's background scheduler.
Personal mode only — automatically suppressed in server mode via GUI guard.

Each job function:
  - Accepts a config dict
  - Calls AI-Prowler Python functions directly (no MCP, no API cost)
  - Returns (subject: str, body_html: str) or None if nothing to report
  - Never raises — all exceptions caught internally
"""
from __future__ import annotations
import datetime, traceback, json
from pathlib import Path


# ── Helpers ──────────────────────────────────────────────────────────────────

def _today() -> str:
    return datetime.date.today().isoformat()

def _now_str() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

def _footer() -> str:
    return f"<hr><p style='color:gray;font-size:11px'>AI-Prowler Proactive Alert · {_now_str()}</p>"

def _ar_aging() -> str:
    try:
        from ai_prowler_mcp import get_ar_aging_report
        return get_ar_aging_report()
    except Exception:
        return ""

def _sms_replies() -> str:
    try:
        from ai_prowler_mcp import list_sms_contacts_with_replies
        return list_sms_contacts_with_replies()
    except Exception:
        return ""

def _weather(location: str) -> str:
    try:
        from ai_prowler_mcp import get_weather
        return get_weather(location=location)
    except Exception:
        return ""

def _job_rows(sheet: str = "Jobs_Schedule") -> list[str]:
    try:
        from ai_prowler_mcp import read_job_spreadsheet
        result = read_job_spreadsheet(sheet_name=sheet)
        return [l for l in result.splitlines()
                if l.strip() and not l.startswith(("=", "-", "#"))]
    except Exception:
        return []

def _todays_jobs_structured() -> list[dict]:
    """Read TODAY's rows directly from the Jobs_Schedule sheet as structured
    dicts (not the pre-formatted multi-line-per-job text _job_rows() returns),
    so callers can access each job's own City/State individually — needed for
    per-job weather cross-referencing in the Morning Briefing.

    Reuses the exact same header-detection (>=3 non-empty cells in the first
    5 rows) and Service-Date matching logic as the tested read_job_spreadsheet
    MCP tool, just returning structured rows instead of formatted text.

    Returns a list of dicts, each with whatever of these keys were found as
    actual columns (missing columns are simply absent, never guessed):
        customer, city, state, service_type, crew
    Empty list on any error, missing file, or no matching rows — callers
    must treat this as "couldn't determine, fall back to the generic
    report", never a hard failure. Personal-mode only, matching this
    module's own scope (see module docstring).
    """
    try:
        from ai_prowler_mcp import _get_default_spreadsheet_path
        import openpyxl as _opx
        import datetime as _dt

        fp = _get_default_spreadsheet_path()
        if not fp or not Path(fp).exists():
            return []

        wb = _opx.load_workbook(fp, data_only=True)
        if "Jobs_Schedule" not in wb.sheetnames:
            return []
        ws = wb["Jobs_Schedule"]

        header_row_idx = None
        headers: list = []
        for r in ws.iter_rows(min_row=1, max_row=5):
            non_empty = [c for c in r if c.value is not None]
            if len(non_empty) >= 3:
                header_row_idx = r[0].row
                headers = [str(c.value).strip().replace('\n', ' ') if c.value else ''
                          for c in r]
                break
        if header_row_idx is None or not headers:
            return []

        # Map the columns we actually care about — tolerant of the
        # "★ AI Route" suffix used on City/Address/ZIP headers.
        col_idx = {}
        for idx, h in enumerate(headers):
            hl = h.lower()
            if 'customer name' in hl or 'company' in hl:
                col_idx.setdefault('customer', idx)
            elif hl.startswith('city'):
                col_idx.setdefault('city', idx)
            elif hl.strip() == 'state':
                col_idx.setdefault('state', idx)
            elif 'service type' in hl:
                col_idx.setdefault('service_type', idx)
            elif 'crew' in hl or 'technician' in hl:
                col_idx.setdefault('crew', idx)
            elif 'service' in hl and 'date' in hl:
                col_idx.setdefault('service_date', idx)

        if 'service_date' not in col_idx:
            return []

        today = _dt.date.today()
        results = []
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            vals = [c.value for c in row]
            if all(v is None or str(v).strip() == '' for v in vals):
                continue

            cell_val = vals[col_idx['service_date']] if col_idx['service_date'] < len(vals) else None
            if cell_val is None:
                continue
            if isinstance(cell_val, (_dt.datetime, _dt.date)):
                cell_date = cell_val.date() if isinstance(cell_val, _dt.datetime) else cell_val
            else:
                cell_date = None
                for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%d/%m/%Y'):
                    try:
                        cell_date = _dt.datetime.strptime(str(cell_val).strip(), fmt).date()
                        break
                    except ValueError:
                        continue
            if cell_date != today:
                continue

            entry = {}
            for key, idx in col_idx.items():
                if key == 'service_date' or idx >= len(vals):
                    continue
                v = vals[idx]
                if v is not None and str(v).strip():
                    entry[key] = str(v).strip()
            if entry:
                results.append(entry)

        return results
    except Exception:
        return []

def _pending_tasks() -> list[dict]:
    try:
        p = Path.home() / ".ai-prowler" / "pending_tasks.json"
        if not p.exists():
            return []
        return json.loads(p.read_text(encoding="utf-8")) or []
    except Exception:
        return []


# ── Job functions ─────────────────────────────────────────────────────────────

def job_morning_briefing(config: dict):
    """Daily: jobs today (with per-job weather by town), overdue invoices,
    unanswered SMS, due tasks.

    Per-job weather (added v8.1.3): each of today's jobs is checked against
    the weather for ITS OWN City/State — not one fixed location — since a
    field-service day's jobs are frequently scattered across several towns.
    Weather is fetched once per UNIQUE town among today's jobs (dedup), not
    once per job, to avoid redundant API calls when multiple jobs share a
    town. Falls back to config['location'] only when a job has no City on
    file, or when the spreadsheet can't be read at all (personal-mode-only
    feature — see _todays_jobs_structured's docstring).
    """
    try:
        name = config.get("name", "David")
        loc  = config.get("location", "New Smyrna Beach, Florida")
        dow  = datetime.date.today().strftime("%A, %B %d")

        parts = [f"<h2>☀️ Good morning, {name}!</h2><p><b>{dow}</b></p><hr>"]

        # Today's jobs — structured, with per-job weather by town
        jobs = _todays_jobs_structured()
        if jobs:
            # Fetch weather once per unique town, not once per job.
            weather_cache: dict[str, str] = {}
            def _job_weather(job: dict) -> str:
                city, state = job.get("city", ""), job.get("state", "")
                job_loc = f"{city}, {state}".strip(", ") if (city or state) else loc
                if job_loc not in weather_cache:
                    weather_cache[job_loc] = _weather(job_loc)
                return weather_cache[job_loc]

            parts.append(f"<h3>📋 Today's Jobs ({len(jobs)})</h3><ul>")
            for j in jobs[:10]:
                label = j.get("customer", "Job")
                where = ", ".join(x for x in (j.get("city"), j.get("state")) if x)
                if where:
                    label += f" — {where}"
                if j.get("service_type"):
                    label += f" ({j['service_type']})"
                w = _job_weather(j)
                rain_flag = ""
                if w and any(x in w for x in ("⚠️", "rain", "Rain")):
                    rain_flag = " <b style='color:#c00'>⚠️ Rain risk</b>"
                parts.append(f"<li>{label}{rain_flag}</li>")
            parts.append("</ul>")
        else:
            parts.append("<p>📋 No jobs scheduled today.</p>")
            # No per-job locations to check — fall back to the single
            # configured location so the briefing still shows something.
            w = _weather(loc)
            if w:
                wl = [l for l in w.splitlines() if l.strip()][:4]
                parts.append(f"<h3>🌤️ Weather — {loc}</h3><p>{'<br>'.join(wl)}</p>")

        # Overdue invoices
        ar = _ar_aging()
        if ar:
            od = [l for l in ar.splitlines()
                  if any(x in l for x in ["31-60", "61-90", "90+"]) and l.strip()]
            if od:
                parts.append("<h3>⚠️ Overdue Invoices</h3><ul>")
                for l in od:
                    parts.append(f"<li style='color:red'>{l}</li>")
                parts.append("</ul>")

        # SMS replies
        sms = _sms_replies()
        if sms and "unread" in sms.lower():
            parts.append(f"<h3>💬 Unanswered Messages</h3><pre>{sms[:500]}</pre>")

        # Due analysis tasks
        tasks = [t for t in _pending_tasks()
                 if t.get("status") == "pending"
                 and t.get("next_due", t.get("created_at",""))[:10] <= _today()]
        if tasks:
            parts.append(f"<h3>🧠 Analysis Tasks Due ({len(tasks)})</h3><ul>")
            for t in tasks:
                parts.append(f"<li>{t.get('label','?')}</li>")
            parts.append("</ul><p><i>Open Claude and press Ctrl+V to run them.</i></p>")

        parts.append(_footer())
        return f"☀️ Morning Briefing — {dow}", "\n".join(parts)
    except Exception:
        return "⚠️ Morning Briefing Error", f"<pre>{traceback.format_exc()}</pre>"


def job_overdue_invoice_alert(config: dict):
    """Daily: silent unless invoices are 31+ days overdue."""
    try:
        ar = _ar_aging()
        if not ar:
            return None
        od = [l for l in ar.splitlines()
              if any(x in l for x in ["31-60", "61-90", "90+"]) and l.strip()]
        if not od:
            return None
        parts = ["<h2>⚠️ Overdue Invoice Alert</h2>",
                 "<table border='1' cellpadding='6'>"]
        for l in od:
            parts.append(f"<tr><td>{l}</td></tr>")
        parts += ["</table>",
                  "<p><i>Consider sending payment reminders via Claude + Square.</i></p>",
                  _footer()]
        return f"⚠️ Overdue Invoices — {_today()}", "\n".join(parts)
    except Exception:
        return None


def job_due_analysis_tasks(config: dict):
    """Daily: alert when scheduled analysis tasks are due or overdue."""
    try:
        due = [t for t in _pending_tasks()
               if t.get("status") == "pending"
               and t.get("next_due", t.get("created_at",""))[:10] <= _today()]
        if not due:
            return None
        parts = [f"<h2>🧠 {len(due)} Analysis Task(s) Due</h2><ul>"]
        for t in due:
            parts.append(f"<li><b>{t.get('label','?')}</b> "
                         f"(due: {t.get('next_due','?')[:10]}, "
                         f"schedule: {t.get('schedule','?')})</li>")
        parts += ["</ul>",
                  "<p><b>Open Claude and press Ctrl+V to run all pending tasks.</b></p>",
                  _footer()]
        return f"🧠 {len(due)} Task(s) Due — {_today()}", "\n".join(parts)
    except Exception:
        return None


def job_sms_reply_monitor(config: dict):
    """Every N hours: alert on unanswered customer messages."""
    try:
        sms = _sms_replies()
        if not sms or "unread" not in sms.lower():
            return None
        parts = ["<h2>💬 Unanswered Customer Messages</h2>",
                 f"<pre>{sms[:1000]}</pre>", _footer()]
        return f"💬 Unanswered Messages — {_now_str()}", "\n".join(parts)
    except Exception:
        return None


def job_weather_watch(config: dict):
    """Sunday: 5-day forecast."""
    try:
        loc = config.get("location", "New Smyrna Beach, Florida")
        w = _weather(loc)
        if not w:
            return None
        parts = [f"<h2>🌤️ Weekly Weather Watch</h2>",
                 f"<p><b>{loc}</b> — 5-day forecast:</p>",
                 f"<pre>{w[:1500]}</pre>",
                 "<p><i>Check your job schedule for outdoor jobs on rainy days.</i></p>",
                 _footer()]
        return f"🌤️ Weather Watch — Week of {_today()}", "\n".join(parts)
    except Exception:
        return None


def job_end_of_day_summary(config: dict):
    """Evening: jobs completed vs scheduled today."""
    try:
        rows  = _job_rows()
        today = [r for r in rows if _today() in r]
        done  = [r for r in today if any(
            w in r.lower() for w in ["complete", "done", "paid", "invoiced"])]
        open_ = [r for r in today if r not in done]

        parts = [f"<h2>🌙 End of Day — {_today()}</h2>"]
        if done:
            parts += [f"<h3>✅ Completed ({len(done)})</h3><ul>"]
            for r in done:
                parts.append(f"<li>{r}</li>")
            parts.append("</ul>")
        if open_:
            parts += [f"<h3>⏳ Still Open ({len(open_)})</h3><ul>"]
            for r in open_:
                parts.append(f"<li>{r}</li>")
            parts += ["</ul>",
                      "<p><i>Don\'t forget to log time entries and mark jobs complete.</i></p>"]
        if not today:
            parts.append("<p>No jobs scheduled today.</p>")
        parts.append(_footer())
        return f"🌙 End of Day — {_today()}", "\n".join(parts)
    except Exception:
        return None


# ── Registry ─────────────────────────────────────────────────────────────────

JOB_REGISTRY: dict[str, dict] = {
    "morning_briefing": {
        "label":        "☀️ Morning Briefing",
        "description":  "Today\'s jobs, weather, overdue invoices, unanswered SMS, due tasks",
        "fn":           job_morning_briefing,
        "default_time": "07:00",
        "default_days": "weekdays",
    },
    "overdue_invoice_alert": {
        "label":        "⚠️ Overdue Invoice Alert",
        "description":  "Silent unless invoices are 31-60, 61-90, or 90+ days overdue",
        "fn":           job_overdue_invoice_alert,
        "default_time": "08:00",
        "default_days": "daily",
    },
    "due_analysis_tasks": {
        "label":        "🧠 Due Analysis Tasks",
        "description":  "Alerts when scheduled AI analysis tasks are due or overdue",
        "fn":           job_due_analysis_tasks,
        "default_time": "08:05",
        "default_days": "daily",
    },
    "sms_reply_monitor": {
        "label":        "💬 SMS Reply Monitor",
        "description":  "Alerts on unanswered customer messages (runs every 2 hours 8am–8pm)",
        "fn":           job_sms_reply_monitor,
        "default_time": "every_2h",
        "default_days": "daily",
    },
    "weather_watch": {
        "label":        "🌤️ Weekly Weather Watch",
        "description":  "Sunday evening 5-day forecast — flag rain days with outdoor jobs",
        "fn":           job_weather_watch,
        "default_time": "19:00",
        "default_days": "sunday",
    },
    "end_of_day_summary": {
        "label":        "🌙 End of Day Summary",
        "description":  "Jobs completed vs scheduled today, missing time entries",
        "fn":           job_end_of_day_summary,
        "default_time": "18:00",
        "default_days": "daily",
    },
}

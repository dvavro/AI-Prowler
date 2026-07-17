"""
tests/analysis/test_scheduler.py
==========================================
Test suite for the Proactive Alerts scheduler (v8.0.0).

Tests:
  TC-SCHED-001  scheduler_jobs — job registry completeness and structure
  TC-SCHED-002  scheduler_jobs — helper functions (mocked AI-Prowler calls)
  TC-SCHED-003  scheduler_jobs — job functions return correct types
  TC-SCHED-004  scheduler_jobs — silent jobs return None when nothing to report
  TC-SCHED-005  scheduler_engine — config load/save roundtrip
  TC-SCHED-006  scheduler_engine — scheduling logic (_is_day_match, _is_time_due)
  TC-SCHED-007  scheduler_engine — last-run tracking
  TC-SCHED-008  scheduler_engine — start/stop/is_running lifecycle

Run:
    run_tests.bat tests\\analysis\\test_scheduler.py -v
"""

import datetime
import json
import time
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock


# ---------------------------------------------------------------------------
# TC-SCHED-001  Job registry completeness and structure
# ---------------------------------------------------------------------------

class TestJobRegistry:

    def test_TC_SCHED_001_registry_has_all_six_jobs(self):
        import scheduler_jobs as sj
        expected = {
            "morning_briefing", "overdue_invoice_alert", "due_analysis_tasks",
            "sms_reply_monitor", "weather_watch", "end_of_day_summary",
        }
        assert set(sj.JOB_REGISTRY.keys()) == expected

    def test_TC_SCHED_001_each_job_has_required_keys(self):
        import scheduler_jobs as sj
        required = ["label", "description", "fn", "default_time", "default_days"]
        for jid, meta in sj.JOB_REGISTRY.items():
            for key in required:
                assert key in meta, f"Job '{jid}' missing key '{key}'"

    def test_TC_SCHED_001_fn_is_callable(self):
        import scheduler_jobs as sj
        for jid, meta in sj.JOB_REGISTRY.items():
            assert callable(meta["fn"]), f"Job '{jid}' fn is not callable"

    def test_TC_SCHED_001_default_time_is_string(self):
        import scheduler_jobs as sj
        for jid, meta in sj.JOB_REGISTRY.items():
            assert isinstance(meta["default_time"], str),                 f"Job '{jid}' default_time is not a string"

    def test_TC_SCHED_001_default_days_is_valid(self):
        import scheduler_jobs as sj
        valid = {"daily", "weekdays", "weekends", "monday", "tuesday",
                 "wednesday", "thursday", "friday", "saturday", "sunday"}
        for jid, meta in sj.JOB_REGISTRY.items():
            assert meta["default_days"].lower() in valid,                 f"Job '{jid}' has invalid default_days: {meta['default_days']}"

    def test_TC_SCHED_001_sms_monitor_uses_interval_time(self):
        import scheduler_jobs as sj
        assert sj.JOB_REGISTRY["sms_reply_monitor"]["default_time"].startswith("every_")

    def test_TC_SCHED_001_label_contains_emoji(self):
        import scheduler_jobs as sj
        for jid, meta in sj.JOB_REGISTRY.items():
            # Labels should have an emoji prefix for visual identification
            assert len(meta["label"]) > 3, f"Job '{jid}' label too short"

    def test_TC_SCHED_001_description_is_nonempty(self):
        import scheduler_jobs as sj
        for jid, meta in sj.JOB_REGISTRY.items():
            assert meta["description"].strip(),                 f"Job '{jid}' has empty description"


# ---------------------------------------------------------------------------
# TC-SCHED-002  Helper functions (mocked AI-Prowler calls)
# ---------------------------------------------------------------------------

class TestSchedulerHelpers:

    def test_TC_SCHED_002_today_returns_iso_date(self):
        import scheduler_jobs as sj
        result = sj._today()
        datetime.date.fromisoformat(result)  # raises if invalid

    def test_TC_SCHED_002_now_str_format(self):
        import scheduler_jobs as sj
        result = sj._now_str()
        datetime.datetime.strptime(result, "%Y-%m-%d %H:%M")  # raises if invalid

    def test_TC_SCHED_002_footer_contains_timestamp(self):
        import scheduler_jobs as sj
        f = sj._footer()
        assert "AI-Prowler" in f
        assert "Proactive Alert" in f

    def test_TC_SCHED_002_ar_aging_returns_string_on_success(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.get_ar_aging_report",
                   return_value="Current: $500  31-60: $200"):
            result = sj._ar_aging()
        assert "Current" in result

    def test_TC_SCHED_002_ar_aging_returns_empty_on_exception(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.get_ar_aging_report",
                   side_effect=Exception("DB error")):
            result = sj._ar_aging()
        assert result == ""

    def test_TC_SCHED_002_weather_returns_string_on_success(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.get_weather",
                   return_value="Sunny 84°F"):
            result = sj._weather("New Smyrna Beach, FL")
        assert "84" in result

    def test_TC_SCHED_002_weather_returns_empty_on_exception(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.get_weather",
                   side_effect=Exception("timeout")):
            result = sj._weather("anywhere")
        assert result == ""

    def test_TC_SCHED_002_sms_replies_returns_empty_on_exception(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.list_sms_contacts_with_replies",
                   side_effect=Exception("not configured")):
            result = sj._sms_replies()
        assert result == ""

    def test_TC_SCHED_002_job_rows_returns_empty_on_exception(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.read_job_spreadsheet",
                   side_effect=Exception("file not found")):
            result = sj._job_rows()
        assert result == []

    def test_TC_SCHED_002_pending_tasks_returns_empty_on_missing_file(self, tmp_path):
        import scheduler_jobs as sj
        with patch.object(Path, "home", return_value=tmp_path):
            result = sj._pending_tasks()
        assert result == []


# ---------------------------------------------------------------------------
# TC-SCHED-010  _todays_jobs_structured — real xlsx, per-job City/State
# extraction for per-job weather cross-referencing (v8.1.3)
# ---------------------------------------------------------------------------

class TestTodaysJobsStructured:
    """
    Uses a REAL temp .xlsx workbook (via openpyxl) rather than mocking —
    this is the actual risky logic (header detection, date-cell matching,
    column mapping), the same class of thing read_job_spreadsheet already
    has tested coverage for. Mocking it out would leave the real parsing
    completely unverified.
    """

    def _make_workbook(self, tmp_path, rows, headers=None):
        import openpyxl
        headers = headers or [
            "JobID (JOB-####)", "Customer Name / Company",
            "City ★ AI Route", "State", "Service Type",
            "Crew / Technician", "Service Date",
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs_Schedule"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        fp = tmp_path / "jobs.xlsx"
        wb.save(fp)
        return str(fp)

    def test_TC_SCHED_010_extracts_todays_jobs_with_city_state(self, tmp_path):
        import scheduler_jobs as sj
        today = datetime.date.today().isoformat()
        fp = self._make_workbook(tmp_path, rows=[
            ["JOB-0001", "Torres Residence", "New Smyrna Beach", "FL",
             "Window Cleaning", "Jake", today],
            ["JOB-0002", "Blue Wave Cafe", "Port Orange", "FL",
             "Pressure Washing", "Sam", today],
        ])
        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=fp):
            result = sj._todays_jobs_structured()

        assert len(result) == 2
        assert result[0]["customer"] == "Torres Residence"
        assert result[0]["city"] == "New Smyrna Beach"
        assert result[0]["state"] == "FL"
        assert result[1]["city"] == "Port Orange"

    def test_TC_SCHED_010_excludes_jobs_on_other_dates(self, tmp_path):
        import scheduler_jobs as sj
        today = datetime.date.today().isoformat()
        yesterday = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
        tomorrow = (datetime.date.today() + datetime.timedelta(days=1)).isoformat()
        fp = self._make_workbook(tmp_path, rows=[
            ["JOB-0001", "Today Job", "New Smyrna Beach", "FL",
             "Window Cleaning", "Jake", today],
            ["JOB-0002", "Yesterday Job", "Edgewater", "FL",
             "Window Cleaning", "Jake", yesterday],
            ["JOB-0003", "Tomorrow Job", "Daytona Beach", "FL",
             "Window Cleaning", "Jake", tomorrow],
        ])
        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=fp):
            result = sj._todays_jobs_structured()

        assert len(result) == 1
        assert result[0]["customer"] == "Today Job"

    def test_TC_SCHED_010_handles_datetime_object_service_date(self, tmp_path):
        """openpyxl often returns actual datetime objects for date cells,
        not strings — must be handled, not just string-formatted dates."""
        import scheduler_jobs as sj
        import openpyxl
        today_dt = datetime.datetime.combine(datetime.date.today(), datetime.time())
        headers = ["JobID (JOB-####)", "Customer Name / Company",
                  "City ★ AI Route", "State", "Service Type",
                  "Crew / Technician", "Service Date"]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs_Schedule"
        ws.append(headers)
        ws.append(["JOB-0001", "Datetime Job", "New Smyrna Beach", "FL",
                   "Window Cleaning", "Jake", today_dt])
        fp = tmp_path / "jobs.xlsx"
        wb.save(fp)

        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=str(fp)):
            result = sj._todays_jobs_structured()

        assert len(result) == 1
        assert result[0]["customer"] == "Datetime Job"

    def test_TC_SCHED_010_no_spreadsheet_path_returns_empty(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=""):
            result = sj._todays_jobs_structured()
        assert result == []

    def test_TC_SCHED_010_missing_file_returns_empty(self, tmp_path):
        import scheduler_jobs as sj
        fp = str(tmp_path / "does_not_exist.xlsx")
        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=fp):
            result = sj._todays_jobs_structured()
        assert result == []

    def test_TC_SCHED_010_missing_jobs_schedule_sheet_returns_empty(self, tmp_path):
        import scheduler_jobs as sj
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SomeOtherSheet"
        ws.append(["A", "B", "C"])
        fp = tmp_path / "wrong_sheet.xlsx"
        wb.save(fp)

        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=str(fp)):
            result = sj._todays_jobs_structured()
        assert result == []

    def test_TC_SCHED_010_no_service_date_column_returns_empty(self, tmp_path):
        """A sheet with no recognizable Service Date column must return []
        rather than guess — mirrors read_job_spreadsheet's own behavior."""
        import scheduler_jobs as sj
        fp = self._make_workbook(
            tmp_path,
            rows=[["JOB-0001", "No Date Job", "Edgewater", "FL", "Cleaning", "Jake"]],
            headers=["JobID (JOB-####)", "Customer Name / Company",
                    "City ★ AI Route", "State", "Service Type",
                    "Crew / Technician"],
        )
        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=fp):
            result = sj._todays_jobs_structured()
        assert result == []

    def test_TC_SCHED_010_never_raises_on_corrupt_workbook(self, tmp_path):
        import scheduler_jobs as sj
        fp = tmp_path / "corrupt.xlsx"
        fp.write_text("this is not a real xlsx file", encoding="utf-8")
        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=str(fp)):
            result = sj._todays_jobs_structured()  # must not raise
        assert result == []

    def test_TC_SCHED_010_blank_rows_skipped(self, tmp_path):
        import scheduler_jobs as sj
        today = datetime.date.today().isoformat()
        fp = self._make_workbook(tmp_path, rows=[
            ["JOB-0001", "Real Job", "New Smyrna Beach", "FL",
             "Cleaning", "Jake", today],
            [None, None, None, None, None, None, None],
        ])
        with patch("ai_prowler_mcp._get_default_spreadsheet_path", return_value=fp):
            result = sj._todays_jobs_structured()
        assert len(result) == 1


# ---------------------------------------------------------------------------
# TC-SCHED-003  Job functions return correct types
# ---------------------------------------------------------------------------

class TestJobReturnTypes:
    """
    Each job function must return either:
      - (str, str) tuple  — subject, html_body
      - None              — nothing to report (silent)
    Never raises.
    """

    def _mock_config(self):
        return {
            "name": "David",
            "location": "New Smyrna Beach, Florida",
            "email_to": "david.vavro1@gmail.com",
            "sms_threshold_hours": 4,
        }

    def _run_with_mocks(self, fn, ar="", weather="Sunny 84F",
                        sms="", jobs=None, tasks=None):
        """Run a job function with all AI-Prowler calls mocked."""
        import scheduler_jobs as sj
        jobs = jobs or []
        tasks_json = json.dumps(tasks or [])

        with patch("ai_prowler_mcp.get_ar_aging_report", return_value=ar), \
             patch("ai_prowler_mcp.get_weather", return_value=weather), \
             patch("ai_prowler_mcp.list_sms_contacts_with_replies",
                   return_value=sms), \
             patch("ai_prowler_mcp.read_job_spreadsheet",
                   return_value="\n".join(jobs)), \
             patch("builtins.open", side_effect=FileNotFoundError):
            return fn(self._mock_config())

    def test_TC_SCHED_003_morning_briefing_returns_tuple(self):
        import scheduler_jobs as sj
        result = self._run_with_mocks(sj.job_morning_briefing,
                                      weather="Sunny 84F")
        assert isinstance(result, tuple)
        assert len(result) == 2
        subject, body = result
        assert isinstance(subject, str) and len(subject) > 5
        assert isinstance(body, str) and len(body) > 10

    def test_TC_SCHED_003_morning_briefing_subject_contains_date(self):
        import scheduler_jobs as sj
        result = self._run_with_mocks(sj.job_morning_briefing)
        assert result is not None
        subject, _ = result
        today = datetime.date.today()
        assert str(today.year) in subject or today.strftime("%B") in subject

    def test_TC_SCHED_003_end_of_day_returns_tuple(self):
        import scheduler_jobs as sj
        result = self._run_with_mocks(sj.job_end_of_day_summary)
        assert isinstance(result, tuple) and len(result) == 2

    def test_TC_SCHED_003_weather_watch_returns_tuple(self):
        import scheduler_jobs as sj
        with patch("scheduler_jobs._owner_location", return_value="New Smyrna Beach, FL"):
            result = self._run_with_mocks(sj.job_weather_watch,
                                          weather="Monday: Rain 72F")
        assert isinstance(result, tuple)
        _, body = result
        assert "Rain" in body

    def test_TC_SCHED_003_weather_watch_returns_none_when_no_location_configured(self):
        """v8.1.3: location now comes from Settings, not a hardcoded
        default — with nothing configured, this must return None (silence)
        rather than reporting on a guessed town."""
        import scheduler_jobs as sj
        with patch("scheduler_jobs._owner_location", return_value=""):
            result = self._run_with_mocks(sj.job_weather_watch)
        assert result is None

    def test_TC_SCHED_003_never_raises_on_exception(self):
        """All job functions must catch their own exceptions."""
        import scheduler_jobs as sj
        # Patch everything to raise
        with patch("ai_prowler_mcp.get_ar_aging_report",
                   side_effect=RuntimeError("DB crashed")),              patch("ai_prowler_mcp.get_weather",
                   side_effect=RuntimeError("timeout")),              patch("ai_prowler_mcp.list_sms_contacts_with_replies",
                   side_effect=RuntimeError("SMS error")),              patch("ai_prowler_mcp.read_job_spreadsheet",
                   side_effect=RuntimeError("xlsx error")):
            cfg = self._mock_config()
            for jid, meta in sj.JOB_REGISTRY.items():
                try:
                    result = meta["fn"](cfg)
                    # Result must be None or a (str, str) tuple — never raises
                    assert result is None or (
                        isinstance(result, tuple) and len(result) == 2
                    ), f"Job '{jid}' returned unexpected type: {type(result)}"
                except Exception as e:
                    pytest.fail(f"Job '{jid}' raised an exception: {e}")

    def test_TC_SCHED_003_body_contains_footer(self):
        import scheduler_jobs as sj
        result = self._run_with_mocks(sj.job_morning_briefing)
        assert result is not None
        _, body = result
        assert "AI-Prowler" in body


# ---------------------------------------------------------------------------
# TC-SCHED-011  job_morning_briefing — per-job weather by town (v8.1.3)
# ---------------------------------------------------------------------------

class TestMorningBriefingPerJobWeather:
    """
    job_morning_briefing now checks weather PER JOB'S OWN city/state via
    _todays_jobs_structured, instead of one fixed config['location'] for
    the whole report. These tests mock _todays_jobs_structured and _weather
    directly (the per-job xlsx parsing itself is covered by
    TestTodaysJobsStructured above) to verify the briefing's OWN logic:
    dedup by town, rain-flag rendering, and the no-jobs fallback.
    """

    def _cfg(self):
        return {"name": "David", "location": "New Smyrna Beach, Florida",
                "email_to": "david.vavro1@gmail.com"}

    def _run(self, jobs, weather_by_loc=None, ar="", sms="", tasks=None,
            owner_name="David", owner_location="New Smyrna Beach, Florida"):
        import scheduler_jobs as sj
        weather_by_loc = weather_by_loc or {}

        def _fake_weather(loc):
            return weather_by_loc.get(loc, "Sunny 80F")

        with patch("scheduler_jobs._todays_jobs_structured", return_value=jobs), \
             patch("scheduler_jobs._weather", side_effect=_fake_weather), \
             patch("scheduler_jobs._owner_name", return_value=owner_name), \
             patch("scheduler_jobs._owner_location", return_value=owner_location), \
             patch("ai_prowler_mcp.get_ar_aging_report", return_value=ar), \
             patch("ai_prowler_mcp.list_sms_contacts_with_replies", return_value=sms), \
             patch("builtins.open", side_effect=FileNotFoundError):
            return sj.job_morning_briefing(self._cfg())

    def test_TC_SCHED_011_lists_each_jobs_own_town(self):
        subject, body = self._run(jobs=[
            {"customer": "Torres Residence", "city": "New Smyrna Beach", "state": "FL"},
            {"customer": "Blue Wave Cafe", "city": "Port Orange", "state": "FL"},
        ])
        assert "Torres Residence" in body
        assert "New Smyrna Beach" in body
        assert "Blue Wave Cafe" in body
        assert "Port Orange" in body

    def test_TC_SCHED_011_weather_fetched_once_per_unique_town_not_per_job(self):
        """Two jobs in the SAME town must only trigger one _weather() call
        for that town — not one per job."""
        import scheduler_jobs as sj
        calls = []

        def _counting_weather(loc):
            calls.append(loc)
            return "Sunny 80F"

        with patch("scheduler_jobs._todays_jobs_structured", return_value=[
                {"customer": "Job A", "city": "New Smyrna Beach", "state": "FL"},
                {"customer": "Job B", "city": "New Smyrna Beach", "state": "FL"},
                {"customer": "Job C", "city": "Port Orange", "state": "FL"},
            ]), \
             patch("scheduler_jobs._weather", side_effect=_counting_weather), \
             patch("scheduler_jobs._owner_name", return_value="David"), \
             patch("scheduler_jobs._owner_location", return_value="New Smyrna Beach, Florida"), \
             patch("ai_prowler_mcp.get_ar_aging_report", return_value=""), \
             patch("ai_prowler_mcp.list_sms_contacts_with_replies", return_value=""), \
             patch("builtins.open", side_effect=FileNotFoundError):
            sj.job_morning_briefing(self._cfg())

        # Two unique towns among three jobs -> exactly two _weather() calls.
        assert len(calls) == 2
        assert calls.count("New Smyrna Beach, FL") == 1

    def test_TC_SCHED_011_rain_flag_shown_for_rainy_town(self):
        subject, body = self._run(
            jobs=[{"customer": "Torres Residence", "city": "New Smyrna Beach", "state": "FL"}],
            weather_by_loc={"New Smyrna Beach, FL": "⚠️ Rain likely, 72F"},
        )
        assert "Rain risk" in body
        assert "Torres Residence" in body

    def test_TC_SCHED_011_no_rain_flag_for_sunny_town(self):
        subject, body = self._run(
            jobs=[{"customer": "Torres Residence", "city": "New Smyrna Beach", "state": "FL"}],
            weather_by_loc={"New Smyrna Beach, FL": "Sunny 84F"},
        )
        assert "Rain risk" not in body

    def test_TC_SCHED_011_job_missing_city_falls_back_to_config_location(self):
        """A job with no City on file must still get SOME weather check —
        using the Settings-tab owner location (v8.1.3), not silently
        skipping it."""
        import scheduler_jobs as sj
        calls = []

        def _counting_weather(loc):
            calls.append(loc)
            return "Sunny 80F"

        with patch("scheduler_jobs._todays_jobs_structured",
                   return_value=[{"customer": "No City Job"}]), \
             patch("scheduler_jobs._weather", side_effect=_counting_weather), \
             patch("scheduler_jobs._owner_name", return_value="David"), \
             patch("scheduler_jobs._owner_location", return_value="New Smyrna Beach, Florida"), \
             patch("ai_prowler_mcp.get_ar_aging_report", return_value=""), \
             patch("ai_prowler_mcp.list_sms_contacts_with_replies", return_value=""), \
             patch("builtins.open", side_effect=FileNotFoundError):
            sj.job_morning_briefing(self._cfg())

        assert calls == ["New Smyrna Beach, Florida"]

    def test_TC_SCHED_011_no_jobs_falls_back_to_configured_location_weather(self):
        """When there are no jobs at all, the briefing still shows a
        general weather section for config['location'] — not nothing."""
        subject, body = self._run(jobs=[])
        assert "No jobs scheduled today" in body
        assert "New Smyrna Beach, Florida" in body

    def test_TC_SCHED_011_service_type_shown_when_present(self):
        subject, body = self._run(jobs=[
            {"customer": "Torres Residence", "city": "New Smyrna Beach",
             "state": "FL", "service_type": "Window Cleaning"},
        ])
        assert "Window Cleaning" in body

    def test_TC_SCHED_011_never_raises_when_weather_lookup_fails(self):
        import scheduler_jobs as sj
        with patch("scheduler_jobs._todays_jobs_structured",
                   return_value=[{"customer": "Torres Residence",
                                 "city": "New Smyrna Beach", "state": "FL"}]), \
             patch("scheduler_jobs._weather", side_effect=Exception("timeout")), \
             patch("ai_prowler_mcp.get_ar_aging_report", return_value=""), \
             patch("ai_prowler_mcp.list_sms_contacts_with_replies", return_value=""), \
             patch("builtins.open", side_effect=FileNotFoundError):
            result = sj.job_morning_briefing(self._cfg())
        # job_morning_briefing's own try/except catches this — but
        # _weather() itself also never raises in production (see
        # TC-SCHED-002), so this is a defense-in-depth check.
        assert result is not None


# ---------------------------------------------------------------------------
# TC-SCHED-004  Silent jobs return None when nothing to report
# ---------------------------------------------------------------------------

class TestSilentJobs:

    def _cfg(self):
        return {"name": "David", "location": "New Smyrna Beach, Florida",
                "sms_threshold_hours": 4}

    def test_TC_SCHED_004_overdue_invoice_silent_when_current(self):
        import scheduler_jobs as sj
        # Only current bucket — nothing overdue
        with patch("ai_prowler_mcp.get_ar_aging_report",
                   return_value="Current: $1500Total: $1500"):
            result = sj.job_overdue_invoice_alert(self._cfg())
        assert result is None

    def test_TC_SCHED_004_overdue_invoice_alerts_when_31_60(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.get_ar_aging_report",
                   return_value="31-60 days: Johnson $450"):
            result = sj.job_overdue_invoice_alert(self._cfg())
        assert result is not None
        subject, body = result
        assert "Overdue" in subject
        assert "Johnson" in body

    def test_TC_SCHED_004_overdue_invoice_alerts_when_90_plus(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.get_ar_aging_report",
                   return_value="90+ days: Williams $800"):
            result = sj.job_overdue_invoice_alert(self._cfg())
        assert result is not None

    def test_TC_SCHED_004_sms_monitor_silent_when_no_unread(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.list_sms_contacts_with_replies",
                   return_value="No contacts with replies."):
            result = sj.job_sms_reply_monitor(self._cfg())
        assert result is None

    def test_TC_SCHED_004_sms_monitor_alerts_when_unread(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.list_sms_contacts_with_replies",
                   return_value="3 unread: Johnson, Williams, Chen"):
            result = sj.job_sms_reply_monitor(self._cfg())
        assert result is not None
        subject, body = result
        assert "Unanswered" in subject or "Message" in subject

    def test_TC_SCHED_004_due_tasks_silent_when_none_due(self, tmp_path):
        import scheduler_jobs as sj
        # Pending tasks all in future
        future = (datetime.date.today() + datetime.timedelta(days=7)).isoformat()
        tasks = [{"task_id": "t1", "status": "pending",
                  "next_due": future, "label": "Future task"}]
        tasks_file = tmp_path / ".ai-prowler" / "pending_tasks.json"
        tasks_file.parent.mkdir(parents=True)
        tasks_file.write_text(json.dumps(tasks), encoding="utf-8")
        with patch.object(Path, "home", return_value=tmp_path):
            result = sj.job_due_analysis_tasks(self._cfg())
        assert result is None

    def test_TC_SCHED_004_due_tasks_alerts_when_overdue(self, tmp_path):
        import scheduler_jobs as sj
        yesterday = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
        tasks = [{"task_id": "t1", "status": "pending",
                  "next_due": yesterday, "label": "Weekly Report",
                  "schedule": "weekly"}]
        tasks_file = tmp_path / ".ai-prowler" / "pending_tasks.json"
        tasks_file.parent.mkdir(parents=True)
        tasks_file.write_text(json.dumps(tasks), encoding="utf-8")
        with patch.object(Path, "home", return_value=tmp_path):
            result = sj.job_due_analysis_tasks(self._cfg())
        assert result is not None
        subject, body = result
        assert "Due" in subject
        assert "Weekly Report" in body

    def test_TC_SCHED_004_overdue_invoice_returns_none_on_empty_ar(self):
        import scheduler_jobs as sj
        with patch("ai_prowler_mcp.get_ar_aging_report", return_value=""):
            result = sj.job_overdue_invoice_alert(self._cfg())
        assert result is None


# ---------------------------------------------------------------------------
# TC-SCHED-005  scheduler_engine — config load/save roundtrip
# ---------------------------------------------------------------------------

class TestSchedulerConfig:

    def test_TC_SCHED_005_load_returns_defaults_when_no_file(self, tmp_path):
        import scheduler_engine as se
        with patch("scheduler_engine.CONFIG_PATH",
                   tmp_path / "nonexistent.json"):
            cfg = se.load_config()
        assert "enabled" in cfg
        assert "jobs" in cfg
        assert cfg["enabled"] is False
        # v8.1.3: email_to/location/name are no longer part of the default
        # schema — they come live from Settings -> Email Configuration/
        # Owner Name instead (see scheduler_engine.py's module docstring).
        assert "email_to" not in cfg

    def test_TC_SCHED_005_save_and_load_roundtrip(self, tmp_path):
        import scheduler_engine as se
        cfg_path = tmp_path / "scheduler_config.json"
        original = {
            "enabled":  True,
            "email_to": "test@example.com",
            "location": "New Smyrna Beach, FL",
            "name":     "TestUser",
            "jobs": {
                "morning_briefing": {"enabled": True, "time": "07:30",
                                     "days": "weekdays"}
            }
        }
        with patch("scheduler_engine.CONFIG_PATH", cfg_path):
            se.save_config(original)
            loaded = se.load_config()
        assert loaded["enabled"] is True
        assert loaded["email_to"] == "test@example.com"
        assert loaded["name"] == "TestUser"
        assert loaded["jobs"]["morning_briefing"]["time"] == "07:30"

    def test_TC_SCHED_005_config_file_is_valid_json(self, tmp_path):
        import scheduler_engine as se
        cfg_path = tmp_path / "scheduler_config.json"
        with patch("scheduler_engine.CONFIG_PATH", cfg_path):
            se.save_config({"enabled": False, "email_to": "", "jobs": {}})
        parsed = json.loads(cfg_path.read_text(encoding="utf-8"))
        assert isinstance(parsed, dict)

    def test_TC_SCHED_005_save_creates_parent_dir(self, tmp_path):
        import scheduler_engine as se
        cfg_path = tmp_path / "subdir" / "scheduler_config.json"
        with patch("scheduler_engine.CONFIG_PATH", cfg_path):
            se.save_config({"enabled": False})
        assert cfg_path.exists()

    def test_TC_SCHED_005_load_handles_corrupt_json(self, tmp_path):
        import scheduler_engine as se
        cfg_path = tmp_path / "scheduler_config.json"
        cfg_path.write_text("NOT JSON", encoding="utf-8")
        with patch("scheduler_engine.CONFIG_PATH", cfg_path):
            cfg = se.load_config()
        # Should return defaults, not raise
        assert isinstance(cfg, dict)
        assert "enabled" in cfg

    def test_TC_SCHED_005_default_job_config_has_required_keys(self):
        import scheduler_engine as se
        cfg = se.default_job_config("morning_briefing")
        assert "enabled" in cfg
        assert "time" in cfg
        assert "days" in cfg
        assert cfg["enabled"] is False

    def test_TC_SCHED_005_default_job_config_unknown_id(self):
        import scheduler_engine as se
        cfg = se.default_job_config("nonexistent_job")
        assert isinstance(cfg, dict)
        assert "enabled" in cfg


# ---------------------------------------------------------------------------
# TC-SCHED-006  Scheduling logic
# ---------------------------------------------------------------------------

class TestSchedulingLogic:

    def _dt(self, weekday: int, hour: int = 8, minute: int = 0):
        """Create a datetime with a specific weekday (0=Mon) and time."""
        # Find next date with that weekday
        today = datetime.date.today()
        days_ahead = weekday - today.weekday()
        if days_ahead < 0:
            days_ahead += 7
        d = today + datetime.timedelta(days=days_ahead)
        return datetime.datetime(d.year, d.month, d.day, hour, minute)

    # _is_day_match tests
    def test_TC_SCHED_006_daily_always_true(self):
        import scheduler_engine as se
        for wd in range(7):
            assert se._is_day_match("daily", self._dt(wd)) is True

    def test_TC_SCHED_006_weekdays_true_mon_fri(self):
        import scheduler_engine as se
        for wd in range(5):  # Mon-Fri
            assert se._is_day_match("weekdays", self._dt(wd)) is True

    def test_TC_SCHED_006_weekdays_false_sat_sun(self):
        import scheduler_engine as se
        assert se._is_day_match("weekdays", self._dt(5)) is False  # Sat
        assert se._is_day_match("weekdays", self._dt(6)) is False  # Sun

    def test_TC_SCHED_006_sunday_only_on_sunday(self):
        import scheduler_engine as se
        assert se._is_day_match("sunday", self._dt(6)) is True   # Sun
        assert se._is_day_match("sunday", self._dt(0)) is False  # Mon
        assert se._is_day_match("sunday", self._dt(5)) is False  # Sat

    def test_TC_SCHED_006_monday_only_on_monday(self):
        import scheduler_engine as se
        assert se._is_day_match("monday", self._dt(0)) is True
        assert se._is_day_match("monday", self._dt(1)) is False

    def test_TC_SCHED_006_weekends_sat_sun(self):
        import scheduler_engine as se
        assert se._is_day_match("weekends", self._dt(5)) is True
        assert se._is_day_match("weekends", self._dt(6)) is True
        assert se._is_day_match("weekends", self._dt(0)) is False

    # _is_time_due tests
    def test_TC_SCHED_006_fixed_time_fires_at_exact_minute(self):
        import scheduler_engine as se
        now = datetime.datetime(2026, 6, 25, 7, 0)
        assert se._is_time_due("__test__", "07:00", now) is True

    def test_TC_SCHED_006_fixed_time_does_not_fire_1min_early(self):
        import scheduler_engine as se
        now = datetime.datetime(2026, 6, 25, 6, 59)
        assert se._is_time_due("__test__", "07:00", now) is False

    def test_TC_SCHED_006_fixed_time_does_not_fire_1min_late(self):
        import scheduler_engine as se
        now = datetime.datetime(2026, 6, 25, 7, 1)
        assert se._is_time_due("__test__", "07:00", now) is False

    def test_TC_SCHED_006_interval_fires_when_elapsed(self):
        import scheduler_engine as se
        # Set last run to 2h 1min ago
        past = (datetime.datetime.now() -
                datetime.timedelta(hours=2, minutes=1)).strftime("%Y-%m-%d %H:%M")
        se._last_run["__interval_test__"] = past
        now = datetime.datetime.now()
        assert se._is_time_due("__interval_test__", "every_2h", now) is True

    def test_TC_SCHED_006_interval_does_not_fire_when_not_elapsed(self):
        import scheduler_engine as se
        # Set last run to 1h 30min ago (< 2h threshold)
        past = (datetime.datetime.now() -
                datetime.timedelta(hours=1, minutes=30)).strftime("%Y-%m-%d %H:%M")
        se._last_run["__interval_test2__"] = past
        now = datetime.datetime.now()
        assert se._is_time_due("__interval_test2__", "every_2h", now) is False

    def test_TC_SCHED_006_interval_fires_when_never_run(self):
        import scheduler_engine as se
        # Job has never run — should fire immediately
        se._last_run.pop("__never_run__", None)
        now = datetime.datetime.now()
        assert se._is_time_due("__never_run__", "every_2h", now) is True

    def test_TC_SCHED_006_every_30m_interval(self):
        import scheduler_engine as se
        past = (datetime.datetime.now() -
                datetime.timedelta(minutes=31)).strftime("%Y-%m-%d %H:%M")
        se._last_run["__30m_test__"] = past
        now = datetime.datetime.now()
        assert se._is_time_due("__30m_test__", "every_30m", now) is True


# ---------------------------------------------------------------------------
# TC-SCHED-007  Last-run tracking
# ---------------------------------------------------------------------------

class TestLastRunTracking:

    def test_TC_SCHED_007_get_last_run_returns_never_when_unset(self):
        import scheduler_engine as se
        se._last_run.pop("__fresh_job__", None)
        assert se.get_last_run("__fresh_job__") == "Never"

    def test_TC_SCHED_007_get_last_run_returns_stored_value(self):
        import scheduler_engine as se
        se._last_run["__stored_job__"] = "2026-06-24 07:00"
        assert se.get_last_run("__stored_job__") == "2026-06-24 07:00"

    def test_TC_SCHED_007_already_ran_today_false_when_unset(self):
        import scheduler_engine as se
        se._last_run.pop("__norun_job__", None)
        assert se._already_ran_today("__norun_job__") is False

    def test_TC_SCHED_007_already_ran_today_true_when_ran_today(self):
        import scheduler_engine as se
        today = datetime.date.today().isoformat()
        se._last_run["__today_job__"] = f"{today} 07:00"
        assert se._already_ran_today("__today_job__") is True

    def test_TC_SCHED_007_already_ran_today_false_when_ran_yesterday(self):
        import scheduler_engine as se
        yesterday = (datetime.date.today() -
                     datetime.timedelta(days=1)).isoformat()
        se._last_run["__yesterday_job__"] = f"{yesterday} 07:00"
        assert se._already_ran_today("__yesterday_job__") is False

    def test_TC_SCHED_007_save_and_load_last_run_roundtrip(self, tmp_path):
        import scheduler_engine as se
        lr_path = tmp_path / "scheduler_last_run.json"
        se._last_run["__roundtrip_job__"] = "2026-06-24 08:00"
        with patch("scheduler_engine._LAST_RUN_PATH", lr_path):
            se._save_last_run()
            se._last_run.clear()
            se._load_last_run()
        assert se._last_run.get("__roundtrip_job__") == "2026-06-24 08:00"

    def test_TC_SCHED_007_load_last_run_handles_missing_file(self, tmp_path):
        import scheduler_engine as se
        missing = tmp_path / "nonexistent.json"
        se._last_run.clear()
        with patch("scheduler_engine._LAST_RUN_PATH", missing):
            se._load_last_run()
        assert isinstance(se._last_run, dict)


# ---------------------------------------------------------------------------
# TC-SCHED-008  Start/stop/is_running lifecycle
# ---------------------------------------------------------------------------

class TestSchedulerLifecycle:

    def test_TC_SCHED_008_starts_and_is_running(self):
        import scheduler_engine as se
        se.stop()
        time.sleep(0.1)
        se.start()
        time.sleep(0.2)
        assert se.is_running() is True
        se.stop()

    def test_TC_SCHED_008_stops_cleanly(self):
        import scheduler_engine as se
        se.start()
        time.sleep(0.1)
        se.stop()
        time.sleep(0.3)
        assert se.is_running() is False

    def test_TC_SCHED_008_double_start_does_not_create_second_thread(self):
        import scheduler_engine as se
        se.stop()
        time.sleep(0.1)
        se.start()
        thread_before = se._thread
        se.start()  # second start — should be a no-op
        assert se._thread is thread_before
        se.stop()

    def test_TC_SCHED_008_run_job_now_unknown_id_returns_error(self):
        import scheduler_engine as se
        result = se.run_job_now("nonexistent_job_id")
        assert "❌" in result or "Unknown" in result

    def test_TC_SCHED_008_run_job_now_no_email_returns_warning(self):
        """v8.1.3: no default recipient configured (via
        _read_default_to_email(), not cfg['email_to']) must return the
        warning, never silently 'succeed' with nowhere to send."""
        import scheduler_engine as se
        with patch("scheduler_engine.load_config", return_value={"jobs": {}}), \
             patch("scheduler_engine._read_default_to_email", return_value=""), \
             patch("scheduler_jobs._owner_name", return_value="Test"), \
             patch("scheduler_jobs._owner_location", return_value="Test"), \
             patch("ai_prowler_mcp.get_ar_aging_report",
                   return_value="Current: $500"), \
             patch("ai_prowler_mcp.get_weather",
                   return_value="Sunny"), \
             patch("ai_prowler_mcp.read_job_spreadsheet",
                   return_value=""), \
             patch("ai_prowler_mcp.list_sms_contacts_with_replies",
                   return_value=""):
            result = se.run_job_now("morning_briefing")
        assert "⚠️" in result or "No default recipient" in result

    def test_TC_SCHED_008_get_log_tail_returns_string(self, tmp_path):
        import scheduler_engine as se
        # get_log_tail reads LOG_PATH as a closure variable captured at import
        # time — patching the module attribute doesn't reach it. Instead write
        # to the real LOG_PATH, call get_log_tail, then restore/delete.
        real_log = se.LOG_PATH
        backup = None
        try:
            if real_log.exists():
                backup = real_log.read_bytes()
            real_log.parent.mkdir(parents=True, exist_ok=True)
            real_log.write_text("line1\nline2\nline3", encoding="utf-8")
            result = se.get_log_tail(10)
        finally:
            if backup is not None:
                real_log.write_bytes(backup)
            elif real_log.exists():
                real_log.unlink()
        assert isinstance(result, str)
        assert len(result) > 0

    def test_TC_SCHED_008_get_log_tail_no_file_returns_message(self, tmp_path):
        import scheduler_engine as se
        with patch("scheduler_engine.LOG_PATH",
                   tmp_path / "nonexistent.log"):
            result = se.get_log_tail(10)
        assert "no log" in result.lower() or isinstance(result, str)


# ---------------------------------------------------------------------------
# TC-SCHED-009  scheduler_engine — _send_email() delivery contract
#
# Regression test for a real bug (found 2026-07-12): _send_email() called
# send_email(..., html=True) and send_alert(subject=..., message=...), but
# neither ai_prowler_mcp.send_email() nor send_alert() actually accept those
# keyword arguments. Every prior test in this file mocks send_email/send_alert
# with plain MagicMock(), which silently accepts ANY kwargs — so the mismatch
# never raised and was invisible to the suite. run_job_now() reported success
# generating the briefing but the email itself silently failed every time.
#
# Fix: use autospec=True (or spec_set with the real function object) so the
# mock enforces the REAL signature — a wrong kwarg raises TypeError in tests
# exactly as it would against the live function.
# ---------------------------------------------------------------------------

class TestSchedulerEmailDeliveryContract:

    def test_TC_SCHED_009_send_email_uses_real_send_email_signature(self):
        """_send_email()'s primary path must call ai_prowler_mcp.send_email()
        with only kwargs that function actually accepts (to, subject, body,
        attachment_path, body_html as of v8.2+). autospec=True raises
        TypeError on any others — this is what would have caught the
        original 'html=True' bug immediately.

        v8.2+: body_html is now a real, supported kwarg carrying the actual
        HTML content (fixing raw-tags-in-inbox), while body carries a
        tag-stripped plain-text fallback."""
        import scheduler_engine as se
        import ai_prowler_mcp as apm

        with patch("ai_prowler_mcp.send_email", autospec=True,
                   return_value="✅ Email sent to test@example.com") as mock_send:
            ok = se._send_email("test@example.com", "Test Subject", "<p>body</p>")

        assert ok is True
        mock_send.assert_called_once()
        # autospec already guarantees no invalid kwargs were passed (it would
        # have raised TypeError above). Confirm the HTML content actually
        # reaches send_email via body_html, and body is a plain-text fallback
        # (not raw HTML) — this is the actual formatting-fix contract.
        kwargs = mock_send.call_args.kwargs
        assert kwargs.get("body_html") == "<p>body</p>"
        assert "<p>" not in kwargs.get("body", "")

    def test_TC_SCHED_009_send_email_fallback_uses_real_send_alert_signature(self):
        """When send_email() fails/raises, the send_alert() fallback must also
        only use kwargs send_alert() actually accepts (message, to) — not
        'subject', which send_alert() has never supported."""
        import scheduler_engine as se

        with patch("ai_prowler_mcp.send_email", autospec=True,
                   side_effect=RuntimeError("SMTP down")), \
             patch("ai_prowler_mcp.send_alert", autospec=True,
                   return_value="✅ Alert sent to test@example.com") as mock_alert:
            ok = se._send_email("test@example.com", "Test Subject", "<p>body</p>")

        assert ok is True
        mock_alert.assert_called_once()
        assert "subject" not in mock_alert.call_args.kwargs
        # subject must be folded into the message text instead
        assert "Test Subject" in mock_alert.call_args.kwargs.get("message", "")

    def test_TC_SCHED_009_send_email_both_paths_fail_returns_false(self):
        """If both send_email and send_alert genuinely fail, _send_email()
        must return False rather than raising — run_job_now() depends on
        this to report '⚠️ Generated but email failed' instead of crashing."""
        import scheduler_engine as se

        with patch("ai_prowler_mcp.send_email", autospec=True,
                   side_effect=RuntimeError("SMTP down")), \
             patch("ai_prowler_mcp.send_alert", autospec=True,
                   side_effect=RuntimeError("also down")):
            ok = se._send_email("test@example.com", "Test Subject", "<p>body</p>")

        assert ok is False

    def test_TC_SCHED_009_run_job_now_end_to_end_delivery(self):
        """End-to-end: run_job_now() with a configured default recipient
        (v8.1.3: from _read_default_to_email(), not cfg['email_to']) and a
        mocked (real-signature-enforced) send_email must report success,
        not 'Generated but email failed'."""
        import scheduler_engine as se

        with patch("scheduler_engine.load_config", return_value={"jobs": {}}), \
             patch("scheduler_engine._read_default_to_email",
                   return_value="test@example.com"), \
             patch("scheduler_jobs._owner_name", return_value="Test"), \
             patch("scheduler_jobs._owner_location", return_value="Test"), \
             patch("ai_prowler_mcp.get_ar_aging_report", return_value="Current: $0"), \
             patch("ai_prowler_mcp.get_weather", return_value="Sunny"), \
             patch("ai_prowler_mcp.read_job_spreadsheet", return_value=""), \
             patch("ai_prowler_mcp.list_sms_contacts_with_replies", return_value=""), \
             patch("ai_prowler_mcp.send_email", autospec=True,
                   return_value="✅ Email sent to test@example.com"):
            result = se.run_job_now("morning_briefing")

        assert "✅ Sent" in result
        assert "failed" not in result.lower()

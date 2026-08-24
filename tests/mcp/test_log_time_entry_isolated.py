"""
test_log_time_entry_isolated.py
================================
Functional tests for _log_time_entry_impl() (clock in/out, GPS logging,
header canonicalization) that exercise the REAL production function — not a
reimplemented copy — while staying fully isolated from:

  1. The real job tracker spreadsheets (production or dev-folder copy).
     Every test builds its own scratch .xlsx inside pytest's tmp_path,
     which pytest deletes automatically after the test. Nothing here ever
     opens a file under Documents/AI-Prowler or the dev-folder install.

  2. A live AI-Prowler server process, if one happens to be running.
     Importing ai_prowler_mcp.py normally opens
     ~/.ai-prowler/logs/mcp_server.log in truncate ("w") mode at import
     time — if a live server already has that file open, the two
     processes fight over it and can corrupt the log (this actually
     happened during manual testing on 2026-08-13). To avoid that, each
     test runs the real function in a SEPARATE SUBPROCESS with USERPROFILE
     (Windows' equivalent of $HOME) redirected to a scratch directory
     inside tmp_path — so Path.home() inside that subprocess resolves to
     the scratch dir, and mcp_server.log gets created there instead of
     under the real ~/.ai-prowler. The live server's log is never touched.

Safe to run at any time, including while AI-Prowler is running live.

Run with:
    pytest tests/mcp/test_log_time_entry_isolated.py -v
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
import textwrap
from pathlib import Path

import openpyxl
import pytest

_SRC = os.environ.get("AI_PROWLER_SRC")
SRC_ROOT = Path(_SRC).resolve() if _SRC else Path(__file__).resolve().parent.parent.parent
MCP_FILE = SRC_ROOT / "ai_prowler_mcp.py"

# The REAL decorated header format found on the production spreadsheet —
# testing against clean "Clock In" / "Clock Out" headers would not actually
# exercise the canonicalization fix, so we replicate the messy real one.
TIMELOG_HEADERS = [
    "EntryID", "JobID", "Customer Name / Company", "Date",
    "Clock In\n(HH:MM:SS)", "Clock Out\n(HH:MM:SS)",
    "Elapsed\n(min)\n=(Out-In)*1440", "Crew /\nTechnician", "Notes",
    "Clock In GPS", "Clock Out GPS",
]

JOBS_HEADERS = [
    "JobID (JOB-####)", "Customer Name / Company", "Crew / Technician",
]


def _build_scratch_tracker(path: Path, jobs=None):
    """Build a minimal but realistic job tracker at `path`. `jobs` is a
    list of (job_id, customer, crew) tuples; defaults to one test job."""
    jobs = jobs or [("JOB-TEST-01", "Test Customer LLC", "Test Crew")]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_log = wb.create_sheet("TimeLog")
    ws_log.append(["⏱️  TIME LOG — Job Clock In / Clock Out"])
    ws_log.append(TIMELOG_HEADERS)

    ws_jobs = wb.create_sheet("Jobs_Schedule")
    ws_jobs.append(["JOBS SCHEDULE"])
    ws_jobs.append(JOBS_HEADERS)
    for jid, cust, crew in jobs:
        ws_jobs.append([jid, cust, crew])

    wb.save(path)


def _run_clock_action(tmp_path: Path, xlsx_path: Path, job_identifier: str,
                       action: str, gps_coords: str = "") -> str:
    """
    Run the REAL _log_time_entry_impl() in an isolated subprocess.
    Returns the function's string result. Raises AssertionError with full
    stderr on any subprocess-level failure (import error, exception, etc.)
    so failures are legible in pytest output rather than a bare timeout.
    """
    scratch_home = tmp_path / "scratch_home"
    scratch_home.mkdir(exist_ok=True)

    script = textwrap.dedent(f"""
        import sys, json
        sys.path.insert(0, {str(SRC_ROOT)!r})
        import ai_prowler_mcp as m
        result = m._log_time_entry_impl(
            {job_identifier!r}, {action!r}, {str(xlsx_path)!r}, None,
            gps_coords={gps_coords!r},
        )
        print("RESULT_JSON_START" + json.dumps(result) + "RESULT_JSON_END")
    """)

    env = os.environ.copy()
    env["USERPROFILE"] = str(scratch_home)
    env["HOME"] = str(scratch_home)

    proc = subprocess.run(
        [sys.executable, "-c", script],
        cwd=str(SRC_ROOT), env=env,
        capture_output=True, text=True, timeout=90,
    )

    if "RESULT_JSON_START" not in proc.stdout:
        raise AssertionError(
            f"Subprocess did not return a result.\n"
            f"--- stdout ---\n{proc.stdout}\n"
            f"--- stderr ---\n{proc.stderr}\n"
            f"--- returncode --- {proc.returncode}"
        )

    payload = proc.stdout.split("RESULT_JSON_START", 1)[1].split("RESULT_JSON_END", 1)[0]
    return json.loads(payload)


def _read_row(xlsx_path: Path, entry_id: str) -> dict:
    """Read back a TimeLog row by EntryID as {header: value}."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["TimeLog"]
    hdrs = [c.value for c in ws[2]]
    for row in ws.iter_rows(min_row=3):
        if row[0].value == entry_id:
            return dict(zip(hdrs, [c.value for c in row]))
    raise AssertionError(f"EntryID {entry_id!r} not found in TimeLog.")


# ══════════════════════════════════════════════════════════════════════════
# TESTS
# ══════════════════════════════════════════════════════════════════════════

@pytest.fixture
def tracker(tmp_path):
    """Fresh scratch job tracker per test — never the real spreadsheet."""
    path = tmp_path / "scratch_tracker.xlsx"
    _build_scratch_tracker(path)
    return path


def _read_hyperlink(xlsx_path: Path, entry_id: str, col_name: str) -> str | None:
    """Read back the hyperlink target (not the display value) for a
    given TimeLog cell, or None if the cell has no hyperlink."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["TimeLog"]
    hdrs = [c.value for c in ws[2]]
    col_idx = hdrs.index(col_name) + 1
    for row in ws.iter_rows(min_row=3):
        if row[0].value == entry_id:
            cell = row[col_idx - 1]
            return cell.hyperlink.target if cell.hyperlink else None
    raise AssertionError(f"EntryID {entry_id!r} not found in TimeLog.")


def test_source_exists():
    assert MCP_FILE.exists(), f"ai_prowler_mcp.py not found at {MCP_FILE}"


def test_clock_in_writes_expected_fields(tmp_path, tracker):
    result = _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start",
                                gps_coords="29.0219,-80.9270")
    assert "Clocked IN" in result

    row = _read_row(tracker, "TE-0001")
    assert row["JobID"] == "JOB-TEST-01"
    assert row["Customer Name / Company"] == "Test Customer LLC"
    # The whole point of the fix: real decorated headers must resolve.
    assert row["Clock In\n(HH:MM:SS)"] not in (None, "")
    assert row["Clock In GPS"] == "29.0219,-80.9270"
    assert row["Clock Out GPS"] in (None, "")  # not set yet


def test_clock_out_writes_elapsed_and_gps(tmp_path, tracker):
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start",
                       gps_coords="29.0219,-80.9270")
    result = _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "stop",
                                gps_coords="29.0221,-80.9268")
    assert "❌" not in result, f"Unexpected error: {result}"

    row = _read_row(tracker, "TE-0001")
    assert row["Clock Out\n(HH:MM:SS)"] not in (None, "")
    assert row["Elapsed\n(min)\n=(Out-In)*1440"] is not None
    assert row["Clock In GPS"] == "29.0219,-80.9270"
    assert row["Clock Out GPS"] == "29.0221,-80.9268"


def test_clock_out_without_gps_leaves_field_blank(tmp_path, tracker):
    """GPS must stay optional — denied/unavailable location on the phone
    should never block or break a clock action."""
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start", gps_coords="")
    result = _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "stop", gps_coords="")
    assert "❌" not in result, f"Unexpected error: {result}"

    row = _read_row(tracker, "TE-0001")
    assert row["Clock Out\n(HH:MM:SS)"] not in (None, "")
    assert row["Clock In GPS"] in (None, "")
    assert row["Clock Out GPS"] in (None, "")


def test_duplicate_clock_in_blocked(tmp_path, tracker):
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start")
    result = _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start")
    assert "already open" in result


def test_clock_out_without_open_entry_errors_cleanly(tmp_path, tracker):
    result = _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "stop")
    assert "❌" in result
    assert "No open clock-in" in result
    # THE ORIGINAL BUG this whole fix was for — must never reappear:
    assert "Could not parse Clock In time: None" not in result


def test_gps_columns_never_collide_with_base_columns(tmp_path, tracker):
    """Regression guard for the canonicalization fix's GPS exclusion —
    if this ever regresses, GPS values would silently land in the wrong
    column or overwrite Clock In / Clock Out."""
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start",
                       gps_coords="1.111111,2.222222")
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "stop",
                       gps_coords="3.333333,4.444444")
    row = _read_row(tracker, "TE-0001")

    # Clock In/Out timestamps must be real timestamp strings, not GPS coords
    assert "," not in str(row["Clock In\n(HH:MM:SS)"])
    assert "," not in str(row["Clock Out\n(HH:MM:SS)"])
    # GPS columns must hold GPS coords, not timestamps
    assert row["Clock In GPS"] == "1.111111,2.222222"
    assert row["Clock Out GPS"] == "3.333333,4.444444"


def test_unrelated_headers_untouched_by_canonicalization(tmp_path, tracker):
    """Crew / Technician and Notes must never be folded into a Clock
    In/Out/Elapsed canonical name — guards against an over-broad prefix
    match in the canonicalization logic."""
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start")
    row = _read_row(tracker, "TE-0001")
    assert row["Crew /\nTechnician"] == "Test Crew"


def test_nonexistent_job_returns_clear_error(tmp_path, tracker):
    result = _run_clock_action(tmp_path, tracker, "JOB-DOES-NOT-EXIST", "start")
    assert "❌" in result
    assert "No job found" in result


def test_scratch_tracker_never_touches_real_files(tmp_path, tracker):
    """Sanity check on the test design itself: confirm the scratch tracker
    is genuinely under tmp_path, nowhere near the real spreadsheets."""
    assert "tmp" in str(tracker).lower() or str(tmp_path) in str(tracker)
    assert "Documents\\AI-Prowler" not in str(tracker)
    assert "AI-Prowler-V900_to_V910_work" not in str(tracker)


def test_clock_in_gps_gets_maps_hyperlink(tmp_path, tracker):
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start",
                       gps_coords="29.049377,-80.994279")
    link = _read_hyperlink(tracker, "TE-0001", "Clock In GPS")
    assert link == "https://www.google.com/maps?q=29.049377,-80.994279"
    # Display text must be untouched — still the raw, copyable coordinates
    row = _read_row(tracker, "TE-0001")
    assert row["Clock In GPS"] == "29.049377,-80.994279"


def test_clock_out_gps_gets_maps_hyperlink(tmp_path, tracker):
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start")
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "stop",
                       gps_coords="29.049385,-80.994275")
    link = _read_hyperlink(tracker, "TE-0001", "Clock Out GPS")
    assert link == "https://www.google.com/maps?q=29.049385,-80.994275"


def test_no_gps_means_no_hyperlink(tmp_path, tracker):
    """Blank GPS must not produce a broken/empty hyperlink."""
    _run_clock_action(tmp_path, tracker, "JOB-TEST-01", "start", gps_coords="")
    link = _read_hyperlink(tracker, "TE-0001", "Clock In GPS")
    assert link is None


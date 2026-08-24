"""
tests/mcp/test_create_invoice_isolated.py
=============================================
Functional tests for _create_invoice_impl() (the new create_invoice tool,
2026-08-24) that exercise the REAL production function — not a
reimplemented copy — while staying fully isolated, matching the same
subprocess-isolation pattern as test_log_time_entry_isolated.py:

  1. Every test builds its own scratch .xlsx inside pytest's tmp_path.
     Nothing here ever opens a file under Documents/AI-Prowler or the
     dev-folder install's real spreadsheet.

  2. Each test runs the real function in a SEPARATE SUBPROCESS with
     USERPROFILE (Windows' $HOME) redirected into tmp_path, so importing
     ai_prowler_mcp.py — which opens ~/.ai-prowler/logs/mcp_server.log in
     truncate mode at import time — can never fight a live server process
     for that file.

Scratch workbooks use the REAL decorated header strings (multi-line,
formula-documentation suffixes like "Taxable\\nAmt ($)\\n=K-L") pulled
directly off the actual production spreadsheet, not clean single-line
headers — otherwise these tests wouldn't actually exercise
_join_header_lines()'s canonicalization at all.

Why create_invoice independently computes dollar amounts rather than
reading Jobs_Schedule's own formula cells (Actual Amount/Tax/Invoice
Total): openpyxl never evaluates formulas, so those cells' cached values
are only as fresh as the last time a real spreadsheet app opened and
saved the file — which never happens in a mobile/voice workflow. Several
tests below specifically build workbooks where those formula cells are
stale or entirely blank, to prove the tool never depends on them.

Safe to run at any time, including while AI-Prowler is running live.

Run with:
    pytest tests/mcp/test_create_invoice_isolated.py -v
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
import textwrap
from pathlib import Path

import openpyxl
import pytest

_SRC = os.environ.get("AI_PROWLER_SRC")
SRC_ROOT = Path(_SRC).resolve() if _SRC else Path(__file__).resolve().parent.parent.parent
MCP_FILE = SRC_ROOT / "ai_prowler_mcp.py"

# Real decorated headers, copied verbatim off the production spreadsheet —
# includes the formula-documentation third line on the three columns that
# have one (Actual Amount, Taxable Amt is on Invoices not Jobs, etc.),
# which _join_header_lines() must strip.
JOBS_HEADERS = [
    "JobID\n(JOB-####)", "CustomerID\n(Customers!A)", "Customer Name\n/ Company",
    "Customer\nType", "Street Address\n★ AI Route", "City\n★ AI Route", "State",
    "ZIP\n★ AI Route", "Latitude\n(AI Geocode)", "Longitude\n(AI Geocode)",
    "Service\nDate", "End Date\n(blank = single-day job)", "Day of\nWeek",
    "Start\nTime", "End\nTime", "Service\nType", "Service\nDetails / Notes",
    "Crew /\nTechnician", "Est.\nDuration", "Est.\nDuration\nUnit",
    "Actual\nDuration", "Actual\nDuration\nUnit", "Route\nStop #\n★ AI Route",
    "Route\nMap URL\n★ AI Prowler", "Weather\nCheck\n★ AI Prowler", "Job\nStatus",
    "Quote\nAmount ($)", "Discount\nApplied ($)", "Actual\nAmount ($)\n=Quote-Discount",
    "Tax\n(7%)", "Invoice\nTotal ($)", "Recurrence", "InvoiceID\n(INV-####)",
    "Invoice\nSent Date", "Payment\nStatus",
]

INVOICES_HEADERS = [
    "InvoiceID\n(INV-####)", "JobID\n(JOB-####)", "CustomerID",
    "Customer Name\n/ Company", "Customer\nType", "Invoice\nDate",
    "Due\nDate\n(Net 30)", "Service\nDate", "Service\nType", "Description",
    "Subtotal\n($)", "Discount\n($)", "Taxable\nAmt ($)\n=K-L",
    "Tax 7%\n($)\n=M*0.07", "TOTAL\nDUE ($)\n=M+N", "Amount\nPaid ($)",
    "Balance\nDue ($)\n=O-P", "Payment\nStatus", "Payment\nDate",
    "Payment\nMethod", "Days\nOverdue\n(AI-AR)",
]


# Canonical (post _join_header_lines) name -> 0-based index into JOBS_HEADERS.
_JOBS_CANON = [
    "JobID (JOB-####)", "CustomerID (Customers!A)", "Customer Name / Company",
    "Customer Type", "Street Address ★ AI Route", "City ★ AI Route", "State",
    "ZIP ★ AI Route", "Latitude (AI Geocode)", "Longitude (AI Geocode)",
    "Service Date", "End Date (blank = single-day job)", "Day of Week",
    "Start Time", "End Time", "Service Type", "Service Details / Notes",
    "Crew / Technician", "Est. Duration", "Est. Duration Unit",
    "Actual Duration", "Actual Duration Unit", "Route Stop # ★ AI Route",
    "Route Map URL ★ AI Prowler", "Weather Check ★ AI Prowler", "Job Status",
    "Quote Amount ($)", "Discount Applied ($)", "Actual Amount ($)",
    "Tax (7%)", "Invoice Total ($)", "Recurrence", "InvoiceID (INV-####)",
    "Invoice Sent Date", "Payment Status",
]
_INV_CANON = [
    "InvoiceID (INV-####)", "JobID (JOB-####)", "CustomerID",
    "Customer Name / Company", "Customer Type", "Invoice Date",
    "Due Date (Net 30)", "Service Date", "Service Type", "Description",
    "Subtotal ($)", "Discount ($)", "Taxable Amt ($)", "Tax 7% ($)",
    "TOTAL DUE ($)", "Amount Paid ($)", "Balance Due ($)", "Payment Status",
    "Payment Date", "Payment Method", "Days Overdue (AI-AR)",
]


def _build_scratch_tracker(path: Path, jobs: list[dict], include_invoices_sheet=True):
    """
    Build a minimal but realistic job tracker at `path`, using the real
    decorated headers. `jobs` is a list of dicts keyed by CANONICAL column
    name (from _JOBS_CANON) — missing keys are left blank, matching how a
    real partially-filled job row looks.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_jobs = wb.create_sheet("Jobs_Schedule")
    ws_jobs.append(["📅  JOBS & SCHEDULE"])
    ws_jobs.append(JOBS_HEADERS)
    for job in jobs:
        row = [job.get(canon, "") for canon in _JOBS_CANON]
        ws_jobs.append(row)

    if include_invoices_sheet:
        ws_inv = wb.create_sheet("Invoices")
        ws_inv.append(["🧾  INVOICES"])
        ws_inv.append(INVOICES_HEADERS)

    wb.save(path)


def _run_create_invoice(
    tmp_path: Path, xlsx_path: Path, job_identifier: str,
    quote_amount=None, discount=None, description="", service_type="",
    tax_rate=0.07, due_days=30,
) -> str:
    """Run the REAL _create_invoice_impl() in an isolated subprocess (ctx=None,
    i.e. personal mode — server-mode crew scoping is covered structurally,
    see TestCrewScopingWired below, matching this codebase's existing
    convention of not fully mocking server-mode ctx in isolated tests)."""
    scratch_home = tmp_path / "scratch_home"
    scratch_home.mkdir(exist_ok=True)

    script = textwrap.dedent(f"""
        import sys, json
        sys.path.insert(0, {str(SRC_ROOT)!r})
        import ai_prowler_mcp as m
        result = m._create_invoice_impl(
            {job_identifier!r}, {quote_amount!r}, {discount!r},
            {description!r}, {service_type!r}, {tax_rate!r}, {due_days!r},
            {str(xlsx_path)!r}, False, None,
        )
        print("RESULT_JSON_START" + json.dumps(result) + "RESULT_JSON_END")
    """)

    env = os.environ.copy()
    env["USERPROFILE"] = str(scratch_home)
    env["HOME"] = str(scratch_home)

    proc = subprocess.run(
        [sys.executable, "-c", script],
        cwd=str(SRC_ROOT), env=env,
        capture_output=True, text=True, timeout=90,
    )

    if "RESULT_JSON_START" not in proc.stdout:
        raise AssertionError(
            f"Subprocess did not return a result.\n"
            f"--- stdout ---\n{proc.stdout}\n"
            f"--- stderr ---\n{proc.stderr}\n"
            f"--- returncode --- {proc.returncode}"
        )

    payload = proc.stdout.split("RESULT_JSON_START", 1)[1].split("RESULT_JSON_END", 1)[0]
    return json.loads(payload)


def _read_job_row(xlsx_path: Path, job_id: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Jobs_Schedule"]
    for row in ws.iter_rows(min_row=3):
        if row[0].value == job_id:
            return dict(zip(_JOBS_CANON, [c.value for c in row]))
    raise AssertionError(f"JobID {job_id!r} not found in Jobs_Schedule.")


def _read_invoice_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Invoices"]
    out = []
    for row in ws.iter_rows(min_row=3):
        if row[0].value is None:
            continue
        out.append(dict(zip(_INV_CANON, [c.value for c in row])))
    return out


def _read_invoice_by_id(xlsx_path: Path, inv_id: str) -> dict:
    for r in _read_invoice_rows(xlsx_path):
        if r["InvoiceID (INV-####)"] == inv_id:
            return r
    raise AssertionError(f"InvoiceID {inv_id!r} not found in Invoices.")


# ══════════════════════════════════════════════════════════════════════════
# FIXTURES
# ══════════════════════════════════════════════════════════════════════════

@pytest.fixture
def priced_job_tracker(tmp_path):
    """One job that already has a Quote Amount/Discount on file (the
    common case: a technician priced the job earlier and is now invoicing
    without needing to change anything)."""
    path = tmp_path / "scratch_tracker.xlsx"
    _build_scratch_tracker(path, [{
        "JobID (JOB-####)": "JOB-0001",
        "CustomerID (Customers!A)": "CUST-0001",
        "Customer Name / Company": "Torres Residence",
        "Customer Type": "Residential",
        "Service Date": "2026-08-24",
        "Service Type": "Window",
        "Service Details / Notes": "Full exterior window cleaning",
        "Crew / Technician": "Jake R.",
        "Quote Amount ($)": 200,
        "Discount Applied ($)": 20,
        # Deliberately STALE/wrong formula-cell values — real production
        # cells like these only update when Excel itself opens+saves the
        # file. create_invoice must never read these.
        "Actual Amount ($)": 99999,
        "Tax (7%)": 99999,
        "Invoice Total ($)": 99999,
    }])
    return path


@pytest.fixture
def unpriced_job_tracker(tmp_path):
    """A job with no price set anywhere — the on-the-spot case where a
    quote_amount override is required."""
    path = tmp_path / "scratch_tracker.xlsx"
    _build_scratch_tracker(path, [{
        "JobID (JOB-####)": "JOB-0002",
        "Customer Name / Company": "Blue Wave Cafe",
        "Service Date": "2026-08-24",
        "Service Type": "Window",
        "Crew / Technician": "Mike C.",
    }])
    return path


@pytest.fixture
def already_invoiced_job_tracker(tmp_path):
    path = tmp_path / "scratch_tracker.xlsx"
    _build_scratch_tracker(path, [{
        "JobID (JOB-####)": "JOB-0003",
        "Customer Name / Company": "Sunshine Realty LLC",
        "Quote Amount ($)": 150,
        "InvoiceID (INV-####)": "INV-0007",
    }])
    return path


# ══════════════════════════════════════════════════════════════════════════
# TESTS — happy path, using the job's own stored price
# ══════════════════════════════════════════════════════════════════════════

class TestUsesJobsOwnStoredPrice:
    def test_success_message(self, tmp_path, priced_job_tracker):
        result = _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001")
        assert "❌" not in result, f"Unexpected error: {result}"
        assert "✅ Invoice created: INV-0001" in result
        assert "NEW_INVOICE_ID=INV-0001" in result

    def test_amounts_computed_from_job_not_stale_formula_cells(self, tmp_path, priced_job_tracker):
        """The whole point of the design: 200 - 20 = 180 subtotal-after-
        discount * 1.07 = 192.60, NOT the poisoned 99999 formula cells."""
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001")
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        assert inv["Subtotal ($)"] == 200
        assert inv["Discount ($)"] == 20
        assert inv["Taxable Amt ($)"] == 180
        assert inv["Tax 7% ($)"] == pytest.approx(12.6)
        assert inv["TOTAL DUE ($)"] == pytest.approx(192.6)
        assert inv["Balance Due ($)"] == pytest.approx(192.6)
        assert inv["Amount Paid ($)"] == 0

    def test_customer_and_job_fields_copied_across(self, tmp_path, priced_job_tracker):
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001")
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        assert inv["JobID (JOB-####)"] == "JOB-0001"
        assert inv["CustomerID"] == "CUST-0001"
        assert inv["Customer Name / Company"] == "Torres Residence"
        assert inv["Customer Type"] == "Residential"
        assert inv["Service Type"] == "Window"
        assert inv["Description"] == "Full exterior window cleaning"

    def test_payment_status_defaults_unpaid(self, tmp_path, priced_job_tracker):
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001")
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        assert inv["Payment Status"] == "Unpaid"

    def test_due_date_defaults_net_30(self, tmp_path, priced_job_tracker):
        import datetime
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001")
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        expected = (datetime.date.today() + datetime.timedelta(days=30)).isoformat()
        assert inv["Due Date (Net 30)"] == expected

    def test_invoice_id_written_back_onto_job_row(self, tmp_path, priced_job_tracker):
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001")
        job = _read_job_row(priced_job_tracker, "JOB-0001")
        assert job["InvoiceID (INV-####)"] == "INV-0001"

    def test_job_row_not_mutated_when_no_override_given(self, tmp_path, priced_job_tracker):
        """No override passed → the job's own Quote/Discount/description
        must be left exactly as they were, not rewritten with themselves
        (rewriting is harmless here but the intent is 'untouched unless
        the caller explicitly overrides')."""
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001")
        job = _read_job_row(priced_job_tracker, "JOB-0001")
        assert job["Quote Amount ($)"] == 200
        assert job["Discount Applied ($)"] == 20

    def test_matches_by_partial_customer_name(self, tmp_path, priced_job_tracker):
        result = _run_create_invoice(tmp_path, priced_job_tracker, "Torres")
        assert "✅ Invoice created" in result


# ══════════════════════════════════════════════════════════════════════════
# TESTS — the "technician adjusts price on the spot" override path
# ══════════════════════════════════════════════════════════════════════════

class TestOnTheSpotOverrides:
    def test_quote_amount_override_used_over_job_value(self, tmp_path, priced_job_tracker):
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001", quote_amount=300)
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        assert inv["Subtotal ($)"] == 300

    def test_quote_amount_override_written_back_to_job(self, tmp_path, priced_job_tracker):
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001", quote_amount=300)
        job = _read_job_row(priced_job_tracker, "JOB-0001")
        assert job["Quote Amount ($)"] == 300

    def test_discount_override(self, tmp_path, priced_job_tracker):
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001",
                             quote_amount=100, discount=10)
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        assert inv["Discount ($)"] == 10
        assert inv["Taxable Amt ($)"] == 90

    def test_description_and_service_type_override(self, tmp_path, priced_job_tracker):
        _run_create_invoice(
            tmp_path, priced_job_tracker, "JOB-0001",
            quote_amount=250, description="Pressure wash driveway",
            service_type="Pressure Wash",
        )
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        assert inv["Description"] == "Pressure wash driveway"
        assert inv["Service Type"] == "Pressure Wash"
        job = _read_job_row(priced_job_tracker, "JOB-0001")
        assert job["Service Details / Notes"] == "Pressure wash driveway"
        assert job["Service Type"] == "Pressure Wash"

    def test_unpriced_job_requires_override(self, tmp_path, unpriced_job_tracker):
        result = _run_create_invoice(tmp_path, unpriced_job_tracker, "JOB-0002")
        assert "❌" in result
        assert "No price to invoice" in result

    def test_unpriced_job_succeeds_with_override(self, tmp_path, unpriced_job_tracker):
        result = _run_create_invoice(tmp_path, unpriced_job_tracker, "JOB-0002",
                                      quote_amount=220)
        assert "✅ Invoice created" in result
        inv = _read_invoice_by_id(unpriced_job_tracker, "INV-0001")
        assert inv["Subtotal ($)"] == 220

    def test_custom_tax_rate(self, tmp_path, priced_job_tracker):
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001",
                             quote_amount=100, discount=0, tax_rate=0.10)
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        assert inv["Tax 7% ($)"] == pytest.approx(10.0)
        assert inv["TOTAL DUE ($)"] == pytest.approx(110.0)

    def test_custom_due_days(self, tmp_path, priced_job_tracker):
        import datetime
        _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001", due_days=15)
        inv = _read_invoice_by_id(priced_job_tracker, "INV-0001")
        expected = (datetime.date.today() + datetime.timedelta(days=15)).isoformat()
        assert inv["Due Date (Net 30)"] == expected


# ══════════════════════════════════════════════════════════════════════════
# TESTS — validation / error paths
# ══════════════════════════════════════════════════════════════════════════

class TestValidation:
    def test_blank_job_identifier_rejected(self, tmp_path, priced_job_tracker):
        result = _run_create_invoice(tmp_path, priced_job_tracker, "")
        assert "❌" in result
        assert "cannot be blank" in result

    def test_nonexistent_job_returns_clear_error(self, tmp_path, priced_job_tracker):
        result = _run_create_invoice(tmp_path, priced_job_tracker, "JOB-DOES-NOT-EXIST")
        assert "❌" in result
        assert "No job found" in result

    def test_ambiguous_match_lists_candidates_and_refuses(self, tmp_path, tmp_path_factory):
        path = tmp_path / "scratch_tracker.xlsx"
        _build_scratch_tracker(path, [
            {"JobID (JOB-####)": "JOB-0001", "Customer Name / Company": "Window Co A",
             "Quote Amount ($)": 100},
            {"JobID (JOB-####)": "JOB-0002", "Customer Name / Company": "Window Co B",
             "Quote Amount ($)": 100},
        ])
        result = _run_create_invoice(tmp_path, path, "Window")
        assert "❌" in result
        assert "matches 2 jobs" in result
        assert "JOB-0001" in result and "JOB-0002" in result

    def test_negative_quote_amount_rejected(self, tmp_path, priced_job_tracker):
        result = _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001",
                                      quote_amount=-50)
        assert "❌" in result
        assert "cannot be negative" in result

    def test_discount_exceeding_quote_rejected(self, tmp_path, priced_job_tracker):
        result = _run_create_invoice(tmp_path, priced_job_tracker, "JOB-0001",
                                      quote_amount=100, discount=150)
        assert "❌" in result
        assert "cannot exceed" in result

    def test_already_invoiced_job_refused(self, tmp_path, already_invoiced_job_tracker):
        result = _run_create_invoice(tmp_path, already_invoiced_job_tracker, "JOB-0003")
        assert "❌" in result
        assert "INV-0007" in result
        assert "already has an invoice" in result

    def test_already_invoiced_job_creates_no_new_row(self, tmp_path, already_invoiced_job_tracker):
        _run_create_invoice(tmp_path, already_invoiced_job_tracker, "JOB-0003")
        assert _read_invoice_rows(already_invoiced_job_tracker) == []

    def test_missing_jobs_schedule_sheet_errors_cleanly(self, tmp_path):
        path = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.save(path)
        result = _run_create_invoice(tmp_path, path, "JOB-0001")
        assert "❌" in result
        assert "Jobs_Schedule" in result


# ══════════════════════════════════════════════════════════════════════════
# TESTS — sequencing, auto-creation, multi-invoice behavior
# ══════════════════════════════════════════════════════════════════════════

class TestSequencingAndSheetCreation:
    def test_invoice_ids_increment_sequentially(self, tmp_path):
        path = tmp_path / "scratch_tracker.xlsx"
        _build_scratch_tracker(path, [
            {"JobID (JOB-####)": "JOB-0001", "Customer Name / Company": "A",
             "Quote Amount ($)": 100},
            {"JobID (JOB-####)": "JOB-0002", "Customer Name / Company": "B",
             "Quote Amount ($)": 100},
        ])
        r1 = _run_create_invoice(tmp_path, path, "JOB-0001")
        r2 = _run_create_invoice(tmp_path, path, "JOB-0002")
        assert "NEW_INVOICE_ID=INV-0001" in r1
        assert "NEW_INVOICE_ID=INV-0002" in r2

    def test_invoices_sheet_auto_created_when_missing(self, tmp_path):
        path = tmp_path / "scratch_tracker.xlsx"
        _build_scratch_tracker(
            path,
            [{"JobID (JOB-####)": "JOB-0001", "Customer Name / Company": "A",
              "Quote Amount ($)": 100}],
            include_invoices_sheet=False,
        )
        result = _run_create_invoice(tmp_path, path, "JOB-0001")
        assert "✅ Invoice created: INV-0001" in result
        inv = _read_invoice_by_id(path, "INV-0001")
        assert inv["Subtotal ($)"] == 100

    def test_next_id_continues_from_existing_invoices(self, tmp_path):
        """A tracker that already has INV-0001..INV-0004 (e.g. seeded data)
        must continue from INV-0005, not restart at INV-0001."""
        path = tmp_path / "scratch_tracker.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws_jobs = wb.create_sheet("Jobs_Schedule")
        ws_jobs.append(["📅  JOBS & SCHEDULE"])
        ws_jobs.append(JOBS_HEADERS)
        row = ["" for _ in _JOBS_CANON]
        row[_JOBS_CANON.index("JobID (JOB-####)")] = "JOB-0009"
        row[_JOBS_CANON.index("Customer Name / Company")] = "Late Entry"
        row[_JOBS_CANON.index("Quote Amount ($)")] = 100
        ws_jobs.append(row)

        ws_inv = wb.create_sheet("Invoices")
        ws_inv.append(["🧾  INVOICES"])
        ws_inv.append(INVOICES_HEADERS)
        for n in range(1, 5):
            ws_inv.append([f"INV-{n:04d}"] + [""] * (len(_INV_CANON) - 1))
        wb.save(path)

        result = _run_create_invoice(tmp_path, path, "JOB-0009")
        assert "NEW_INVOICE_ID=INV-0005" in result


def test_source_exists():
    assert MCP_FILE.exists(), f"ai_prowler_mcp.py not found at {MCP_FILE}"


# ══════════════════════════════════════════════════════════════════════════
# STRUCTURAL TESTS — crew scoping wiring, tool registration
# (matching this codebase's existing convention — see
#  TestSharedCrewScopeHelper in test_server_mode_jobs_pwa.py and
#  TestBackendToolAllowedInBothModes in
#  test_pwa_invoice_button_labels_and_sms_gating.py — of verifying
#  server-mode wiring structurally rather than fully mocking a server-mode
#  ctx/user/crew-assignment in isolated subprocess tests.)
# ══════════════════════════════════════════════════════════════════════════

@pytest.fixture(scope="module")
def mcp_source():
    return MCP_FILE.read_text(encoding="utf-8")


class TestCrewScopingWired:
    def test_create_invoice_uses_shared_crew_scope_helper(self, mcp_source):
        idx = mcp_source.index("def _create_invoice_impl(")
        end_idx = mcp_source.index("\n@mcp.tool()\ndef create_invoice(", idx)
        body = mcp_source[idx:end_idx]
        assert "_job_crew_scope(ctx, fp)" in body
        assert "_crew_name_in_cell(" in body

    def test_refuses_when_not_assigned_to_caller(self, mcp_source):
        idx = mcp_source.index("def _create_invoice_impl(")
        end_idx = mcp_source.index("\n@mcp.tool()\ndef create_invoice(", idx)
        body = mcp_source[idx:end_idx]
        assert "not assigned to you" in body


class TestToolRegisteredInBothPwaModes:
    def test_allowed_in_personal_mode(self, mcp_source):
        idx = mcp_source.index("_allowed_tools = {")
        nearby = mcp_source[idx:idx + 800]
        assert '"create_invoice"' in nearby

    def test_allowed_in_server_mode(self, mcp_source):
        idx = mcp_source.index("_srv_pa_allowed = {")
        nearby = mcp_source[idx:idx + 800]
        assert '"create_invoice"' in nearby

    def test_tool_is_mcp_registered(self, mcp_source):
        assert "@mcp.tool()\ndef create_invoice(" in mcp_source

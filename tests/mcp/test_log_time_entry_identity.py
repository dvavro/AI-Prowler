"""
tests/mcp/test_log_time_entry_identity.py
============================================
Tests for log_time_entry()'s server-mode identity and job-specificity fixes.

Background
----------
Two related problems were fixed:

1. Job identification was too loose: an unmatched job_identifier silently
   fell back to using the raw string itself as both JobID and customer
   name (never erroring), and an ambiguous identifier matching MULTIPLE
   jobs silently took the first row found. Neither "clock in" nor
   "clock out" was actually pinned to a real, specific job.

2. No caller identity at all: the "Crew / Technician" field was copied
   from whatever the job's spreadsheet assignment already said, not from
   who actually called the tool — and "stop" could close out ANY open
   entry for a job, including one a different employee opened, since
   ownership was never checked.

Fixed by:
  - Requiring an unambiguous single match against Jobs_Schedule for both
    start and stop (zero matches -> error; 2+ matches -> error listing
    candidates).
  - A new "Logged By (User ID)" column on the TimeLog sheet.
  - Server mode: "start" stamps the CALLER's own name into
    Crew / Technician (not the job's pre-assigned crew), and the
    already-open check is scoped to the caller's own entries only.
  - Server mode: "stop" only finds/closes an open entry the SAME caller
    opened — a coworker's open entry for the same job is invisible to
    "stop", with a message distinguishing "no entry" from
    "someone else has one."
  - Personal mode is completely unchanged in behavior (single user, so
    ownership scoping is a no-op — verified by the pre-existing
    test_contractor_tools.py CT-18..22 suite, which still passes as-is).
"""

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
import openpyxl

_SRC = Path(__file__).resolve().parent.parent.parent
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))


@pytest.fixture(scope="module")
def mcp_mod():
    import ai_prowler_mcp as ap
    ap._prewarm_event.set()
    return ap


def _make_ctx(user):
    if user is None:
        return None
    ctx = MagicMock()
    ctx.request_context.request.state.user = user
    return ctx


def _user(role, uid, name):
    return {"id": uid, "name": name, "role": role, "status": "active"}


def _make_spreadsheet(tmp_path, jobs):
    """jobs: list of (JobID, Customer, Crew) tuples."""
    fp = tmp_path / "tracker.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs_Schedule"
    ws.append(["JobID (JOB-####)", "Customer Name / Company", "Crew / Technician", "Service Type"])
    for jid, cust, crew in jobs:
        ws.append([jid, cust, crew, "Window Washing"])
    wb.save(str(fp))
    return str(fp)


class TestAmbiguousJobRejected:

    def test_no_match_at_all_rejected(self, mcp_mod, monkeypatch, tmp_path):
        fp = _make_spreadsheet(tmp_path, [("JOB-0001", "Crabby's Daytona", "")])
        monkeypatch.setattr(mcp_mod, "_current_user", lambda ctx: None)
        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_mod.log_time_entry(
                job_identifier="totally nonexistent job", action="start", filepath=fp, ctx=None
            )
        assert "❌" in result
        assert "No job found" in result

    def test_ambiguous_match_rejected_with_candidates(self, mcp_mod, monkeypatch, tmp_path):
        """Two jobs both containing 'Daytona' — must not silently pick one."""
        fp = _make_spreadsheet(tmp_path, [
            ("JOB-0001", "Crabby's Daytona", ""),
            ("JOB-0002", "Sunshine Realty Daytona", ""),
        ])
        monkeypatch.setattr(mcp_mod, "_current_user", lambda ctx: None)
        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_mod.log_time_entry(
                job_identifier="Daytona", action="start", filepath=fp, ctx=None
            )
        assert "❌" in result
        assert "matches 2 jobs" in result
        assert "JOB-0001" in result
        assert "JOB-0002" in result

    def test_exact_unique_match_succeeds(self, mcp_mod, monkeypatch, tmp_path):
        fp = _make_spreadsheet(tmp_path, [
            ("JOB-0001", "Crabby's Daytona", ""),
            ("JOB-0002", "Sunshine Realty Daytona", ""),
        ])
        monkeypatch.setattr(mcp_mod, "_current_user", lambda ctx: None)
        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="start", filepath=fp, ctx=None
            )
        assert "Clocked IN" in result


class TestServerModeIdentity:

    def test_crew_field_uses_caller_not_job_assignment(self, mcp_mod, monkeypatch, tmp_path):
        """Core fix: Crew / Technician gets the ACTUAL caller's name, not
        whatever the job's spreadsheet assignment says."""
        fp = _make_spreadsheet(tmp_path, [("JOB-0001", "Crabby's Daytona", "Someone Else Entirely")])
        monkeypatch.setattr(mcp_mod, "_get_default_spreadsheet_path", lambda: fp)
        jake = _user("field_crew", "jake-r", "Jake R")
        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="start", filepath=fp, ctx=_make_ctx(jake)
            )
        assert "Jake R" in result
        assert "Someone Else Entirely" not in result

        wb = openpyxl.load_workbook(fp)
        ws = wb["TimeLog"]
        row = [c.value for c in ws[3]]  # row 1 = title, row 2 = headers, row 3 = first entry
        assert "jake-r" in row  # Logged By (User ID)

    def test_second_user_can_start_own_entry_same_job(self, mcp_mod, monkeypatch, tmp_path):
        """Two employees working the SAME job must each be able to clock
        in independently — one's open entry must not block the other."""
        fp = _make_spreadsheet(tmp_path, [("JOB-0001", "Crabby's Daytona", "")])
        monkeypatch.setattr(mcp_mod, "_get_default_spreadsheet_path", lambda: fp)
        jake = _user("field_crew", "jake-r", "Jake R")
        karen = _user("staff", "karen-s", "Karen S")

        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            r1 = mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="start", filepath=fp, ctx=_make_ctx(jake)
            )
            r2 = mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="start", filepath=fp, ctx=_make_ctx(karen)
            )
        assert "Clocked IN" in r1
        assert "Clocked IN" in r2  # NOT blocked by Jake's still-open entry

    def test_same_user_double_start_still_blocked(self, mcp_mod, monkeypatch, tmp_path):
        fp = _make_spreadsheet(tmp_path, [("JOB-0001", "Crabby's Daytona", "")])
        monkeypatch.setattr(mcp_mod, "_get_default_spreadsheet_path", lambda: fp)
        jake = _user("field_crew", "jake-r", "Jake R")
        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="start", filepath=fp, ctx=_make_ctx(jake)
            )
            r2 = mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="start", filepath=fp, ctx=_make_ctx(jake)
            )
        assert "already open" in r2

    def test_cannot_stop_coworkers_open_entry(self, mcp_mod, monkeypatch, tmp_path):
        """The actual leak this closes: Karen must NOT be able to clock out
        Jake's open shift, even for the same job."""
        fp = _make_spreadsheet(tmp_path, [("JOB-0001", "Crabby's Daytona", "")])
        monkeypatch.setattr(mcp_mod, "_get_default_spreadsheet_path", lambda: fp)
        jake = _user("field_crew", "jake-r", "Jake R")
        karen = _user("staff", "karen-s", "Karen S")

        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="start", filepath=fp, ctx=_make_ctx(jake)
            )
            r_karen_stop = mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="stop", filepath=fp, ctx=_make_ctx(karen)
            )
        assert "❌" in r_karen_stop
        assert "someone else" in r_karen_stop.lower()

        # Jake himself must still be able to stop his own entry.
        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            r_jake_stop = mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="stop", filepath=fp, ctx=_make_ctx(jake)
            )
        assert "Clocked OUT" in r_jake_stop or "⏱️" in r_jake_stop

    def test_personal_mode_completely_unaffected(self, mcp_mod, tmp_path):
        """Personal mode: identical behavior to before — job's assigned
        crew is used, no per-user ownership scoping applies at all."""
        fp = _make_spreadsheet(tmp_path, [("JOB-0001", "Crabby's Daytona", "David Vavro")])
        with patch.object(mcp_mod, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_mod.log_time_entry(
                job_identifier="JOB-0001", action="start", filepath=fp, ctx=None
            )
        assert "David Vavro" in result

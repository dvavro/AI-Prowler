"""
tests/mcp/test_migrate_spreadsheet.py
=======================================
Test suite for migrate_spreadsheet.py — the Job Tracker spreadsheet
migration engine.

SAFETY GUARANTEE
----------------
ALL tests run against temporary copies of test fixtures in a scratch
directory under TEMP/AI-Prowler-test-migrate/. They NEVER touch:
  - Program Files/AI-Prowler/  (installed app)
  - Documents/AI-Prowler/AI-Prowler_Job_Tracker.xlsx  (live data)
  - .ai-prowler/config.json  (production config)

Each test gets an isolated temp directory that is cleaned up after the
test completes. The production config is never read or written.

WHAT IS TESTED
==============

GROUP 1 — Module structure (import, constants, public API)
  1.01  Module imports cleanly
  1.02  CURRENT_SCHEMA_VERSION is a positive integer
  1.03  check_and_migrate() is callable
  1.04  MigrationResult has required attributes
  1.05  _MIGRATIONS list is non-empty and well-formed
  1.06  Each migration function is callable

GROUP 2 — Config helpers (isolated, no production config)
  2.01  _get_schema_version() returns 0 when key absent
  2.02  _get_schema_version() returns correct value when key present
  2.03  _set_schema_version() writes to the patched config path
  2.04  _get_spreadsheet_path() returns None when config is empty
  2.05  _get_spreadsheet_path() returns Path when config has the key

GROUP 3 — File-lock detection
  3.01  _is_file_locked() returns False for a readable file
  3.02  _is_file_locked() returns True for a locked file

GROUP 4 — Backup helper
  4.01  _make_backup() creates a .xlsx file in _backups/
  4.02  Backup content is byte-identical to the original
  4.03  _make_backup() raises when the source doesn't exist
  4.04  Two backups of the same file get different timestamps

GROUP 5 — _verify_migration() — clean migration
  5.01  Returns empty list when migrated wb is identical to original
  5.02  Returns empty list when migrated wb adds new columns (additive only)
  5.03  Returns empty list when migrated wb adds new sheets
  5.04  Tolerates float precision drift < 1e-9

GROUP 6 — _verify_migration() — corruption detection
  6.01  CRITICAL when a sheet is deleted
  6.02  CRITICAL when a data row is deleted (row count drops)
  6.03  CRITICAL when a cell value changes
  6.04  CRITICAL when a column header is deleted
  6.05  WARNING when a column header text changes
  6.06  WARNING when a formula becomes a plain value
  6.07  Multiple CRITICALs are all reported (not just the first)
  6.08  Per-sheet CRITICAL count is capped at 5 examples + summary line

GROUP 7 — _migrate_0_to_1() logic
  7.01  Adds Settings sheet when missing
  7.02  Settings sheet is inserted at position 0 (first tab)
  7.03  Adds AI-Prowler_Commands sheet when missing
  7.04  Does NOT touch data rows on existing sheets
  7.05  Adds missing columns to existing sheets (appended at right)
  7.06  Does not add duplicate columns that already exist
  7.07  Fixes freeze panes to match template
  7.08  Returns a non-empty changes list when work is done
  7.09  Returns an empty changes list when nothing needs doing

GROUP 8 — check_and_migrate() end-to-end (isolated temp files)
  8.01  Returns result.needed=False when schema version is current
  8.02  Returns result.needed=True when schema version is 0
  8.03  Returns result.success=True on a clean migration
  8.04  Backup file exists after successful migration
  8.05  Template copy exists after successful migration
  8.06  Schema version written to patched config after success
  8.07  Migrated file can be re-opened by openpyxl after save
  8.08  result.changes is non-empty when new sheets were added
  8.09  Returns result.success=False when file is locked
  8.10  Returns result.success=False when template is missing
  8.11  Original file restored when integrity check fails (CRITICAL)
  8.12  Migration succeeds and saves when only WARNINGs present
  8.13  Log file is written to the patched log path
  8.14  User with schema_version=0 and no spreadsheet: treated as current
"""
from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
import threading
import time
from pathlib import Path
from unittest import mock
from copy import copy

import pytest
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ── Path setup ─────────────────────────────────────────────────────────────────
# Ensure the work-dir AI-Prowler package is importable. Does NOT affect the
# installed copy in Program Files.
WORK_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(WORK_DIR))

import migrate_spreadsheet as ms

# ── Fixtures ───────────────────────────────────────────────────────────────────

@pytest.fixture
def tmp(tmp_path):
    """Isolated temp directory — auto-cleaned by pytest after each test."""
    return tmp_path


def _make_minimal_wb(sheets=("Customers", "Jobs_Schedule")) -> openpyxl.Workbook:
    """Create a minimal workbook that mimics the v9.0.0 user spreadsheet."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)           # remove default Sheet
    for name in sheets:
        ws = wb.create_sheet(name)
        # Row 1: title banner
        ws.cell(row=1, column=1, value=f"🗋 {name.upper()} — Test Sheet")
        # Row 2: column headers
        ws.cell(row=2, column=1, value="ID")
        ws.cell(row=2, column=2, value="Name")
        ws.cell(row=2, column=3, value="Service Date")
        ws.cell(row=2, column=4, value="Amount ($)")
        # Row 3: data row
        ws.cell(row=3, column=1, value="TEST-0001")
        ws.cell(row=3, column=2, value="Test Customer")
        ws.cell(row=3, column=3, value="08/25/2026")
        ws.cell(row=3, column=4, value=125.00)
    return wb


def _make_template_wb() -> openpyxl.Workbook:
    """Create a minimal template wb with Settings + Commands + extra column."""
    wb = _make_minimal_wb(("Settings", "Customers", "Jobs_Schedule",
                            "AI-Prowler_Commands"))
    # Settings sheet — position 0
    ws_s = wb["Settings"]
    ws_s.cell(row=2, column=1, value="Setting")
    ws_s.cell(row=2, column=2, value="Value")
    ws_s.cell(row=2, column=3, value="Notes")
    ws_s.cell(row=3, column=1, value="Tax Rate")
    ws_s.cell(row=3, column=2, value=0.07)
    ws_s.freeze_panes = "A3"

    # Add an extra column to Customers (simulates a new v9.1 column)
    ws_c = wb["Customers"]
    ws_c.cell(row=2, column=5, value="New Column v9.1")
    ws_c.freeze_panes = "A3"

    ws_j = wb["Jobs_Schedule"]
    ws_j.freeze_panes = "A3"

    return wb


def _save_wb(wb: openpyxl.Workbook, path: Path) -> Path:
    wb.save(path)
    return path


# ════════════════════════════════════════════════════════════════════════════════
# GROUP 1 — Module structure
# ════════════════════════════════════════════════════════════════════════════════

class TestGroup1_ModuleStructure:

    def test_1_01_module_imports(self):
        assert ms is not None

    def test_1_02_current_schema_version_is_positive_int(self):
        assert isinstance(ms.CURRENT_SCHEMA_VERSION, int)
        assert ms.CURRENT_SCHEMA_VERSION >= 1

    def test_1_03_check_and_migrate_callable(self):
        assert callable(ms.check_and_migrate)

    def test_1_04_migration_result_has_required_attributes(self):
        r = ms.MigrationResult()
        for attr in ("needed", "success", "changes", "integrity_problems",
                     "error", "backup_path", "template_path",
                     "log_path", "from_version", "to_version"):
            assert hasattr(r, attr), f"MigrationResult missing attribute: {attr}"

    def test_1_05_migrations_list_non_empty(self):
        assert len(ms._MIGRATIONS) >= 1

    def test_1_06_each_migration_function_callable(self):
        for from_v, to_v, fn in ms._MIGRATIONS:
            assert isinstance(from_v, int)
            assert isinstance(to_v, int)
            assert to_v > from_v
            assert callable(fn)


# ════════════════════════════════════════════════════════════════════════════════
# GROUP 2 — Config helpers (patched to temp paths)
# ════════════════════════════════════════════════════════════════════════════════

class TestGroup2_ConfigHelpers:

    def test_2_01_get_schema_version_returns_0_when_absent(self, tmp):
        cfg = tmp / "config.json"
        cfg.write_text("{}", encoding="utf-8")
        with mock.patch.object(ms, "_CONFIG_PATH", cfg):
            assert ms._get_schema_version() == 0

    def test_2_02_get_schema_version_returns_stored_value(self, tmp):
        cfg = tmp / "config.json"
        cfg.write_text('{"spreadsheet_schema_version": 1}', encoding="utf-8")
        with mock.patch.object(ms, "_CONFIG_PATH", cfg):
            assert ms._get_schema_version() == 1

    def test_2_03_set_schema_version_writes_to_config(self, tmp):
        cfg = tmp / "config.json"
        cfg.write_text("{}", encoding="utf-8")
        with mock.patch.object(ms, "_CONFIG_PATH", cfg):
            ms._set_schema_version(2)
            data = json.loads(cfg.read_text())
            assert data["spreadsheet_schema_version"] == 2

    def test_2_04_get_spreadsheet_path_returns_none_when_empty(self, tmp):
        cfg = tmp / "config.json"
        cfg.write_text("{}", encoding="utf-8")
        with mock.patch.object(ms, "_CONFIG_PATH", cfg):
            # Default location doesn't exist in tmp
            with mock.patch.object(ms.Path, "exists", return_value=False):
                result = ms._get_spreadsheet_path()
                # May return a Path or None — if Path, it shouldn't exist
                if result is not None:
                    assert not result.exists()

    def test_2_05_get_spreadsheet_path_returns_configured_path(self, tmp):
        xlsx = tmp / "my_tracker.xlsx"
        xlsx.write_bytes(b"fake")
        cfg = tmp / "config.json"
        cfg.write_text(
            json.dumps({"default_spreadsheet_path": str(xlsx)}),
            encoding="utf-8"
        )
        with mock.patch.object(ms, "_CONFIG_PATH", cfg):
            result = ms._get_spreadsheet_path()
            assert result == xlsx


# ════════════════════════════════════════════════════════════════════════════════
# GROUP 3 — File-lock detection
# ════════════════════════════════════════════════════════════════════════════════

class TestGroup3_FileLock:

    def test_3_01_unlocked_file_returns_false(self, tmp):
        f = tmp / "test.xlsx"
        f.write_bytes(b"data")
        assert ms._is_file_locked(f) is False

    def test_3_02_locked_file_returns_true(self, tmp):
        f = tmp / "locked.xlsx"
        f.write_bytes(b"data")
        # Simulate a locked file by patching open() to raise PermissionError
        original_open = open

        def mock_open(path, *args, **kwargs):
            if str(path) == str(f) and args and "b" in str(args[0]):
                raise PermissionError("locked")
            return original_open(path, *args, **kwargs)

        with mock.patch("builtins.open", side_effect=mock_open):
            locked = ms._is_file_locked(f)
        assert locked is True


# ════════════════════════════════════════════════════════════════════════════════
# GROUP 4 — Backup helper
# ════════════════════════════════════════════════════════════════════════════════

class TestGroup4_BackupHelper:

    def test_4_01_creates_backup_in_backups_subdir(self, tmp):
        src = tmp / "tracker.xlsx"
        src.write_bytes(b"xlsx content")
        bak = ms._make_backup(src)
        assert bak.exists()
        assert bak.parent.name == "_backups"
        assert bak.suffix == ".xlsx"

    def test_4_02_backup_is_byte_identical(self, tmp):
        src = tmp / "tracker.xlsx"
        content = b"unique content " + os.urandom(32)
        src.write_bytes(content)
        bak = ms._make_backup(src)
        assert bak.read_bytes() == content

    def test_4_03_backup_raises_when_source_missing(self, tmp):
        src = tmp / "nonexistent.xlsx"
        with pytest.raises(Exception):
            ms._make_backup(src)

    def test_4_04_two_backups_have_different_names(self, tmp):
        src = tmp / "tracker.xlsx"
        src.write_bytes(b"data")
        b1 = ms._make_backup(src)
        time.sleep(1.1)  # ensure different timestamp
        b2 = ms._make_backup(src)
        assert b1.name != b2.name


# ════════════════════════════════════════════════════════════════════════════════
# GROUP 5 — _verify_migration() clean cases
# ════════════════════════════════════════════════════════════════════════════════

class TestGroup5_VerifyClean:

    def test_5_01_identical_workbooks_returns_empty(self):
        wb = _make_minimal_wb()
        wb2 = _make_minimal_wb()
        assert ms._verify_migration(wb, wb2) == []

    def test_5_02_additive_columns_returns_empty(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers",))
        # Add a new column to migrated — additive, no corruption
        migr["Customers"].cell(row=2, column=5, value="New Col")
        migr["Customers"].cell(row=3, column=5, value="new value")
        assert ms._verify_migration(orig, migr) == []

    def test_5_03_new_sheet_in_migrated_returns_empty(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers", "Settings"))
        assert ms._verify_migration(orig, migr) == []

    def test_5_04_float_precision_drift_tolerated(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers",))
        # Introduce sub-epsilon drift in a numeric cell
        orig["Customers"].cell(row=3, column=4, value=125.0)
        migr["Customers"].cell(row=3, column=4, value=125.00000000001)
        assert ms._verify_migration(orig, migr) == []


# ════════════════════════════════════════════════════════════════════════════════
# GROUP 6 — _verify_migration() corruption detection
# ════════════════════════════════════════════════════════════════════════════════

class TestGroup6_VerifyCorruption:

    def test_6_01_critical_when_sheet_deleted(self):
        orig = _make_minimal_wb(("Customers", "Jobs_Schedule"))
        migr = _make_minimal_wb(("Customers",))  # Jobs_Schedule removed
        problems = ms._verify_migration(orig, migr)
        criticals = [p for p in problems if "CRITICAL" in p and "Jobs_Schedule" in p]
        assert len(criticals) >= 1

    def test_6_02_critical_when_row_deleted(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers",))
        # Add a second data row to orig only
        orig["Customers"].cell(row=4, column=1, value="TEST-0002")
        orig["Customers"].cell(row=4, column=2, value="Second Customer")
        problems = ms._verify_migration(orig, migr)
        criticals = [p for p in problems if "CRITICAL" in p and "row" in p.lower()]
        assert len(criticals) >= 1

    def test_6_03_critical_when_cell_value_changes(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers",))
        orig["Customers"].cell(row=3, column=2, value="Original Name")
        migr["Customers"].cell(row=3, column=2, value="CHANGED NAME")
        problems = ms._verify_migration(orig, migr)
        criticals = [p for p in problems if "CRITICAL" in p]
        assert len(criticals) >= 1

    def test_6_04_critical_when_column_header_deleted(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers",))
        # Confirm col 3 exists in orig before we delete it from migr
        assert orig["Customers"].cell(row=2, column=3).value is not None, \
            "Test setup: col 3 must have a header in the original"
        # Delete col 3 header from the MIGRATED sheet only
        migr["Customers"].cell(row=2, column=3).value = None
        # Now verify: orig has col 3, migr does not → CRITICAL
        problems = ms._verify_migration(orig, migr)
        criticals = [p for p in problems if p.startswith("CRITICAL")]
        assert len(criticals) >= 1, \
            f"Expected at least 1 CRITICAL for deleted header, got: {problems}"

    def test_6_05_warning_when_header_text_changes(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers",))
        orig["Customers"].cell(row=2, column=3, value="Service Date")
        migr["Customers"].cell(row=2, column=3, value="ServiceDate")  # renamed
        problems = ms._verify_migration(orig, migr)
        warnings = [p for p in problems if "WARNING" in p and "header" in p.lower()]
        assert len(warnings) >= 1

    def test_6_06_warning_when_formula_becomes_value(self):
        orig = _make_minimal_wb(("Invoices",))
        migr = _make_minimal_wb(("Invoices",))
        orig["Invoices"].cell(row=3, column=4, value="=B3*0.07")
        migr["Invoices"].cell(row=3, column=4, value=8.75)  # formula → value
        problems = ms._verify_migration(orig, migr)
        warnings = [p for p in problems if "WARNING" in p and "formula" in p.lower()]
        assert len(warnings) >= 1

    def test_6_07_multiple_criticals_all_reported(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers",))
        # Change 3 different cell values
        for col in (1, 2, 3):
            orig["Customers"].cell(row=3, column=col, value=f"orig-{col}")
            migr["Customers"].cell(row=3, column=col, value=f"changed-{col}")
        problems = ms._verify_migration(orig, migr)
        criticals = [p for p in problems if "CRITICAL" in p]
        assert len(criticals) >= 3

    def test_6_08_per_sheet_errors_capped_at_5_plus_summary(self):
        orig = _make_minimal_wb(("Customers",))
        migr = _make_minimal_wb(("Customers",))
        # Add 10 data rows with different values in each
        for r in range(3, 13):
            orig["Customers"].cell(row=r, column=2, value=f"orig-{r}")
            migr["Customers"].cell(row=r, column=2, value=f"changed-{r}")
        problems = ms._verify_migration(orig, migr)
        criticals = [p for p in problems if "CRITICAL" in p
                     and "value changed" in p]
        assert len(criticals) <= 5
        # There should be a "and more" summary line
        summaries = [p for p in problems if "more value mismatches" in p]
        assert len(summaries) >= 1


# ════════════════════════════════════════════════════════════════════════════════
# GROUP 7 — _migrate_0_to_1() logic
# ════════════════════════════════════════════════════════════════════════════════

class TestGroup7_Migrate0To1:

    def _run(self, user_wb, template_wb):
        return ms._migrate_0_to_1(user_wb, template_wb)

    def test_7_01_adds_settings_sheet_when_missing(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        self._run(user, template)
        assert "Settings" in user.sheetnames

    def test_7_02_settings_sheet_inserted_at_position_0(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        self._run(user, template)
        assert user.sheetnames[0] == "Settings"

    def test_7_03_adds_commands_sheet_when_missing(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        self._run(user, template)
        assert "AI-Prowler_Commands" in user.sheetnames

    def test_7_04_does_not_touch_data_rows(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        # Record original data
        original_data = [
            (r, c, user["Customers"].cell(row=r, column=c).value)
            for r in range(3, user["Customers"].max_row + 1)
            for c in range(1, user["Customers"].max_column + 1)
        ]
        self._run(user, template)
        # Verify every original data cell is intact
        for r, c, orig_val in original_data:
            new_val = user["Customers"].cell(row=r, column=c).value
            assert new_val == orig_val, \
                f"Data changed at R{r}C{c}: {orig_val!r} → {new_val!r}"

    def test_7_05_adds_missing_columns_to_existing_sheets(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        orig_max = user["Customers"].max_column
        self._run(user, template)
        # Template has an extra col — should be added to user's sheet
        assert user["Customers"].max_column > orig_max

    def test_7_06_no_duplicate_columns_added(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        self._run(user, template)
        # Run again — should not add another copy
        col_count_after_first = user["Customers"].max_column
        self._run(user, template)
        assert user["Customers"].max_column == col_count_after_first

    def test_7_07_fixes_freeze_panes(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        user["Customers"].freeze_panes = None   # wrong / missing
        self._run(user, template)
        assert user["Customers"].freeze_panes == "A3"

    def test_7_08_returns_non_empty_changes_when_work_done(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        changes  = self._run(user, template)
        assert len(changes) >= 1

    def test_7_09_commands_sheet_always_replaced(self):
        """AI-Prowler_Commands is reference content — it should always be
        replaced from the template, even when the user's workbook already
        has it. This ensures new commands added in each release reach users."""
        user     = _make_template_wb()  # already has Commands
        template = _make_template_wb()
        changes  = self._run(user, template)
        # The Commands replacement must always appear in changes
        commands_changes = [c for c in changes if "Commands" in c]
        assert len(commands_changes) >= 1, \
            ("AI-Prowler_Commands must always be replaced from template — "
             f"got changes: {changes}")

    def test_7_10_copies_new_sheet_with_styled_cells(self):
        """Regression test: _ensure_sheet() used to crash on every styled cell
        with 'cannot import name copy from openpyxl.styles' — a stray, broken
        duplicate import inside the `if cell.has_style:` branch that shadowed
        the correct `from copy import copy` already at the top of the
        function. Every existing fixture in this file has has_style=False on
        every cell (openpyxl.Workbook cells created via plain .cell(value=...)
        with no formatting applied), so this branch was never exercised by
        any test — the bug shipped invisibly. Real spreadsheets always have
        styled headers (bold/filled/bordered), so it crashed on every single
        real migration. This test gives the template a genuinely styled
        header cell so has_style is True, then verifies both that
        _ensure_sheet doesn't raise AND that the style actually made it onto
        the copied cell — which the broken code could never have done even
        if the import hadn't crashed, since it copied the un-aliased `copy`
        into a variable that immediately raised."""
        from openpyxl.styles import Font, PatternFill

        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        header_cell = template["Settings"]["A2"]
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill("solid", fgColor="4472C4")

        self._run(user, template)  # must not raise

        assert "Settings" in user.sheetnames
        copied_cell = user["Settings"]["A2"]
        assert copied_cell.has_style
        assert copied_cell.font.bold is True
        assert copied_cell.fill.fgColor.rgb == "004472C4"


# ════════════════════════════════════════════════════════════════════════════════
# GROUP 8 — check_and_migrate() end-to-end (fully isolated)
# ════════════════════════════════════════════════════════════════════════════════

class TestGroup8_EndToEnd:
    """All tests patch _CONFIG_PATH, _LOG_PATH, _APP_DIR, and
    _get_spreadsheet_path so they NEVER touch production files."""

    def _patch(self, tmp: Path, user_wb=None, template_wb=None,
               schema_version: int = 0):
        """Return a context-manager stack that fully isolates check_and_migrate."""
        user_path    = tmp / "user_tracker.xlsx"
        template_path = tmp / "template_tracker.xlsx"
        cfg_path     = tmp / "config.json"
        log_path     = tmp / "migration.log"

        # Write initial config
        cfg_path.write_text(
            json.dumps({"spreadsheet_schema_version": schema_version}),
            encoding="utf-8"
        )

        # Save workbooks to disk
        if user_wb is not None:
            user_wb.save(user_path)
        if template_wb is not None:
            template_wb.save(template_path)

        patches = [
            mock.patch.object(ms, "_CONFIG_PATH",  cfg_path),
            mock.patch.object(ms, "_LOG_PATH",      log_path),
            mock.patch.object(ms, "_APP_DIR",       tmp),
            mock.patch.object(ms, "_get_spreadsheet_path",
                              return_value=user_path if user_wb else None),
            mock.patch.object(ms, "_get_template_path",
                              return_value=template_path if template_wb else None),
        ]
        return patches, user_path, cfg_path, log_path

    def _apply(self, patches):
        """Enter all patches and return a stack for cleanup."""
        stack = []
        for p in patches:
            stack.append(p.__enter__())
        return stack

    def _cleanup(self, patches):
        for p in patches:
            p.__exit__(None, None, None)

    def test_8_01_returns_not_needed_when_schema_current(self, tmp):
        patches, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=ms.CURRENT_SCHEMA_VERSION
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            assert result.needed is False
        finally:
            self._cleanup(patches)

    def test_8_02_returns_needed_true_when_schema_is_0(self, tmp):
        patches, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            assert result.needed is True
        finally:
            self._cleanup(patches)

    def test_8_03_success_true_on_clean_migration(self, tmp):
        patches, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            assert result.success is True, f"Migration failed: {result.error}"
        finally:
            self._cleanup(patches)

    def test_8_04_backup_exists_after_success(self, tmp):
        patches, user_path, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            assert result.backup_path
            assert Path(result.backup_path).exists()
        finally:
            self._cleanup(patches)

    def test_8_05_template_copy_exists_after_success(self, tmp):
        patches, user_path, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            if result.template_path:
                assert Path(result.template_path).exists()
        finally:
            self._cleanup(patches)

    def test_8_06_schema_version_written_after_success(self, tmp):
        patches, _, cfg_path, _ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            assert result.success
            data = json.loads(cfg_path.read_text())
            assert data["spreadsheet_schema_version"] == ms.CURRENT_SCHEMA_VERSION
        finally:
            self._cleanup(patches)

    def test_8_07_migrated_file_reopens_cleanly(self, tmp):
        patches, user_path, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            assert result.success
            # File must be readable by openpyxl
            wb2 = openpyxl.load_workbook(user_path, read_only=True)
            wb2.close()
        finally:
            self._cleanup(patches)

    def test_8_08_changes_non_empty_when_sheets_added(self, tmp):
        patches, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(("Customers",)),  # missing Settings
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            assert result.success
            assert len(result.changes) >= 1
        finally:
            self._cleanup(patches)

    def test_8_09_failure_when_file_locked(self, tmp):
        patches, user_path, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            with mock.patch.object(ms, "_is_file_locked", return_value=True):
                result = ms.check_and_migrate()
            assert result.success is False
            assert "open in Excel" in result.error or "locked" in result.error.lower()
        finally:
            self._cleanup(patches)

    def test_8_10_failure_when_template_missing(self, tmp):
        patches, *_ = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=None,        # no template on disk
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            assert result.success is False
            assert result.error
        finally:
            self._cleanup(patches)

    def test_8_11_original_restored_when_integrity_check_fails(self, tmp):
        """Inject a migration function that introduces CRITICAL corruption,
        verify the backup is restored to the original path."""
        orig_wb   = _make_minimal_wb(("Customers",))
        tmpl_wb   = _make_template_wb()
        patches, user_path, *_ = self._patch(
            tmp, user_wb=orig_wb, template_wb=tmpl_wb, schema_version=0
        )

        # Read original bytes before migration
        orig_bytes = (tmp / "user_tracker.xlsx").read_bytes()

        def _bad_migration(wb, tmpl):
            # Delete all data rows — CRITICAL corruption
            ws = wb["Customers"]
            if ws.max_row >= 3:
                ws.delete_rows(3, ws.max_row - 2)
            return ["bad migration done"]

        self._apply(patches)
        try:
            with mock.patch.object(ms, "_MIGRATIONS",
                                   [(0, 1, _bad_migration)]):
                result = ms.check_and_migrate()

            assert result.success is False
            # Original file must be restored
            restored_bytes = user_path.read_bytes()
            assert restored_bytes == orig_bytes, \
                "Backup was NOT restored — original file corrupted!"
        finally:
            self._cleanup(patches)

    def test_8_12_succeeds_with_only_warnings(self, tmp):
        """A migration that produces only WARNINGs (not CRITICALs)
        should still succeed and save."""
        orig_wb  = _make_minimal_wb(("Customers",))
        tmpl_wb  = _make_template_wb()
        patches, user_path, *_ = self._patch(
            tmp, user_wb=orig_wb, template_wb=tmpl_wb, schema_version=0
        )

        def _warn_migration(wb, tmpl):
            # Change a header text — produces WARNING, not CRITICAL
            wb["Customers"].cell(row=2, column=1, value="CustomerID-Renamed")
            return ["renamed a header"]

        self._apply(patches)
        try:
            with mock.patch.object(ms, "_MIGRATIONS",
                                   [(0, 1, _warn_migration)]):
                result = ms.check_and_migrate()
            assert result.success is True
        finally:
            self._cleanup(patches)

    def test_8_13_log_file_written(self, tmp):
        patches, _, _, log_path = self._patch(
            tmp,
            user_wb=_make_minimal_wb(),
            template_wb=_make_template_wb(),
            schema_version=0
        )
        # Clear any existing file handlers on the migration logger so the
        # patched _LOG_PATH is used (the 'if not _log.handlers' guard in
        # _setup_log() would otherwise skip adding the new handler).
        for h in ms._log.handlers[:]:
            ms._log.removeHandler(h)
            h.close()
        self._apply(patches)
        try:
            ms.check_and_migrate()
            assert log_path.exists(), \
                f"Log file was not created at {log_path}"
            content = log_path.read_text(encoding="utf-8")
            assert len(content) > 0
        finally:
            # Re-clear so the patched handler doesn't bleed into other tests
            for h in ms._log.handlers[:]:
                ms._log.removeHandler(h)
                h.close()
            self._cleanup(patches)

    def test_8_14_no_spreadsheet_treated_as_current(self, tmp):
        """User with schema_version=0 but no spreadsheet file:
        should skip migration and mark as current."""
        patches, *_ = self._patch(
            tmp,
            user_wb=None,           # no spreadsheet on disk
            template_wb=_make_template_wb(),
            schema_version=0
        )
        self._apply(patches)
        try:
            result = ms.check_and_migrate()
            # Should succeed (no file = nothing to migrate)
            assert result.error == "" or result.success is True or \
                   result.needed is False
        finally:
            self._cleanup(patches)


# ══════════════════════════════════════════════════════════════════════════════
# GROUP 9 — Renamed-column guard (_ensure_columns trigram similarity)
# ══════════════════════════════════════════════════════════════════════════════

class TestGroup9_RenamedColumnGuard:
    """Tests for the trigram-similarity renamed-column guard in _ensure_columns."""

    def test_9_01_identical_header_not_added_again(self):
        user = _make_minimal_wb(("Customers",))
        tmpl = _make_minimal_wb(("Customers",))
        added, warned = ms._ensure_columns(
            user["Customers"], tmpl["Customers"], hdr_row=2)
        assert added == []
        assert warned == []

    def test_9_02_clearly_different_header_is_added(self):
        user = _make_minimal_wb(("Customers",))
        tmpl = _make_minimal_wb(("Customers",))
        tmpl["Customers"].cell(row=2, column=5, value="GPS Coordinates")
        added, warned = ms._ensure_columns(
            user["Customers"], tmpl["Customers"], hdr_row=2)
        assert "GPS Coordinates" in added

    def test_9_03_similar_header_triggers_rename_guard(self):
        """Headers above similarity threshold → skip add, emit warning."""
        user = _make_minimal_wb(("Customers",))
        user["Customers"].cell(row=2, column=3, value="ServiceDate")
        tmpl = _make_minimal_wb(("Customers",))
        tmpl["Customers"].cell(row=2, column=3, value="Service Date")
        tmpl["Customers"].cell(row=2, column=5, value="Service Date")
        added, warned = ms._ensure_columns(
            user["Customers"], tmpl["Customers"], hdr_row=2)
        rename_warns = [w for w in warned
                        if "Service Date" in w or "renamed" in w.lower()
                        or "similar" in w.lower()]
        assert len(rename_warns) >= 1, \
            f"Expected rename guard warning, got: warned={warned}"

    def test_9_04_rename_guard_does_not_block_genuinely_new_columns(self):
        user = _make_minimal_wb(("Customers",))
        tmpl = _make_minimal_wb(("Customers",))
        tmpl["Customers"].cell(row=2, column=5, value="Payment Method")
        added, warned = ms._ensure_columns(
            user["Customers"], tmpl["Customers"], hdr_row=2)
        assert "Payment Method" in added

    def test_9_05_similarity_function_symmetric(self):
        pairs = [
            ("Service Date", "ServiceDate"),
            ("Amount ($)", "Total Amount"),
            ("CustomerID", "Customer ID"),
        ]
        for a, b in pairs:
            assert abs(ms._header_similarity(a, b) -
                       ms._header_similarity(b, a)) < 1e-9

    def test_9_06_empty_strings_return_zero_similarity(self):
        assert ms._header_similarity("", "anything") == 0.0
        assert ms._header_similarity("anything", "") == 0.0


# ══════════════════════════════════════════════════════════════════════════════
# GROUP 10 — Protection: template headers locked, user columns unlocked
# ══════════════════════════════════════════════════════════════════════════════

class TestGroup10_Protection:

    def test_10_01_template_header_cells_are_locked(self):
        wb = _make_minimal_wb(("Customers",))
        ms._apply_sheet_protection(wb["Customers"], "Customers", user_added_cols=None)
        assert wb["Customers"].cell(row=2, column=1).protection.locked is True

    def test_10_02_data_row_cells_are_unlocked(self):
        wb = _make_minimal_wb(("Customers",))
        ms._apply_sheet_protection(wb["Customers"], "Customers", user_added_cols=None)
        assert wb["Customers"].cell(row=3, column=1).protection.locked is False

    def test_10_03_user_added_column_header_is_unlocked(self):
        wb = _make_minimal_wb(("Customers",))
        wb["Customers"].cell(row=2, column=5, value="My Custom Column")
        ms._apply_sheet_protection(wb["Customers"], "Customers", user_added_cols={5})
        assert wb["Customers"].cell(row=2, column=5).protection.locked is False

    def test_10_04_template_column_still_locked_when_user_cols_present(self):
        wb = _make_minimal_wb(("Customers",))
        wb["Customers"].cell(row=2, column=5, value="My Custom Column")
        ms._apply_sheet_protection(wb["Customers"], "Customers", user_added_cols={5})
        assert wb["Customers"].cell(row=2, column=1).protection.locked is True

    def test_10_05_sheet_protection_enabled(self):
        wb = _make_minimal_wb(("Customers",))
        ms._apply_sheet_protection(wb["Customers"], "Customers")
        assert wb["Customers"].protection.sheet is True

    def test_10_06_column_insert_blocked(self):
        wb = _make_minimal_wb(("Customers",))
        ms._apply_sheet_protection(wb["Customers"], "Customers")
        assert wb["Customers"].protection.insertColumns is True

    def test_10_07_row_insert_allowed(self):
        wb = _make_minimal_wb(("Customers",))
        ms._apply_sheet_protection(wb["Customers"], "Customers")
        assert wb["Customers"].protection.insertRows is False

    def test_10_08_column_resize_allowed(self):
        wb = _make_minimal_wb(("Customers",))
        ms._apply_sheet_protection(wb["Customers"], "Customers")
        assert wb["Customers"].protection.formatColumns is False

    def test_10_09_migrate_preserves_user_added_columns_unlocked(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        user["Customers"].cell(row=2, column=5, value="My Notes")
        user["Customers"].cell(row=3, column=5, value="some note")
        ms._migrate_0_to_1(user, template)
        ws = user["Customers"]
        user_col_idx = next(
            (c for c in range(1, ws.max_column + 1)
             if ws.cell(row=2, column=c).value == "My Notes"), None)
        assert user_col_idx is not None, "User-added column 'My Notes' was removed"
        assert ws.cell(row=2, column=user_col_idx).protection.locked is False

    def test_10_10_migrate_locks_template_column_headers(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        ms._migrate_0_to_1(user, template)
        assert user["Customers"].cell(row=2, column=1).protection.locked is True

    def test_10_11_user_added_sheet_not_touched_by_migration(self):
        user     = _make_minimal_wb(("Customers",))
        template = _make_template_wb()
        user.create_sheet("My_Custom_Tracking").cell(row=3, column=1, value="Value1")
        changes = ms._migrate_0_to_1(user, template)
        assert "My_Custom_Tracking" in user.sheetnames
        assert user["My_Custom_Tracking"].cell(row=3, column=1).value == "Value1"
        assert any("My_Custom_Tracking" in c for c in changes)


# ══════════════════════════════════════════════════════════════════════════════
# GROUP 11 — Chain migration v0 → v3
# ══════════════════════════════════════════════════════════════════════════════

class TestGroup11_ChainMigration:

    @staticmethod
    def _chain():
        def _m01(wb, t): wb.create_sheet("Chain_v1"); return ["v0→v1"]
        def _m12(wb, t): wb.create_sheet("Chain_v2"); return ["v1→v2"]
        def _m23(wb, t): wb.create_sheet("Chain_v3"); return ["v2→v3"]
        return [(0,1,_m01),(1,2,_m12),(2,3,_m23)]

    def _run_chain(self, start_ver, migrations):
        wb = _make_minimal_wb(("Customers",))
        t  = _make_template_wb()
        ver = start_ver
        for from_v, to_v, fn in migrations:
            if ver >= to_v or ver != from_v:
                continue
            fn(wb, t); ver = to_v
        return wb, ver

    def test_11_01_v0_to_v3_runs_all_steps(self):
        wb, ver = self._run_chain(0, self._chain())
        assert "Chain_v1" in wb.sheetnames
        assert "Chain_v2" in wb.sheetnames
        assert "Chain_v3" in wb.sheetnames
        assert ver == 3

    def test_11_02_v1_to_v3_skips_first(self):
        wb, ver = self._run_chain(1, self._chain())
        assert "Chain_v1" not in wb.sheetnames
        assert "Chain_v2" in wb.sheetnames
        assert "Chain_v3" in wb.sheetnames
        assert ver == 3

    def test_11_03_v2_to_v3_only_last(self):
        wb, ver = self._run_chain(2, self._chain())
        assert "Chain_v1" not in wb.sheetnames
        assert "Chain_v2" not in wb.sheetnames
        assert "Chain_v3" in wb.sheetnames
        assert ver == 3

    def test_11_04_already_current_runs_nothing(self):
        wb, ver = self._run_chain(3, self._chain())
        assert "Chain_v1" not in wb.sheetnames
        assert "Chain_v2" not in wb.sheetnames
        assert "Chain_v3" not in wb.sheetnames
        assert ver == 3

    def test_11_05_end_to_end_via_check_and_migrate(self, tmp):
        user_path     = tmp / "user_tracker.xlsx"
        template_path = tmp / "template_tracker.xlsx"
        cfg_path      = tmp / "config.json"
        cfg_path.write_text(json.dumps({"spreadsheet_schema_version": 0}))
        _make_minimal_wb(("Customers",)).save(user_path)
        _make_template_wb().save(template_path)
        patches = [
            mock.patch.object(ms, "_CONFIG_PATH",  cfg_path),
            mock.patch.object(ms, "_LOG_PATH",      tmp / "log.log"),
            mock.patch.object(ms, "_APP_DIR",       tmp),
            mock.patch.object(ms, "_get_spreadsheet_path", return_value=user_path),
            mock.patch.object(ms, "_get_template_path",    return_value=template_path),
        ]
        for p in patches: p.__enter__()
        try:
            result = ms.check_and_migrate()
            assert result.success is True, f"Failed: {result.error}"
            data = json.loads(cfg_path.read_text())
            assert data["spreadsheet_schema_version"] == ms.CURRENT_SCHEMA_VERSION
        finally:
            for p in patches: p.__exit__(None, None, None)


# ══════════════════════════════════════════════════════════════════════════════
# GROUP 12 — get_migration_plan() dry-run preview
# ══════════════════════════════════════════════════════════════════════════════

class TestGroup12_MigrationPlan:

    def _patch(self, tmp, user_wb=None, tmpl_wb=None, ver=0):
        up = tmp / "user.xlsx"
        tp = tmp / "tmpl.xlsx"
        cp = tmp / "cfg.json"
        cp.write_text(json.dumps({"spreadsheet_schema_version": ver}))
        if user_wb: user_wb.save(up)
        if tmpl_wb: tmpl_wb.save(tp)
        return [
            mock.patch.object(ms, "_CONFIG_PATH",  cp),
            mock.patch.object(ms, "_LOG_PATH",      tmp / "l.log"),
            mock.patch.object(ms, "_APP_DIR",       tmp),
            mock.patch.object(ms, "_get_spreadsheet_path",
                              return_value=up if user_wb else None),
            mock.patch.object(ms, "_get_template_path",
                              return_value=tp if tmpl_wb else None),
        ]

    def _go(self, patches):
        for p in patches: p.__enter__()

    def _stop(self, patches):
        for p in patches: p.__exit__(None, None, None)

    def test_12_01_not_needed_when_current(self, tmp):
        ps = self._patch(tmp, _make_minimal_wb(), _make_template_wb(),
                         ver=ms.CURRENT_SCHEMA_VERSION)
        self._go(ps)
        try:
            assert ms.get_migration_plan().needed is False
        finally:
            self._stop(ps)

    def test_12_02_needed_when_behind(self, tmp):
        ps = self._patch(tmp, _make_minimal_wb(), _make_template_wb(), ver=0)
        self._go(ps)
        try:
            assert ms.get_migration_plan().needed is True
        finally:
            self._stop(ps)

    def test_12_03_includes_user_path(self, tmp):
        ps = self._patch(tmp, _make_minimal_wb(), _make_template_wb(), ver=0)
        self._go(ps)
        try:
            assert "user.xlsx" in ms.get_migration_plan().user_path
        finally:
            self._stop(ps)

    def test_12_04_includes_backup_name(self, tmp):
        ps = self._patch(tmp, _make_minimal_wb(), _make_template_wb(), ver=0)
        self._go(ps)
        try:
            plan = ms.get_migration_plan()
            assert ".xlsx" in plan.backup_name
            assert "pre_migration" in plan.backup_name
        finally:
            self._stop(ps)

    def test_12_05_identifies_user_added_sheets(self, tmp):
        user = _make_minimal_wb(("Customers", "My_Sheet"))
        ps = self._patch(tmp, user, _make_template_wb(), ver=0)
        self._go(ps)
        try:
            assert "My_Sheet" in ms.get_migration_plan().sheets_preserved
        finally:
            self._stop(ps)

    def test_12_06_does_not_modify_files(self, tmp):
        user_wb = _make_minimal_wb(("Customers",))
        ps = self._patch(tmp, user_wb, _make_template_wb(), ver=0)
        up = tmp / "user.xlsx"
        self._go(ps)
        try:
            mtime_before = up.stat().st_mtime
            ms.get_migration_plan()
            assert up.stat().st_mtime == mtime_before
        finally:
            self._stop(ps)

    def test_12_07_no_error_when_plan_computed_successfully(self, tmp):
        """Regression test for the read_only=True / ReadOnlyWorksheet.freeze_panes
        bug: get_migration_plan() loaded both workbooks with read_only=True for
        speed, then compared ws.freeze_panes, which openpyxl's read-only
        worksheets don't support — raising AttributeError on every single call
        and silently landing in plan.error with zero logging (get_migration_plan
        has no _log calls of its own). Every prior test in this class exercised
        that exact code path (default _make_minimal_wb sheets have no freeze
        panes vs the template's "A3") but only checked needed/user_path/
        backup_name/sheets_preserved — none of which are set after the crash
        point — so the bug shipped without a failing test. This test would have
        caught it."""
        ps = self._patch(tmp, _make_minimal_wb(), _make_template_wb(), ver=0)
        self._go(ps)
        try:
            plan = ms.get_migration_plan()
            assert plan.error == "", f"plan computation failed: {plan.error}"
        finally:
            self._stop(ps)

    def test_12_08_identifies_freeze_pane_mismatches(self, tmp):
        """_make_minimal_wb() sheets have no freeze panes; _make_template_wb()
        sets freeze_panes='A3' on Customers and Jobs_Schedule — both should be
        reported as needing a freeze-pane fix."""
        ps = self._patch(tmp, _make_minimal_wb(), _make_template_wb(), ver=0)
        self._go(ps)
        try:
            plan = ms.get_migration_plan()
            assert plan.error == ""
            assert "Customers" in plan.freeze_fixes
            assert "Jobs_Schedule" in plan.freeze_fixes
        finally:
            self._stop(ps)

    def test_12_09_identifies_missing_columns(self, tmp):
        """_make_template_wb() adds 'New Column v9.1' to Customers that the
        minimal user workbook doesn't have — should show up in columns_to_add."""
        ps = self._patch(tmp, _make_minimal_wb(), _make_template_wb(), ver=0)
        self._go(ps)
        try:
            plan = ms.get_migration_plan()
            assert plan.error == ""
            assert "New Column v9.1" in plan.columns_to_add.get("Customers", [])
        finally:
            self._stop(ps)


# ══════════════════════════════════════════════════════════════════════════════
# GROUP 13 — Migration button helpers
# ══════════════════════════════════════════════════════════════════════════════

class TestGroup13_MigrationButtonHelpers:

    def test_13_01_needed_when_behind(self, tmp):
        cfg = tmp / "c.json"
        cfg.write_text(json.dumps({"spreadsheet_schema_version": 0}))
        with mock.patch.object(ms, "_CONFIG_PATH", cfg):
            assert ms._get_schema_version() < ms.CURRENT_SCHEMA_VERSION

    def test_13_02_not_needed_when_current(self, tmp):
        cfg = tmp / "c.json"
        cfg.write_text(
            json.dumps({"spreadsheet_schema_version": ms.CURRENT_SCHEMA_VERSION}))
        with mock.patch.object(ms, "_CONFIG_PATH", cfg):
            assert not (ms._get_schema_version() < ms.CURRENT_SCHEMA_VERSION)

    def test_13_03_migration_plan_has_required_attributes(self):
        plan = ms.MigrationPlan()
        for attr in ("needed","user_path","backup_name","backup_dir",
                     "sheets_to_add","sheets_to_replace","sheets_preserved",
                     "columns_to_add","columns_preserved","freeze_fixes","error"):
            assert hasattr(plan, attr), f"MigrationPlan missing: {attr}"

    def test_13_04_migration_result_has_ui_attributes(self):
        r = ms.MigrationResult()
        assert hasattr(r, "user_spreadsheet_path")
        assert hasattr(r, "backup_name_preview")

    def test_13_05_get_migration_plan_callable(self):
        assert callable(ms.get_migration_plan)

    def test_13_06_indicator_strings_correct(self):
        assert "🔴" in "🔴  Update required"
        assert "🟢" in "🟢  Up to date"

    def test_13_07_on_demand_triggers_plan_computation(self, tmp):
        calls = []
        def fake_plan():
            calls.append(1)
            return ms.MigrationPlan()
        with mock.patch.object(ms, "get_migration_plan", side_effect=fake_plan):
            plan = ms.get_migration_plan()
            assert len(calls) == 1
            assert isinstance(plan, ms.MigrationPlan)

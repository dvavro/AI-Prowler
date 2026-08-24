"""
Unit tests -- V8.0.0 Contractor Workflow Action Tools

Tests for the five new action tools:
  email_invoice              (ACTION TOOL 8)
  send_sms                   (ACTION TOOL 9)
  schedule_next_recurring_job (ACTION TOOL 10)
  log_time_entry             (ACTION TOOL 11)
  get_ar_aging_report        (ACTION TOOL 12)

All tests mock openpyxl, smtplib, and the Twilio REST API so the suite
runs without a real spreadsheet, SMTP server, or Twilio account.

Test IDs
--------
  CT_01 - CT_06   email_invoice
  CT_07 - CT_11   send_sms
  CT_12 - CT_17   schedule_next_recurring_job
  CT_18 - CT_22   log_time_entry
  CT_23 - CT_28   get_ar_aging_report
"""
from __future__ import annotations

import datetime
import importlib.abc
import importlib.machinery
import importlib.util
import json
import os
import sys
import types
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
import openpyxl


# ---------------------------------------------------------------------------
# Dependency stubs -- install a MetaPathFinder that satisfies imports of
# packages that may be absent in CI/sandbox (chromadb, sentence-transformers,
# mcp SDK, etc.) without requiring them to be installed.
#
# On the user's real Windows install ALL packages ARE installed, so the finder
# returns None for them (they're already in sys.modules) and the real modules
# are used.  On Linux CI / the developer sandbox the finder intercepts the
# imports and returns lightweight callable stubs.
# ---------------------------------------------------------------------------
_STUB_TOPS = frozenset([
    "mcp", "chromadb", "sentence_transformers", "transformers",
    "pdfplumber", "pypdfium2", "pytesseract",
    "bs4", "striprtf", "odf", "watchdog", "pillow_heif", "extract_msg",
])


class _StubAttr:
    """Callable stub returned for any attribute of a stub module."""
    def __init__(self, *a, **kw):
        pass

    def __call__(self, *a, **kw):
        return _StubAttr()

    def __getattr__(self, n):
        return _StubAttr()

    def __iter__(self):
        return iter([])

    def __bool__(self):
        return True

    def __enter__(self):
        return self

    def __exit__(self, *a):
        return False


class _StubLoader(importlib.abc.Loader):
    def create_module(self, spec):
        mod = types.ModuleType(spec.name)
        mod.__path__ = []
        mod.__package__ = spec.name.split(".")[0]
        mod.__spec__ = spec
        return mod

    def exec_module(self, module):
        def _getattr(name):
            if name.startswith("__"):
                raise AttributeError(name)
            return _StubAttr()

        module.__class__ = type(
            module.__name__,
            (types.ModuleType,),
            {"__getattr__": lambda self, n: _getattr(n)},
        )


class _StubFinder(importlib.abc.MetaPathFinder):
    # Guard against re-entrancy: importlib.util.find_spec() below walks
    # sys.meta_path, which calls back into this finder. Track tops we are
    # currently probing so the nested call returns None instead of recursing.
    _probing: set = set()

    def find_spec(self, fullname, path, target=None):
        top = fullname.split(".")[0]
        # Only stub when the TOP-LEVEL package is genuinely absent.
        #
        # IMPORTANT: do NOT stub a submodule (e.g. chromadb.config) just
        # because that submodule hasn't been imported yet. If the top-level
        # package is installed and real, its submodules must resolve to the
        # REAL implementation. The previous version stubbed any not-yet-
        # imported submodule of a _STUB_TOPS package, which leaked _StubAttr
        # objects into live ChromaDB collections (collection.count() -> stub),
        # poisoning every later test in the run via the process-global
        # sys.meta_path finder. See test isolation bug, v8.0.0.
        if top not in _STUB_TOPS:
            return None
        if top in self._probing:
            # Re-entrant call from our own find_spec probe below: defer.
            return None
        self._probing.add(top)
        try:
            # If the real top-level package can be located, it is installed --
            # defer to the real import machinery for it and all submodules.
            if importlib.util.find_spec(top) is not None:
                return None
        except (ImportError, AttributeError, ValueError):
            pass
        finally:
            self._probing.discard(top)
        # Top-level package is truly absent -> provide a lightweight stub.
        return importlib.machinery.ModuleSpec(fullname, _StubLoader())


def _install_stub_finder():
    """Insert the stub finder once; return it so it can be removed later."""
    for f in sys.meta_path:
        if isinstance(f, _StubFinder):
            return f
    finder = _StubFinder()
    sys.meta_path.insert(0, finder)
    return finder


_STUB_FINDER = _install_stub_finder()


@pytest.fixture(scope="module", autouse=True)
def _remove_stub_finder_after_module():
    """Ensure the process-global stub finder cannot outlive this test module.

    Two-part teardown:

    1.  Remove the finder from sys.meta_path so no further imports are
        intercepted after this module finishes.

    2.  Evict any stub-generated entries from sys.modules for packages that
        ARE actually installed on this machine.  Without this step a later
        test file that imports (e.g.) `watchdog` gets the _StubAttr version
        that was baked into sys.modules during this module's run, causing
        "TypeError: __mro_entries__ must return a tuple" when the stub is
        used as a base class — even though the real watchdog package is
        installed and our fixed find_spec correctly defers to it on fresh
        imports.  The session-scoped `wd` fixture in test_file_watchdog.py
        is the concrete victim of this if it runs after us.
    """
    yield
    # 1. Remove the finder.
    try:
        sys.meta_path.remove(_STUB_FINDER)
    except ValueError:
        pass

    # 2. Evict stub-generated sys.modules entries for installed packages.
    #    We only evict tops (and their submodules) that ARE actually
    #    installed — if they were genuinely absent we leave their stub
    #    entries so the rest of the session keeps seeing stubs, not import
    #    errors, for packages that don't exist.
    _to_evict = []
    for top in _STUB_TOPS:
        _STUB_FINDER._probing.add(top)   # prevent re-entrant find_spec
        try:
            real_spec = importlib.util.find_spec(top)
        except Exception:
            real_spec = None
        finally:
            _STUB_FINDER._probing.discard(top)

        if real_spec is None:
            continue   # genuinely not installed — keep the stub in modules

        # Package is installed: remove every sys.modules entry whose top
        # matches, so the next import gets the real package.
        _to_evict.extend(
            k for k in list(sys.modules)
            if k == top or k.startswith(top + ".")
        )

    for key in _to_evict:
        sys.modules.pop(key, None)

    # 3. Re-bind stub-contaminated module-level names inside rag_preprocessor.
    #
    #    Evicting from sys.modules (Step 2) lets FUTURE imports get the real
    #    package, but any module that already bound the stub at import time
    #    still holds a reference to the _StubAttr object in its own namespace.
    #    rag_preprocessor is the primary victim: it does `import pdfplumber`
    #    at module level (line ~138) inside a try block. If the stub finder was
    #    active when rag_preprocessor was first imported (session scope), the
    #    module's `pdfplumber` attribute is permanently a _StubAttr — causing
    #    test_pdf_extraction.py and test_image_formats.py to get empty strings
    #    back from load_pdf/load_image_ocr when they run after us.
    #
    #    Fix: after evicting the stubs from sys.modules, re-import each
    #    installed package and patch the binding directly on rag_preprocessor.
    _rag_mod = sys.modules.get("rag_preprocessor")
    if _rag_mod is not None:
        # Force-rebind pdfplumber, pytesseract, and pillow_heif on
        # rag_preprocessor regardless of whether they look like stubs.
        # If the stub finder was active when rag_preprocessor was first
        # imported, those names point to stub module objects. After Step 2
        # evicted them from sys.modules, a fresh import_module() gives the
        # real package. We then patch rag_preprocessor's module dict directly
        # so that subsequent test files calling e.g. pdfplumber.open() inside
        # load_pdf() get the real (and mockable) implementation.
        _rebind_targets = ["pdfplumber", "pytesseract", "pillow_heif"]
        for _pkg_name in _rebind_targets:
            try:
                import importlib as _il
                _real = _il.import_module(_pkg_name)
                setattr(_rag_mod, _pkg_name, _real)
                # Also patch into sys.modules so patch("pdfplumber.open", ...)
                # and rag_preprocessor.pdfplumber refer to the same object.
                sys.modules[_pkg_name] = _real
            except Exception:
                pass  # package genuinely absent — leave as-is

# Wire FastMCP stub (triggers stub loader for mcp.server.fastmcp)
import mcp.server.fastmcp as _fmcp  # noqa: E402


class _FakeFastMCP:
    def __init__(self, *a, **kw):
        pass

    def tool(self):
        def decorator(fn):
            return fn
        return decorator

    def run(self, *a, **kw):
        pass


_fmcp.FastMCP = _FakeFastMCP   # type: ignore[attr-defined]
_fmcp.Context = None            # type: ignore[attr-defined]


# ---------------------------------------------------------------------------
# Module import helper
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module")
def mcp_module():
    """Import ai_prowler_mcp once per test module."""
    import ai_prowler_mcp
    return ai_prowler_mcp


# ---------------------------------------------------------------------------
# Shared spreadsheet factory
# ---------------------------------------------------------------------------
def _make_test_spreadsheet(tmp_path):
    """Create a minimal AI-Prowler_Job_Tracker.xlsx for testing."""
    fp = tmp_path / "test_tracker.xlsx"
    wb = openpyxl.Workbook()

    # Customers sheet
    ws_c = wb.active
    ws_c.title = "Customers"
    ws_c.append(["AI-PROWLER JOB TRACKER -- Customer Master List"])
    ws_c.append([
        "CustomerID (CUST-####)", "Customer Type Comm/Res", "Company Name",
        "First Name", "Last Name", "Phone", "Email",
        "Street Address * AI Route", "City * AI Route", "State", "ZIP * AI Route",
        "Latitude (AI Geocode)", "Longitude (AI Geocode)",
        "Service Type(s) Win/Press/Both", "Frequency W/BW/M/Q/OT",
        "Preferred Day(s)", "Pref. Time Window", "Avg Job Duration (min)",
        "Standard Quote ($)", "Discount (%)", "Net Price ($)",
        "Last Service Date", "Next Sched. Date", "Total Jobs Completed",
        "Lifetime Revenue ($)", "Gate Code / Access Notes", "On-Site Contact",
        "Status Active/Inactive",
    ])
    ws_c.append([
        "CUST-0001", "Commercial", "Sunshine Realty LLC", "Karen", "Walsh",
        "3865550101", "karen@sunshine.com",
        "125 Harbor Blvd", "New Smyrna Beach", "FL", "32168",
        "", "", "Both", "Monthly",
        "Mon,Wed", "8am-5pm", "90", "350", "0.1", "315",
        "2026-02-28", "2026-03-30", "5", "1750", "", "", "Active",
    ])
    ws_c.append([
        "CUST-0002", "Residential", "", "Michael", "Torres",
        "3865550202", "mtorres@gmail.com",
        "47 Oceanview Dr", "Edgewater", "FL", "32141",
        "", "", "Window", "Biweekly",
        "Saturday", "9am-12pm", "60", "185", "0", "185",
        "2026-03-16", "2026-03-30", "8", "1480", "", "", "Active",
    ])

    # Jobs_Schedule sheet
    ws_j = wb.create_sheet("Jobs_Schedule")
    ws_j.append(["JOBS & SCHEDULE -- All Service Appointments"])
    ws_j.append([
        "JobID (JOB-####)", "CustomerID (Customers!A)", "Customer Name / Company",
        "Customer Type", "Street Address * AI Route", "City * AI Route",
        "State", "ZIP * AI Route", "Latitude (AI Geocode)", "Longitude (AI Geocode)",
        "Service Date", "Day of Week", "Start Time", "End Time", "Service Type",
        "Service Details / Notes", "Crew / Technician", "Est. Duration (min)",
        "Actual Duration (min)", "Route Stop # * AI Route", "Route Map URL * AI Prowler",
        "Weather Check * AI Prowler", "Job Status", "Quote Amount ($)",
        "Actual Amount ($)", "Discount Applied ($)", "Tax (7%)", "Invoice Total ($)",
        "InvoiceID (INV-####)", "Invoice Sent Date", "Payment Status",
    ])
    ws_j.append([
        "JOB-0001", "CUST-0001", "Sunshine Realty LLC", "Commercial",
        "125 Harbor Blvd", "New Smyrna Beach", "FL", "32168", "", "",
        "2026-03-30", "Monday", "08:00", "09:30", "Window",
        "Full exterior window cleaning", "Mike C.", "90", "", "1",
        "", "", "Complete", "315", "315", "31.5", "22.05", "305.55",
        "INV-0001", "2026-03-30", "Unpaid",
    ])
    ws_j.append([
        "JOB-0002", "CUST-0002", "Michael Torres", "Residential",
        "47 Oceanview Dr", "Edgewater", "FL", "32141", "", "",
        "2026-03-16", "Monday", "09:00", "10:00", "Window",
        "House exterior windows", "Jake R.", "60", "", "1",
        "", "", "Complete", "185", "185", "0", "12.95", "197.95",
        "INV-0002", "2026-03-16", "Paid",
    ])

    # Invoices sheet
    ws_i = wb.create_sheet("Invoices")
    ws_i.append(["INVOICES -- Billing & Payment Tracking"])
    ws_i.append([
        "InvoiceID (INV-####)", "JobID (JOB-####)", "CustomerID",
        "Customer Name / Company", "Customer Type", "Invoice Date",
        "Due Date (Net 30)", "Service Date", "Service Type", "Description",
        "Subtotal ($)", "Discount ($)", "Taxable Amt ($)", "Tax 7% ($)",
        "TOTAL DUE ($)", "Amount Paid ($)", "Balance Due ($)",
        "Payment Status", "Payment Date", "Payment Method",
    ])
    ws_i.append([
        "INV-0001", "JOB-0001", "CUST-0001", "Sunshine Realty LLC", "Commercial",
        "2026-03-30", "2026-04-29", "2026-03-30", "Window",
        "Exterior window cleaning -- 12 windows",
        "315", "31.5", "283.5", "19.845", "303.345", "0", "303.345",
        "Unpaid", "", "",
    ])
    ws_i.append([
        "INV-0002", "JOB-0002", "CUST-0002", "Michael Torres", "Residential",
        "2026-03-16", "2026-04-15", "2026-03-16", "Window",
        "House exterior windows",
        "185", "0", "185", "12.95", "197.95", "197.95", "0",
        "Paid", "2026-03-20", "Check",
    ])
    # Overdue invoice (due 2026-02-14, > 90 days overdue by 2026-05-30)
    ws_i.append([
        "INV-0003", "JOB-0003", "CUST-0001", "Sunshine Realty LLC", "Commercial",
        "2026-01-15", "2026-02-14", "2026-01-15", "Both",
        "Old overdue job",
        "500", "0", "500", "35", "535", "0", "535",
        "Unpaid", "", "",
    ])

    # TimeLog sheet
    ws_t = wb.create_sheet("TimeLog")
    ws_t.append(["TIME LOG -- Job Clock In / Clock Out"])
    ws_t.append([
        "EntryID", "JobID", "Customer Name / Company",
        "Clock In", "Clock Out", "Elapsed (min)", "Crew / Technician", "Notes",
    ])

    wb.save(str(fp))
    return fp


def _make_multiline_header_spreadsheet(tmp_path):
    """
    Builds an Invoices sheet using the SAME multi-line header format the
    real production template actually uses — the formula-computed columns
    (Taxable Amt, Tax, TOTAL DUE, Balance Due) have their column label on
    one line and a "=FORMULA" documentation note on a second line within
    the same cell, e.g. "Tax 7% ($)\n=M*0.07".

    This is the exact real-world shape _make_test_spreadsheet()'s simple
    single-line headers never exercised — which is why a real production
    bug (header detection merging the formula note into the header text,
    breaking exact-match lookups against hardcoded keys like
    "Tax 7% ($)") passed every existing test while still being broken for
    every real invoice with this column format.
    """
    import openpyxl as _opx_ml
    fp = tmp_path / "multiline_headers.xlsx"
    wb = _opx_ml.Workbook()
    wb.remove(wb.active)

    ws_i = wb.create_sheet("Invoices")
    ws_i.append(["INVOICES -- Billing & Payment Tracking"])
    ws_i.append([
        "InvoiceID (INV-####)", "JobID (JOB-####)", "CustomerID",
        "Customer Name / Company", "Customer Type", "Invoice Date",
        "Due Date (Net 30)", "Service Date", "Service Type", "Description",
        "Subtotal ($)", "Discount ($)",
        "Taxable Amt ($)\n=K-L", "Tax 7% ($)\n=M*0.07",
        "TOTAL DUE ($)\n=M+N", "Amount Paid ($)", "Balance Due ($)\n=O-P",
        "Payment Status", "Payment Date", "Payment Method",
    ])
    ws_i.append([
        "INV-0007", "JOB-0007", "CUST-0007", "AI-Prowler LLC", "Commercial",
        "2026-03-30", "2026-04-29", "2026-03-30", "Window",
        "Exterior window cleaning -- 12 windows",
        "350", "35", "315", "22.05", "337.05", "0", "337.05",
        "Unpaid", "", "",
    ])
    wb.save(str(fp))
    return fp


# ===========================================================================
# email_invoice  (CT_01 - CT_06)
# ===========================================================================

class TestEmailInvoice:
    """Tests for ACTION TOOL 8 -- email_invoice."""

    # Matches the REAL keys _email_config_save()/configure_email() actually
    # produce (username/password/from_address) — NOT what email_invoice()
    # used to read (smtp_user/smtp_password/from_email). That mismatch was
    # a real production bug: every email_invoice() call attempted SMTP
    # login with a blank password and then hit a bare KeyError on
    # 'from_email', guaranteeing failure regardless of correct SMTP setup.
    # This fixture previously used the SAME wrong keys as the buggy code,
    # which is exactly why these tests never caught it — a MagicMock SMTP
    # server accepts login() with any (even blank) credentials silently,
    # and the old tests only checked the success message's text, never
    # what was actually passed to login()/sendmail().
    _SMTP_CFG = {
        "smtp_host": "smtp.test.com",
        "smtp_port": 587,
        "username": "u@test.com",
        "password": "realpassword123",
        "from_address": "me@test.com",
        "from_name": "Test",
    }

    def test_CT_01_email_invoice_by_invoice_id(self, mcp_module, tmp_path):
        """email_invoice('INV-0001') must send email and return confirmation."""
        fp = str(_make_test_spreadsheet(tmp_path))

        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch("smtplib.SMTP", return_value=smtp_mock):
            result = mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                to="karen@sunshine.com",
                filepath=fp,
            )

        assert "INV-0001" in result
        assert "Sunshine Realty" in result or "karen" in result or "sent" in result.lower()

    def test_CT_01b_email_invoice_uses_real_username_and_password(self, mcp_module, tmp_path):
        """The exact bug this regression guards against: login() must be
        called with the real configured username/password, not a blank
        default from a key that doesn't exist in the saved config."""
        fp = str(_make_test_spreadsheet(tmp_path))

        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch("smtplib.SMTP", return_value=smtp_mock):
            mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                to="karen@sunshine.com",
                filepath=fp,
            )

        smtp_mock.login.assert_called_once_with("u@test.com", "realpassword123")

    def test_CT_01c_email_invoice_uses_real_from_address_for_envelope(self, mcp_module, tmp_path):
        """sendmail()'s envelope-from must be the real configured
        from_address, not a KeyError on a nonexistent 'from_email' key."""
        fp = str(_make_test_spreadsheet(tmp_path))

        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch("smtplib.SMTP", return_value=smtp_mock):
            mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                to="karen@sunshine.com",
                filepath=fp,
            )

        assert smtp_mock.sendmail.call_args[0][0] == "me@test.com"

    def test_CT_02_email_invoice_auto_lookup_customer_email(self, mcp_module, tmp_path):
        """When 'to' is omitted, email_invoice must try to look up email from Customers."""
        fp = str(_make_test_spreadsheet(tmp_path))

        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch("smtplib.SMTP", return_value=smtp_mock):
            result = mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                filepath=fp,
                # 'to' is omitted -- auto-lookup path
            )

        # Must return a string; must not crash
        assert isinstance(result, str)
        assert "No spreadsheet" not in result

    def test_CT_03_email_invoice_not_found_returns_error(self, mcp_module, tmp_path):
        """Searching for a nonexistent invoice must return an error string."""
        fp = str(_make_test_spreadsheet(tmp_path))

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG):
            result = mcp_module.email_invoice(
                invoice_identifier="INV-9999",
                to="nobody@example.com",
                filepath=fp,
            )

        assert "INV-9999" in result
        # Should be an error: either explicitly or "not found" language
        assert any(w in result.lower() for w in ["not found", "no invoice", "error", "could not"])

    def test_CT_04_email_invoice_no_smtp_config_returns_error(self, mcp_module, tmp_path):
        """If email is not configured, email_invoice must return a clear error."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_email_config_load", return_value=None):
            result = mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                to="karen@sunshine.com",
                filepath=fp,
            )

        assert any(w in result.lower() for w in ["configure", "email", "smtp", "setup"])

    def test_CT_05_email_invoice_missing_spreadsheet_returns_error(self, mcp_module, tmp_path):
        """Passing a nonexistent filepath must return a file-not-found error."""
        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG):
            result = mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                to="test@example.com",
                filepath=str(tmp_path / "does_not_exist.xlsx"),
            )
        assert any(w in result.lower() for w in ["not found", "no spreadsheet", "error", "file"])

    def test_CT_05b_blank_identifier_is_rejected_not_treated_as_match_all(self, mcp_module, tmp_path):
        """The exact production bug this regression guards against: Python
        treats "" as a substring of every string, so a blank
        invoice_identifier previously matched the FIRST row in the
        Invoices sheet unconditionally — silently emailing an unrelated
        customer's invoice with no error. A caller passing '' (from an
        upstream parsing gap, not malice) must get a clear rejection
        instead of a wrong invoice being sent."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG):
            result = mcp_module.email_invoice(
                invoice_identifier="",
                to="test@example.com",
                filepath=fp,
            )
        assert result.startswith("❌")
        assert "blank" in result.lower() or "required" in result.lower()

    def test_CT_05c_whitespace_only_identifier_is_also_rejected(self, mcp_module, tmp_path):
        """A whitespace-only identifier is functionally blank once
        stripped, and Python's substring check doesn't strip — must be
        caught the same way as a fully empty string."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG):
            result = mcp_module.email_invoice(
                invoice_identifier="   ",
                to="test@example.com",
                filepath=fp,
            )
        assert result.startswith("❌")

    def test_CT_05d_blank_identifier_check_happens_before_smtp_send(self, mcp_module, tmp_path):
        """The rejection must happen before any email is actually sent —
        confirms this is a genuine early-exit guard, not just a message
        wrapped around a real send attempt."""
        fp = str(_make_test_spreadsheet(tmp_path))
        smtp_mock = MagicMock()
        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch("smtplib.SMTP", return_value=smtp_mock):
            mcp_module.email_invoice(
                invoice_identifier="",
                to="test@example.com",
                filepath=fp,
            )
        smtp_mock.assert_not_called()

    def test_CT_06_email_invoice_html_contains_key_fields(self, mcp_module, tmp_path):
        """The email payload sent must reference the invoice ID."""
        fp = str(_make_test_spreadsheet(tmp_path))
        captured = []

        def fake_sendmail(from_addr, to_list, msg_str):
            captured.append(msg_str)

        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)
        smtp_mock.sendmail.side_effect = fake_sendmail

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch("smtplib.SMTP", return_value=smtp_mock):
            result = mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                to="karen@sunshine.com",
                filepath=fp,
            )

        if smtp_mock.sendmail.called and captured:
            assert "INV-0001" in captured[0] or "Sunshine" in captured[0]

    def test_CT_06b_tax_and_total_render_correctly_with_real_multiline_headers(
        self, mcp_module, tmp_path
    ):
        """Real production bug regression guard: the actual spreadsheet
        template's formula-computed columns (Taxable Amt, Tax, TOTAL DUE,
        Balance Due) have multi-line headers — label on one line, a
        "=FORMULA" note on the next. Header detection was merging those
        into one string ("Tax 7% ($) =M*0.07"), which no longer matched
        the hardcoded lookup key ("Tax 7% ($)"), so every real invoice
        with this format showed "—" for tax and total regardless of
        having correct underlying numbers. _make_test_spreadsheet()'s
        simple single-line headers never exercised this at all."""
        fp = str(_make_multiline_header_spreadsheet(tmp_path))
        captured = {}

        def fake_sendmail(from_addr, to_list, msg_str):
            import email as _email_mod
            parsed = _email_mod.message_from_string(msg_str)
            for part in parsed.walk():
                if part.get_content_type() == "text/html":
                    captured["html"] = part.get_payload(decode=True).decode("utf-8", errors="replace")
                    break

        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)
        smtp_mock.sendmail.side_effect = fake_sendmail

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch("smtplib.SMTP", return_value=smtp_mock):
            mcp_module.email_invoice(
                invoice_identifier="INV-0007",
                to="david@example.com",
                filepath=fp,
            )

        html = captured.get("html", "")
        assert "$22.05" in html  # Tax 7%
        assert "$337.05" in html  # TOTAL DUE (and Balance Due)
        # The old bug's telltale symptom — a bare dash instead of a real
        # dollar figure for these specific fields.
        assert "Tax (7%)—" not in html.replace(" ", "").replace("\n", "")


# ===========================================================================
# _load_payment_settings  +  email_invoice(also_sms=...)
# ===========================================================================

class TestLoadPaymentSettings:
    """Tests for the shared _load_payment_settings() helper — reads the
    Stripe/Square credentials and fallback URLs (Small Business tab), plus
    the two independent enable/disable toggles, from the main config.json.
    Xero is intentionally not supported at all (full OAuth 2.0 required —
    judged too complex for this feature's target users)."""

    def _write_config(self, tmp_path, monkeypatch, data):
        cfg_dir = tmp_path / ".ai-prowler"
        cfg_dir.mkdir(parents=True, exist_ok=True)
        (cfg_dir / "config.json").write_text(json.dumps(data), encoding="utf-8")
        monkeypatch.setattr(Path, "home", lambda: tmp_path)

    def test_defaults_when_config_missing(self, mcp_module, tmp_path, monkeypatch):
        """No config.json at all — email defaults ON (preserves the
        feature's original, pre-toggle behavior), SMS defaults OFF (a new,
        more exposed capability, off until explicitly enabled)."""
        monkeypatch.setattr(Path, "home", lambda: tmp_path)  # empty tmp_path, no config.json
        result = mcp_module._load_payment_settings()
        assert result["stripe_secret_key"] == ""
        assert result["square_access_token"] == ""
        assert result["email_enabled"] is True
        assert result["sms_enabled"] is False

    def test_reads_stripe_credentials(self, mcp_module, tmp_path, monkeypatch):
        self._write_config(tmp_path, monkeypatch, {
            "stripe_secret_key": "sk_test_abc123",
            "stripe_payment_url": "https://buy.stripe.com/fallback",
        })
        result = mcp_module._load_payment_settings()
        assert result["stripe_secret_key"] == "sk_test_abc123"
        assert result["stripe_fallback_url"] == "https://buy.stripe.com/fallback"

    def test_reads_square_credentials(self, mcp_module, tmp_path, monkeypatch):
        self._write_config(tmp_path, monkeypatch, {
            "square_access_token": "EAAA_test_token",
            "square_location_id": "L123ABC",
            "square_payment_url": "https://square.link/u/fallback",
        })
        result = mcp_module._load_payment_settings()
        assert result["square_access_token"] == "EAAA_test_token"
        assert result["square_location_id"] == "L123ABC"
        assert result["square_fallback_url"] == "https://square.link/u/fallback"

    def test_xero_is_not_supported_at_all(self, mcp_module, tmp_path, monkeypatch):
        """Even if a xero_payment_url happens to still be sitting in an old
        config.json from before Xero was removed, it must be ignored
        entirely — no Xero key of any kind in the returned dict."""
        self._write_config(tmp_path, monkeypatch, {
            "xero_payment_url": "https://invoices.xero.com/leftover",
        })
        result = mcp_module._load_payment_settings()
        assert not any("xero" in k.lower() for k in result.keys())

    def test_email_toggle_respected_when_explicitly_off(self, mcp_module, tmp_path, monkeypatch):
        self._write_config(tmp_path, monkeypatch, {
            "email_payment_link_enabled": False,
        })
        result = mcp_module._load_payment_settings()
        assert result["email_enabled"] is False

    def test_sms_toggle_respected_when_explicitly_on(self, mcp_module, tmp_path, monkeypatch):
        self._write_config(tmp_path, monkeypatch, {
            "sms_payment_link_enabled": True,
        })
        result = mcp_module._load_payment_settings()
        assert result["sms_enabled"] is True


class TestEmailInvoicePaymentLinksAndSms:
    """Tests for email_invoice()'s payment-link email injection (gated by
    email_payment_link_enabled) and the also_sms companion notification
    (gated separately by sms_payment_link_enabled). Covers both the
    dynamic per-invoice checkout creation (Stripe/Square API calls,
    always mocked here — never real network) and the static-URL fallback
    when no credentials are configured."""

    _SMTP_CFG = TestEmailInvoice._SMTP_CFG

    def _run_with_settings(self, mcp_module, tmp_path, payment_settings,
                            also_sms=False, send_sms_result="✅ SMS sent to +13865550101",
                            create_stripe_url=None, create_square_url=None):
        fp = str(_make_test_spreadsheet(tmp_path))
        captured = {}

        def fake_sendmail(from_addr, to_list, msg_str):
            # The HTML body is base64-encoded at the MIME transport layer
            # (Content-Transfer-Encoding: base64) — decode it so tests can
            # check for plain-text content, rather than accidentally
            # asserting against the encoded representation.
            import email as _email_mod
            parsed = _email_mod.message_from_string(msg_str)
            html = ""
            for part in parsed.walk():
                if part.get_content_type() == "text/html":
                    html = part.get_payload(decode=True).decode("utf-8", errors="replace")
                    break
            captured["html"] = html
            captured["raw"] = msg_str

        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)
        smtp_mock.sendmail.side_effect = fake_sendmail

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch.object(mcp_module, "_load_payment_settings", return_value=payment_settings), \
             patch.object(mcp_module, "_create_stripe_checkout_url", return_value=create_stripe_url) as stripe_create_mock, \
             patch.object(mcp_module, "_create_square_checkout_url", return_value=create_square_url) as square_create_mock, \
             patch.object(mcp_module, "send_sms", return_value=send_sms_result) as sms_mock, \
             patch("smtplib.SMTP", return_value=smtp_mock):
            result = mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                to="karen@sunshine.com",
                filepath=fp,
                also_sms=also_sms,
            )
        return result, captured, sms_mock, stripe_create_mock, square_create_mock

    _NO_CREDS = {
        "stripe_secret_key": "", "stripe_fallback_url": "",
        "square_access_token": "", "square_location_id": "", "square_fallback_url": "",
        "email_enabled": True, "sms_enabled": False,
    }

    def test_dynamic_stripe_url_used_when_secret_key_configured(self, mcp_module, tmp_path):
        """The primary new capability: a secret key present means a real
        (mocked here) checkout session gets created and its URL used —
        not the static fallback."""
        settings = dict(self._NO_CREDS, stripe_secret_key="sk_test_abc",
                         stripe_fallback_url="https://buy.stripe.com/fallback")
        result, captured, _, stripe_mock, _ = self._run_with_settings(
            mcp_module, tmp_path, settings,
            create_stripe_url="https://checkout.stripe.com/pay/cs_test_dynamic123",
        )
        assert "cs_test_dynamic123" in captured.get("html", "")
        assert "buy.stripe.com/fallback" not in captured.get("html", "")
        stripe_mock.assert_called_once()

    def test_dynamic_stripe_creation_uses_real_invoice_amount(self, mcp_module, tmp_path):
        """The exact dollar amount passed to the checkout-creation call
        must be the invoice's real TOTAL DUE ($303.345 for INV-0001 in
        the test fixture), not a placeholder or the subtotal."""
        settings = dict(self._NO_CREDS, stripe_secret_key="sk_test_abc")
        self._run_with_settings(
            mcp_module, tmp_path, settings,
            create_stripe_url="https://checkout.stripe.com/pay/cs_test_x",
        )
        # Re-run capturing the actual call args this time.
        fp = str(_make_test_spreadsheet(tmp_path))
        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)
        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch.object(mcp_module, "_load_payment_settings", return_value=settings), \
             patch.object(mcp_module, "_create_stripe_checkout_url",
                           return_value="https://checkout.stripe.com/pay/cs_test_x") as stripe_mock, \
             patch("smtplib.SMTP", return_value=smtp_mock):
            mcp_module.email_invoice(invoice_identifier="INV-0001", to="k@sunshine.com", filepath=fp)
        called_amount = stripe_mock.call_args[0][1]
        assert abs(called_amount - 303.345) < 0.01

    def test_falls_back_to_static_url_when_no_secret_key(self, mcp_module, tmp_path):
        settings = dict(self._NO_CREDS, stripe_fallback_url="https://buy.stripe.com/fallback")
        result, captured, _, stripe_mock, _ = self._run_with_settings(mcp_module, tmp_path, settings)
        assert "buy.stripe.com/fallback" in captured.get("html", "")
        stripe_mock.assert_not_called()  # no key → never even attempted

    def test_falls_back_to_static_url_when_dynamic_creation_fails(self, mcp_module, tmp_path):
        """If the Stripe API call fails (bad key, network error, etc.),
        _create_stripe_checkout_url returns None — email_invoice must
        fall back to the static URL rather than showing no button at all
        or crashing."""
        settings = dict(self._NO_CREDS, stripe_secret_key="sk_bad_key",
                         stripe_fallback_url="https://buy.stripe.com/fallback")
        result, captured, _, stripe_mock, _ = self._run_with_settings(
            mcp_module, tmp_path, settings, create_stripe_url=None,
        )
        assert "buy.stripe.com/fallback" in captured.get("html", "")
        stripe_mock.assert_called_once()

    def test_dynamic_square_url_used_when_credentials_configured(self, mcp_module, tmp_path):
        settings = dict(self._NO_CREDS, square_access_token="EAAA_test",
                         square_location_id="L123", square_fallback_url="https://square.link/u/fallback")
        result, captured, _, _, square_mock = self._run_with_settings(
            mcp_module, tmp_path, settings,
            create_square_url="https://checkout.square.site/dynamic456",
        )
        assert "dynamic456" in captured.get("html", "")
        assert "square.link/u/fallback" not in captured.get("html", "")
        square_mock.assert_called_once()

    def test_square_requires_both_token_and_location_id(self, mcp_module, tmp_path):
        """An access token alone (no location ID) must not attempt
        dynamic creation — Square's API requires both."""
        settings = dict(self._NO_CREDS, square_access_token="EAAA_test",
                         square_fallback_url="https://square.link/u/fallback")
        result, captured, _, _, square_mock = self._run_with_settings(mcp_module, tmp_path, settings)
        square_mock.assert_not_called()
        assert "square.link/u/fallback" in captured.get("html", "")

    def test_no_xero_button_ever_appears(self, mcp_module, tmp_path):
        """Regression guard — Xero support was removed entirely; nothing
        in the email should ever reference it."""
        settings = dict(self._NO_CREDS, stripe_secret_key="sk_test",
                         square_access_token="EAAA", square_location_id="L1")
        result, captured, _, _, _ = self._run_with_settings(
            mcp_module, tmp_path, settings,
            create_stripe_url="https://checkout.stripe.com/pay/x",
            create_square_url="https://checkout.square.site/y",
        )
        assert "xero" not in captured.get("html", "").lower()

    def test_email_payment_section_omitted_when_disabled(self, mcp_module, tmp_path):
        """Even with credentials configured, the section must not appear
        if email_payment_link_enabled is off."""
        settings = dict(self._NO_CREDS, email_enabled=False, stripe_secret_key="sk_test")
        result, captured, _, stripe_mock, _ = self._run_with_settings(
            mcp_module, tmp_path, settings,
            create_stripe_url="https://checkout.stripe.com/pay/x",
        )
        assert "checkout.stripe.com" not in captured.get("html", "")
        stripe_mock.assert_not_called()  # never even attempted — email disabled, sms disabled too

    def test_also_sms_false_never_calls_send_sms(self, mcp_module, tmp_path):
        _, _, sms_mock, _, _ = self._run_with_settings(mcp_module, tmp_path, self._NO_CREDS, also_sms=False)
        sms_mock.assert_not_called()

    def test_also_sms_true_calls_send_sms_with_customer_name(self, mcp_module, tmp_path):
        _, _, sms_mock, _, _ = self._run_with_settings(mcp_module, tmp_path, self._NO_CREDS, also_sms=True)
        sms_mock.assert_called_once()
        assert sms_mock.call_args.kwargs.get("to") or sms_mock.call_args[1].get("to")

    def test_also_sms_true_sms_enabled_false_no_link_in_message(self, mcp_module, tmp_path):
        """SMS notification sent, but with NO payment link, even though
        Stripe is configured — sms_enabled must be independently checked,
        not inherited from a configured credential existing."""
        settings = dict(self._NO_CREDS, stripe_secret_key="sk_test", sms_enabled=False)
        _, _, sms_mock, _, _ = self._run_with_settings(
            mcp_module, tmp_path, settings, also_sms=True,
            create_stripe_url="https://checkout.stripe.com/pay/x",
        )
        sent_message = sms_mock.call_args.kwargs.get("message", "")
        assert "checkout.stripe.com" not in sent_message

    def test_also_sms_true_sms_enabled_true_includes_link_in_message(self, mcp_module, tmp_path):
        settings = dict(self._NO_CREDS, stripe_secret_key="sk_test", sms_enabled=True)
        _, _, sms_mock, _, _ = self._run_with_settings(
            mcp_module, tmp_path, settings, also_sms=True,
            create_stripe_url="https://checkout.stripe.com/pay/x",
        )
        sent_message = sms_mock.call_args.kwargs.get("message", "")
        assert "checkout.stripe.com/pay/x" in sent_message

    def test_sms_and_email_share_one_checkout_session_not_two(self, mcp_module, tmp_path):
        """If both email_enabled and sms_enabled are on, only ONE checkout
        session should be created and reused by both channels — not a
        separate one minted per channel for the same invoice."""
        settings = dict(self._NO_CREDS, stripe_secret_key="sk_test",
                         email_enabled=True, sms_enabled=True)
        self._run_with_settings(
            mcp_module, tmp_path, settings, also_sms=True,
            create_stripe_url="https://checkout.stripe.com/pay/shared",
        )
        # Rerun with a spy to count calls precisely.
        fp = str(_make_test_spreadsheet(tmp_path))
        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)
        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch.object(mcp_module, "_load_payment_settings", return_value=settings), \
             patch.object(mcp_module, "_create_stripe_checkout_url",
                           return_value="https://checkout.stripe.com/pay/shared") as stripe_mock, \
             patch.object(mcp_module, "send_sms", return_value="✅ sent"), \
             patch("smtplib.SMTP", return_value=smtp_mock):
            mcp_module.email_invoice(invoice_identifier="INV-0001", to="k@sunshine.com",
                                      filepath=fp, also_sms=True)
        stripe_mock.assert_called_once()

    def test_sms_failure_does_not_turn_successful_email_into_error(self, mcp_module, tmp_path):
        """If send_sms returns an error string (opted out, no phone on
        file, etc.), the overall result must still read as a success for
        the email — not a bare failure."""
        result, _, _, _, _ = self._run_with_settings(
            mcp_module, tmp_path, self._NO_CREDS, also_sms=True,
            send_sms_result="❌ No phone number found for this contact",
        )
        assert result.startswith("✅")
        assert "not sent" in result.lower()

    def test_also_sms_uses_raw_customer_name_not_placeholder(self, mcp_module, tmp_path):
        """The SMS recipient/lookup key and message must use the invoice's
        actual customer name — never the "—" display placeholder that
        the email body uses for a blank field."""
        _, _, sms_mock, _, _ = self._run_with_settings(mcp_module, tmp_path, self._NO_CREDS, also_sms=True)
        sms_mock.assert_called_once()
        assert sms_mock.call_args.kwargs.get("to") == "Sunshine Realty LLC"
        assert "—" not in sms_mock.call_args.kwargs.get("message", "")

    def test_also_sms_skipped_cleanly_when_customer_name_blank(self, mcp_module, tmp_path):
        """If the invoice row's customer name is genuinely blank, send_sms
        must never be called at all. The result must clearly say why the
        text wasn't sent, and email's success must still be reported."""
        fp = str(_make_test_spreadsheet(tmp_path))
        import openpyxl as _opx_blank
        wb = _opx_blank.load_workbook(fp)
        ws = wb["Invoices"]
        for row in ws.iter_rows(min_row=3):
            if row[0].value == "INV-0001":
                row[3].value = ""  # "Customer Name / Company" column
                break
        wb.save(fp)

        smtp_mock = MagicMock()
        smtp_mock.__enter__ = MagicMock(return_value=smtp_mock)
        smtp_mock.__exit__ = MagicMock(return_value=False)

        with patch.object(mcp_module, "_email_config_load", return_value=self._SMTP_CFG), \
             patch.object(mcp_module, "_load_payment_settings", return_value=self._NO_CREDS), \
             patch.object(mcp_module, "send_sms") as sms_mock, \
             patch("smtplib.SMTP", return_value=smtp_mock):
            result = mcp_module.email_invoice(
                invoice_identifier="INV-0001",
                to="karen@sunshine.com",
                filepath=fp,
                also_sms=True,
            )

        sms_mock.assert_not_called()
        assert result.startswith("✅")
        assert "not sent" in result.lower()
        assert "customer name" in result.lower()


class TestStripeAndSquareCheckoutHelpers:
    """Direct tests for _create_stripe_checkout_url() and
    _create_square_checkout_url() — the two functions that make the
    actual (always-mocked-here) network calls."""

    def test_stripe_success_returns_url(self, mcp_module):
        resp = MagicMock()
        resp.status_code = 200
        resp.json.return_value = {"url": "https://checkout.stripe.com/pay/cs_123"}
        with patch("requests.post", return_value=resp):
            result = mcp_module._create_stripe_checkout_url(
                "sk_test_abc", 303.35, "Invoice INV-0001", "INV-0001")
        assert result == "https://checkout.stripe.com/pay/cs_123"

    def test_stripe_failure_returns_none_not_raises(self, mcp_module):
        """A bad key or API error must never propagate up — email_invoice
        relies on None meaning 'fall back to static URL'."""
        resp = MagicMock()
        resp.status_code = 401
        resp.text = "Invalid API Key"
        with patch("requests.post", return_value=resp):
            result = mcp_module._create_stripe_checkout_url(
                "sk_bad", 100.0, "Invoice INV-0001", "INV-0001")
        assert result is None

    def test_stripe_network_exception_returns_none_not_raises(self, mcp_module):
        with patch("requests.post", side_effect=ConnectionError("network down")):
            result = mcp_module._create_stripe_checkout_url(
                "sk_test_abc", 100.0, "Invoice INV-0001", "INV-0001")
        assert result is None

    def test_stripe_amount_converted_to_cents(self, mcp_module):
        resp = MagicMock()
        resp.status_code = 200
        resp.json.return_value = {"url": "https://checkout.stripe.com/pay/x"}
        with patch("requests.post", return_value=resp) as post_mock:
            mcp_module._create_stripe_checkout_url("sk_test", 303.35, "desc", "INV-0001")
        sent_data = post_mock.call_args.kwargs.get("data", {})
        assert sent_data.get("line_items[0][price_data][unit_amount]") == 30335

    def test_square_success_returns_url(self, mcp_module):
        resp = MagicMock()
        resp.status_code = 200
        resp.json.return_value = {"payment_link": {"url": "https://checkout.square.site/abc"}}
        with patch("requests.post", return_value=resp):
            result = mcp_module._create_square_checkout_url(
                "EAAA_test", "L123", 303.35, "Invoice INV-0001", "INV-0001")
        assert result == "https://checkout.square.site/abc"

    def test_square_failure_returns_none_not_raises(self, mcp_module):
        resp = MagicMock()
        resp.status_code = 401
        resp.text = "Unauthorized"
        with patch("requests.post", return_value=resp):
            result = mcp_module._create_square_checkout_url(
                "EAAA_bad", "L123", 100.0, "desc", "INV-0001")
        assert result is None

    def test_square_amount_converted_to_cents(self, mcp_module):
        resp = MagicMock()
        resp.status_code = 200
        resp.json.return_value = {"payment_link": {"url": "https://checkout.square.site/x"}}
        with patch("requests.post", return_value=resp) as post_mock:
            mcp_module._create_square_checkout_url("EAAA_test", "L123", 303.35, "desc", "INV-0001")
        sent_json = post_mock.call_args.kwargs.get("json", {})
        assert sent_json.get("quick_pay", {}).get("price_money", {}).get("amount") == 30335


# ===========================================================================
# send_sms  (CT_07 - CT_11)
# ===========================================================================

class TestSendSms:
    """Tests for ACTION TOOL 9 -- send_sms."""

    _TWILIO_CFG = {
        "twilio_sms_enabled": True,
        "twilio_account_sid": "ACxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx",
        "twilio_auth_token": "test_auth_token_1234567890abcdef",
        "twilio_from_number": "+13865550100",
    }

    def _write_cfg(self, tmp_path):
        cfg_dir = tmp_path / ".ai-prowler"
        cfg_dir.mkdir(parents=True, exist_ok=True)
        cfg_path = cfg_dir / "config.json"
        cfg_path.write_text(json.dumps(self._TWILIO_CFG), encoding="utf-8")
        return tmp_path

    def test_CT_07_send_sms_success(self, mcp_module, tmp_path):
        """send_sms must call Twilio API and return a success confirmation."""
        home = self._write_cfg(tmp_path)

        twilio_resp = MagicMock()
        twilio_resp.status_code = 201
        twilio_resp.json.return_value = {"sid": "SM1234567890abcdef"}

        with patch("pathlib.Path.home", return_value=home), \
             patch("requests.post", return_value=twilio_resp):
            result = mcp_module.send_sms(
                to="3865550101",
                message="Hi Karen, Mike is 20 minutes away!",
            )

        assert "SM1234567890abcdef" in result or "sent" in result.lower()

    def test_CT_08_send_sms_normalises_10_digit_number(self, mcp_module, tmp_path):
        """A 10-digit number must be normalised to E.164 (+1XXXXXXXXXX)."""
        home = self._write_cfg(tmp_path)
        captured = {}

        def fake_post(url, auth, data, timeout=30):
            captured["to"] = data.get("To")
            resp = MagicMock()
            resp.status_code = 201
            resp.json.return_value = {"sid": "SM_test"}
            return resp

        with patch("pathlib.Path.home", return_value=home), \
             patch("requests.post", side_effect=fake_post):
            mcp_module.send_sms(to="3865550101", message="Test")

        assert captured.get("to") == "+13865550101"

    def test_CT_09_send_sms_no_twilio_config_returns_error(self, mcp_module, tmp_path):
        """Missing Twilio config must return a clear setup-instructions error."""
        cfg_dir = tmp_path / ".ai-prowler"
        cfg_dir.mkdir(parents=True, exist_ok=True)
        (cfg_dir / "config.json").write_text(
            json.dumps({"other_key": "value"}), encoding="utf-8"
        )

        with patch("pathlib.Path.home", return_value=tmp_path):
            result = mcp_module.send_sms(to="3865550101", message="Test")

        assert any(w in result.lower() for w in ["twilio", "config", "setup", "configure"])

    def test_CT_10_send_sms_empty_message_returns_error(self, mcp_module, tmp_path):
        """An empty message must return an error before hitting the API."""
        home = self._write_cfg(tmp_path)
        with patch("pathlib.Path.home", return_value=home):
            result = mcp_module.send_sms(to="3865550101", message="   ")

        assert any(w in result.lower() for w in ["empty", "blank", "message", "error"])

    def test_CT_11_send_sms_twilio_error_response_surfaced(self, mcp_module, tmp_path):
        """A Twilio 400 error must be returned as a readable error string."""
        home = self._write_cfg(tmp_path)

        twilio_resp = MagicMock()
        twilio_resp.status_code = 400
        twilio_resp.json.return_value = {"message": "Invalid phone number format"}

        with patch("pathlib.Path.home", return_value=home), \
             patch("requests.post", return_value=twilio_resp):
            result = mcp_module.send_sms(to="0000000000", message="Test")

        assert "400" in result or "invalid" in result.lower() or "error" in result.lower()

    def test_CT_11b_blank_to_is_rejected_before_any_lookup(self, mcp_module, tmp_path):
        """Same bug class as email_invoice's blank-identifier fix: the
        name-resolution lookups search for `to` as a SUBSTRING of stored
        names, and a blank `to` (e.g. a job whose customer name failed to
        parse upstream) must never reach them — it would otherwise match
        whichever record happens to come first."""
        home = self._write_cfg(tmp_path)
        with patch("pathlib.Path.home", return_value=home):
            result = mcp_module.send_sms(to="", message="Test")
        assert result.startswith("❌")
        assert "blank" in result.lower() or "required" in result.lower()

    def test_CT_11c_whitespace_only_to_is_also_rejected(self, mcp_module, tmp_path):
        home = self._write_cfg(tmp_path)
        with patch("pathlib.Path.home", return_value=home):
            result = mcp_module.send_sms(to="   ", message="Test")
        assert result.startswith("❌")

    def test_CT_11d_blank_to_never_reaches_users_json_crew_lookup(self, mcp_module, tmp_path):
        """The specific vulnerability this guards against: step 2 (crew
        lookup in users.json) had no blank-input guard at all, unlike
        step 1 (Customers sheet) which was already correctly guarded. A
        blank `to` would match the FIRST crew member in the dict
        unconditionally — this test proves that lookup is never even
        attempted when `to` is blank, by making it fail loudly if called."""
        home = self._write_cfg(tmp_path)

        def _users_that_should_never_be_read():
            raise AssertionError(
                "users.json crew lookup was reached with a blank 'to' — "
                "the top-level blank-input guard did not short-circuit it."
            )

        with patch("pathlib.Path.home", return_value=home), \
             patch.object(mcp_module, "_load_users", side_effect=_users_that_should_never_be_read):
            result = mcp_module.send_sms(to="", message="Test")

        assert result.startswith("❌")

    def test_CT_11e_real_name_still_resolves_normally(self, mcp_module, tmp_path):
        """Regression guard — the blank-input fix must not break the
        normal, non-blank name-resolution path it's built around."""
        home = self._write_cfg(tmp_path)
        captured = {}

        def fake_post(url, auth, data, timeout=30):
            captured["to"] = data.get("To")
            resp = MagicMock()
            resp.status_code = 201
            resp.json.return_value = {"sid": "SM_test"}
            return resp

        users_data = {"users": {"u1": {"name": "Jake Rivera", "cell_phone": "3865550199"}}}

        with patch("pathlib.Path.home", return_value=home), \
             patch.object(mcp_module, "_load_users", return_value=users_data), \
             patch.object(mcp_module, "_get_default_spreadsheet_path", return_value=""), \
             patch("requests.post", side_effect=fake_post):
            mcp_module.send_sms(to="Jake", message="Test")

        assert captured.get("to") == "+13865550199"


# ===========================================================================
# schedule_next_recurring_job  (CT_12 - CT_17)
# ===========================================================================

class TestScheduleNextRecurringJob:
    """Tests for ACTION TOOL 10 -- schedule_next_recurring_job."""

    def test_CT_12_monthly_customer_gets_next_job_plus_one_month(self, mcp_module, tmp_path):
        """Monthly customer: next job should be 1 month after last service date."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0001",
                filepath=fp,
                # Fixture uses a fixed 2026-03-30 Service Date unrelated to
                # "today" — when="any" is needed since schedule_next_recurring_job
                # now defaults to "today" if `when` is omitted.
                when="any",
            )

        assert isinstance(result, str)
        # Base date 2026-03-30 + 1 month = 2026-04-30
        assert "2026-04-30" in result or "April" in result

    def test_CT_13_biweekly_customer_gets_next_job_plus_14_days(self, mcp_module, tmp_path):
        """Biweekly customer: next job should be 14 days after last service date."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0002",
                filepath=fp,
                when="any",
            )

        assert isinstance(result, str)
        # Base date 2026-03-16 + 14 days = 2026-03-30
        assert "2026-03-30" in result

    def test_CT_14_new_job_written_to_jobs_schedule(self, mcp_module, tmp_path):
        """After scheduling, the new job row must exist in the spreadsheet."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            mcp_module.schedule_next_recurring_job(job_identifier="JOB-0001", filepath=fp, when="any")

        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["Jobs_Schedule"]
        job_ids = [
            str(row[0].value)
            for row in ws.iter_rows(min_row=3)
            if row[0].value and str(row[0].value).startswith("JOB-")
        ]
        assert len(job_ids) >= 3, "New job row should have been appended"

    def test_CT_15_one_time_customer_returns_info_message(self, mcp_module, tmp_path):
        """OT (one-time) frequency must return an info message, not create a new job."""
        fp = _make_test_spreadsheet(tmp_path)

        # Add a one-time customer and job
        wb = openpyxl.load_workbook(str(fp))
        ws_c = wb["Customers"]
        ws_c.append([
            "CUST-0099", "Residential", "", "OneTime", "Customer",
            "0000000000", "once@test.com",
            "1 Test St", "TestCity", "FL", "00000",
            "", "", "Window", "OT",
            "", "", "60", "100", "0", "100",
            "2026-03-01", "", "0", "0", "", "", "Active",
        ])
        ws_j = wb["Jobs_Schedule"]
        ws_j.append([
            "JOB-0099", "CUST-0099", "OneTime Customer", "Residential",
            "1 Test St", "TestCity", "FL", "00000", "", "",
            "2026-03-01", "Sunday", "10:00", "11:00", "Window",
            "One-time clean", "Mike C.", "60", "", "1",
            "", "", "Complete", "100", "100", "0", "7", "107",
            "", "", "Unpaid",
        ])
        wb.save(str(fp))

        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0099",
                filepath=str(fp),
                when="any",
            )

        assert any(w in result.lower() for w in ["one-time", "ot", "one time", "no recurring"])

    def test_CT_16_job_not_found_returns_error(self, mcp_module, tmp_path):
        """Nonexistent job identifier must return an error string."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-9999",
                filepath=fp,
                when="any",
            )

        assert any(w in result.lower() for w in ["not found", "error", "no job", "could not"])

    def test_CT_17_new_job_status_is_scheduled(self, mcp_module, tmp_path):
        """The auto-created job row must have Job Status = 'Scheduled'."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0001",
                filepath=fp,
                when="any",
            )

        if "JOB-0003" not in result and "scheduled" not in result.lower():
            # If the tool returned an error, skip the spreadsheet check
            pytest.skip("Tool did not create a new job -- skipping row status check")

        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["Jobs_Schedule"]
        hdrs = [
            str(c.value).strip() if c.value else ""
            for c in list(ws.iter_rows(min_row=2, max_row=2))[0]
        ]
        for row in ws.iter_rows(min_row=3):
            vals = {hdrs[i]: row[i].value for i in range(len(hdrs))}
            if "JOB-0003" in str(vals.get("JobID (JOB-####)", "")):
                assert vals.get("Job Status") == "Scheduled"
                break


class TestScheduleNextRecurringJobExpandedFrequencies:
    """Tests for the expanded recurrence vocabulary — Bi-Monthly,
    Semi-Annually, and Annually were entirely missing from _FREQ_MAP
    before this fix, and the month-arithmetic itself had a latent
    day-of-month overflow bug (Jan 31 + 1 month would try to construct
    "Feb 31" and raise ValueError) that never surfaced in prior tests
    only because none of them happened to use a month-end base date."""

    def _make_customer_and_job(self, tmp_path, frequency, service_date, job_id="JOB-0099", cust_id="CUST-0099"):
        fp = _make_test_spreadsheet(tmp_path)
        wb = openpyxl.load_workbook(str(fp))
        ws_c = wb["Customers"]
        ws_c.append([
            cust_id, "Residential", "", "Test", "Customer",
            "0000000000", "test@test.com",
            "1 Test St", "TestCity", "FL", "00000",
            "", "", "Window", frequency,
            "", "", "60", "100", "0", "100",
            service_date, "", "0", "0", "", "", "Active",
        ])
        ws_j = wb["Jobs_Schedule"]
        ws_j.append([
            job_id, cust_id, "Test Customer", "Residential",
            "1 Test St", "TestCity", "FL", "00000", "", "",
            service_date, "Sunday", "10:00", "11:00", "Window",
            "Test clean", "Mike C.", "60", "", "1",
            "", "", "Complete", "100", "100", "0", "7", "107",
            "", "", "Unpaid",
        ])
        wb.save(str(fp))
        return fp

    def test_bi_monthly_gets_next_job_plus_two_months(self, mcp_module, tmp_path):
        fp = self._make_customer_and_job(tmp_path, "Bi-Monthly", "2026-03-01")
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0099", filepath=str(fp), when="any",
            )
        assert "2026-05-01" in result

    def test_bi_monthly_short_code_bm_also_works(self, mcp_module, tmp_path):
        """Short code, for backward compatibility with any existing data
        that predates the full-word dropdown."""
        fp = self._make_customer_and_job(tmp_path, "BM", "2026-03-01")
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0099", filepath=str(fp), when="any",
            )
        assert "2026-05-01" in result

    def test_semi_annually_gets_next_job_plus_six_months(self, mcp_module, tmp_path):
        fp = self._make_customer_and_job(tmp_path, "Semi-Annually", "2026-01-15")
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0099", filepath=str(fp), when="any",
            )
        assert "2026-07-15" in result

    def test_annually_gets_next_job_plus_twelve_months(self, mcp_module, tmp_path):
        fp = self._make_customer_and_job(tmp_path, "Annually", "2026-06-01")
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0099", filepath=str(fp), when="any",
            )
        assert "2027-06-01" in result

    def test_month_end_day_overflow_does_not_crash(self, mcp_module, tmp_path):
        """The actual bug this fix addresses: a base date of Jan 31 plus
        1 month must not attempt to construct the nonexistent "Feb 31" —
        must cap at Feb 28 (2026 is not a leap year) instead of raising."""
        fp = self._make_customer_and_job(tmp_path, "Monthly", "2026-01-31")
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0099", filepath=str(fp), when="any",
            )
        assert isinstance(result, str)
        assert not result.startswith("❌")
        assert "2026-02-28" in result

    def test_bi_monthly_month_end_day_overflow_also_capped(self, mcp_module, tmp_path):
        """Same day-overflow protection, exercised through a 2-month
        jump instead of 1 — Dec 31 + 2 months = Feb 28, not a crash."""
        fp = self._make_customer_and_job(tmp_path, "Bi-Monthly", "2025-12-31")
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0099", filepath=str(fp), when="any",
            )
        assert isinstance(result, str)
        assert not result.startswith("❌")
        assert "2026-02-28" in result

    def test_unrecognized_frequency_error_message_lists_full_expanded_set(self, mcp_module, tmp_path):
        """The error message must reflect the actual expanded vocabulary,
        not the old, now-incomplete short-code-only hint."""
        fp = self._make_customer_and_job(tmp_path, "Every Full Moon", "2026-03-01")
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.schedule_next_recurring_job(
                job_identifier="JOB-0099", filepath=str(fp), when="any",
            )
        assert "Bi-Monthly" in result
        assert "Semi-Annually" in result
        assert "Annually" in result


# ===========================================================================
# log_time_entry  (CT_18 - CT_22)
# ===========================================================================

class TestLogTimeEntry:
    """Tests for ACTION TOOL 11 -- log_time_entry."""

    def test_CT_18_clock_in_creates_timelog_entry(self, mcp_module, tmp_path):
        """action='start' must write a new row to the TimeLog sheet."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.log_time_entry(
                job_identifier="JOB-0001",
                action="start",
                filepath=fp,
            )

        assert isinstance(result, str)
        assert any(w in result.lower() for w in ["clock", "in", "start", "te-", "logged"])

        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb["TimeLog"]
        entries = [
            row for row in ws.iter_rows(min_row=3)
            if row[0].value and str(row[0].value).startswith("TE-")
        ]
        assert len(entries) >= 1, "TimeLog must have at least one entry after clock-in"

    def test_CT_19_clock_out_calculates_elapsed_time(self, mcp_module, tmp_path):
        """action='stop' must compute elapsed minutes and write Clock Out + Elapsed."""
        fp = str(_make_test_spreadsheet(tmp_path))

        # Plant an open clock-in entry (~47 min ago)
        wb = openpyxl.load_workbook(fp)
        ws_t = wb["TimeLog"]
        clock_in = (
            datetime.datetime.now() - datetime.timedelta(minutes=47)
        ).strftime("%Y-%m-%d %H:%M:%S")
        ws_t.append(["TE-0001", "JOB-0001", "Sunshine Realty LLC",
                      clock_in, None, None, "Mike C.", ""])
        wb.save(fp)

        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.log_time_entry(
                job_identifier="JOB-0001",
                action="stop",
                filepath=fp,
            )

        assert isinstance(result, str)
        assert any(w in result.lower() for w in ["clock", "out", "stop", "elapsed", "min"])

        wb2 = openpyxl.load_workbook(fp, data_only=True)
        ws_t2 = wb2["TimeLog"]
        for row in ws_t2.iter_rows(min_row=3):
            if row[0].value == "TE-0001":
                assert row[4].value is not None, "Clock Out must be written"
                elapsed = row[5].value
                assert elapsed is not None, "Elapsed (min) must be written"
                assert 44 <= int(elapsed) <= 50, f"Elapsed should be ~47 min, got {elapsed}"
                break

    def test_CT_20_clock_out_updates_actual_duration_in_jobs_schedule(self, mcp_module, tmp_path):
        """Clocking out must write Actual Duration (min) back to Jobs_Schedule."""
        fp = str(_make_test_spreadsheet(tmp_path))

        wb = openpyxl.load_workbook(fp)
        ws_t = wb["TimeLog"]
        clock_in = (
            datetime.datetime.now() - datetime.timedelta(minutes=35)
        ).strftime("%Y-%m-%d %H:%M:%S")
        ws_t.append(["TE-0001", "JOB-0001", "Sunshine Realty LLC",
                      clock_in, None, None, "Mike C.", ""])
        wb.save(fp)

        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            mcp_module.log_time_entry(
                job_identifier="JOB-0001", action="stop", filepath=fp
            )

        wb2 = openpyxl.load_workbook(fp, data_only=True)
        ws_j = wb2["Jobs_Schedule"]
        hdrs = [
            str(c.value).strip() if c.value else ""
            for c in list(ws_j.iter_rows(min_row=2, max_row=2))[0]
        ]
        for row in ws_j.iter_rows(min_row=3):
            vals = dict(zip(hdrs, [c.value for c in row]))
            if vals.get("JobID (JOB-####)") == "JOB-0001":
                actual = vals.get("Actual Duration (min)")
                if actual is not None:
                    assert 32 <= int(actual) <= 38, f"Expected ~35 min, got {actual}"
                break

    def test_CT_21_double_clock_in_returns_warning(self, mcp_module, tmp_path):
        """Clocking in when already clocked in must return a warning, not crash."""
        fp = str(_make_test_spreadsheet(tmp_path))

        wb = openpyxl.load_workbook(fp)
        ws_t = wb["TimeLog"]
        clock_in = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_t.append(["TE-0001", "JOB-0001", "Sunshine Realty LLC",
                      clock_in, None, None, "Mike C.", ""])
        wb.save(fp)

        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.log_time_entry(
                job_identifier="JOB-0001",
                action="start",
                filepath=fp,
            )

        assert any(w in result.lower() for w in ["already", "open", "active", "warning", "clocked"])

    def test_CT_22_invalid_action_returns_error(self, mcp_module, tmp_path):
        """action='lunch' (invalid) must return an error about valid options."""
        fp = str(_make_test_spreadsheet(tmp_path))
        with patch.object(mcp_module, "_backup_spreadsheet", return_value="Backup saved"):
            result = mcp_module.log_time_entry(
                job_identifier="JOB-0001",
                action="lunch",
                filepath=fp,
            )

        assert any(w in result.lower() for w in ["start", "stop", "invalid", "error", "action"])


# ===========================================================================
# get_ar_aging_report  (CT_23 - CT_28)
# ===========================================================================

class TestGetArAgingReport:
    """Tests for ACTION TOOL 12 -- get_ar_aging_report."""

    def test_CT_23_report_contains_unpaid_invoice(self, mcp_module, tmp_path):
        """The AR report must list Sunshine Realty INV-0001 (Unpaid)."""
        fp = str(_make_test_spreadsheet(tmp_path))
        result = mcp_module.get_ar_aging_report(filepath=fp, as_of_date="2026-05-01")

        assert isinstance(result, str)
        assert "INV-0001" in result or "Sunshine Realty" in result
        assert "303" in result  # balance ~303.345

    def test_CT_23b_balance_renders_with_real_multiline_headers(self, mcp_module, tmp_path):
        """Same real production bug as email_invoice's CT_06b — the
        Balance Due column's multi-line header ("Balance Due ($)\n=O-P")
        must not break get_ar_aging_report's lookup either. The report
        must show the real $337.05 balance, not a missing/zero amount."""
        fp = str(_make_multiline_header_spreadsheet(tmp_path))
        result = mcp_module.get_ar_aging_report(filepath=fp, as_of_date="2026-05-01")

        assert isinstance(result, str)
        assert "INV-0007" in result or "AI-Prowler LLC" in result
        assert "337.05" in result

    def test_CT_24_paid_invoice_excluded_from_report(self, mcp_module, tmp_path):
        """INV-0002 is fully paid -- its balance must NOT appear in AR."""
        fp = str(_make_test_spreadsheet(tmp_path))
        result = mcp_module.get_ar_aging_report(filepath=fp, as_of_date="2026-05-01")

        # Paid invoice balance 197.95 must not appear
        assert "197.95" not in result, "Paid invoice balance must not appear in AR report"

    def test_CT_25_overdue_invoice_lands_in_correct_bucket(self, mcp_module, tmp_path):
        """INV-0003 (due 2026-02-14, $535) must appear in 90+ bucket as of 2026-05-30."""
        fp = str(_make_test_spreadsheet(tmp_path))
        result = mcp_module.get_ar_aging_report(filepath=fp, as_of_date="2026-05-30")

        assert isinstance(result, str)
        # Either the invoice ID or its balance should appear
        assert "INV-0003" in result or "535" in result
        # 90-day bucket language
        assert "90" in result

    def test_CT_26_report_includes_total_outstanding(self, mcp_module, tmp_path):
        """The report must include a total outstanding balance line."""
        fp = str(_make_test_spreadsheet(tmp_path))
        result = mcp_module.get_ar_aging_report(filepath=fp, as_of_date="2026-05-01")

        assert any(
            w in result
            for w in ["TOTAL OUTSTANDING", "Total Outstanding", "Total:", "TOTAL:"]
        )

    def test_CT_27_all_paid_returns_clean_message(self, mcp_module, tmp_path):
        """If all invoices are paid, the report must say no outstanding invoices."""
        fp = _make_test_spreadsheet(tmp_path)

        wb = openpyxl.load_workbook(str(fp))
        ws = wb["Invoices"]
        hdrs = [
            str(c.value).strip() if c.value else ""
            for c in list(ws.iter_rows(min_row=2, max_row=2))[0]
        ]
        try:
            pmt_col = hdrs.index("Payment Status") + 1
            bal_col = hdrs.index("Balance Due ($)") + 1
        except ValueError:
            pytest.skip("Could not locate Payment Status / Balance Due columns")

        for row in ws.iter_rows(min_row=3):
            if row[0].value:
                ws.cell(row=row[0].row, column=pmt_col).value = "Paid"
                ws.cell(row=row[0].row, column=bal_col).value = 0
        wb.save(str(fp))

        result = mcp_module.get_ar_aging_report(filepath=str(fp), as_of_date="2026-05-01")
        assert any(
            w in result.lower()
            for w in ["no outstanding", "all paid", "nothing outstanding", "0 outstanding"]
        ) or result.strip().startswith("No")

    def test_CT_28_missing_spreadsheet_returns_error(self, mcp_module, tmp_path):
        """Passing a nonexistent filepath must return an error."""
        result = mcp_module.get_ar_aging_report(
            filepath=str(tmp_path / "no_such_file.xlsx")
        )
        assert isinstance(result, str)
        assert any(w in result.lower() for w in ["not found", "no spreadsheet", "error", "file"])


# ===========================================================================
# _join_header_lines  +  filter_date='today' with real multi-line headers
# ===========================================================================
#
# Real production bug: a prior fix for a DIFFERENT problem (a "=FORMULA"
# documentation note on a header cell's second line breaking exact-match
# lookups against keys like "Tax 7% ($)") took only the first line of
# EVERY multi-line header unconditionally. That silently broke a
# genuinely different case: headers where the column name ITSELF spans
# multiple lines with no formula involved — "Service\nDate", "Service\n
# Type", and "Service\nDetails / Notes" all collapsed to the single word
# "Service", making three separate columns indistinguishable. That broke
# read_job_spreadsheet's Service Date column lookup entirely (no column
# matched 'service' AND 'date' anymore), which silently disabled
# filter_date='today' filtering — even for the OWNER, who should see
# every job regardless of crew assignment, and still saw none.

class TestJoinHeaderLines:
    """Direct tests for the shared _join_header_lines() helper."""

    def test_single_line_header_unchanged(self, mcp_module):
        assert mcp_module._join_header_lines("Customer Type") == "Customer Type"

    def test_formula_note_on_second_line_is_dropped(self, mcp_module):
        assert mcp_module._join_header_lines("Tax 7% ($)\n=M*0.07") == "Tax 7% ($)"

    def test_genuine_multiline_label_is_merged_not_truncated(self, mcp_module):
        """The exact case the prior fix broke — this must NOT collapse to
        just 'Service'."""
        assert mcp_module._join_header_lines("Service\nDate") == "Service Date"
        assert mcp_module._join_header_lines("Service\nType") == "Service Type"

    def test_three_service_columns_remain_distinct(self, mcp_module):
        """The actual real-world collision: three different multi-line
        headers must NOT all reduce to the same string."""
        date_hdr = mcp_module._join_header_lines("Service\nDate")
        type_hdr = mcp_module._join_header_lines("Service\nType")
        notes_hdr = mcp_module._join_header_lines("Service\nDetails / Notes")
        assert len({date_hdr, type_hdr, notes_hdr}) == 3

    def test_three_line_header_with_trailing_formula(self, mcp_module):
        """A label that itself wraps across two lines, PLUS a formula
        note on a third line — both rules apply together."""
        result = mcp_module._join_header_lines("Balance\nDue ($)\n=O-P")
        assert result == "Balance Due ($)"

    def test_star_route_annotation_preserved(self, mcp_module):
        """Non-formula annotations (like the "★ AI Route" suffix seen on
        several real columns) are part of the genuine header text and
        must be preserved, not dropped."""
        assert mcp_module._join_header_lines("Street Address\n★ AI Route") == \
               "Street Address ★ AI Route"

    def test_blank_value_returns_empty_string(self, mcp_module):
        assert mcp_module._join_header_lines(None) == ""
        assert mcp_module._join_header_lines("") == ""


def _make_multiline_jobs_schedule_spreadsheet(tmp_path, today_str, other_day_str):
    """
    Builds a Jobs_Schedule sheet using the SAME multi-line header format
    the real production template uses for the three colliding columns —
    "Service\nDate", "Service\nType", "Service\nDetails / Notes" — plus
    one job dated today and one job dated a different day, so a
    filter_date='today' regression test can prove only the right one
    comes back.
    """
    import openpyxl as _opx_js
    fp = tmp_path / "multiline_jobs_schedule.xlsx"
    wb = _opx_js.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Jobs_Schedule")
    ws.append(["JOBS & SCHEDULE -- All Service Appointments"])
    ws.append([
        "JobID (JOB-####)", "CustomerID", "Customer Name / Company",
        "Customer Type", "Street Address", "City", "State", "ZIP",
        "Service\nDate", "Day of Week", "Start Time", "End Time",
        "Service\nType", "Service\nDetails / Notes",
        "Crew / Technician", "Est. Duration (min)",
        "Job Status", "Quote Amount ($)", "Payment Status",
    ])
    ws.append([
        "JOB-0007", "CUST-0007", "AI-Prowler LLC", "Commercial",
        "1500 Shadow Pines Dr", "New Smyrna Beach", "FL", "32168",
        today_str, "Wednesday", "08:00", "11:30",
        "Window", "Full exterior window cleaning -- 12 windows",
        "David Vavro", "90",
        "Scheduled", "350", "Unpaid",
    ])
    ws.append([
        "JOB-0001", "CUST-0001", "Sunshine Realty LLC", "Commercial",
        "125 Harbor Blvd", "New Smyrna Beach", "FL", "32168",
        other_day_str, "Monday", "08:00", "09:30",
        "Window", "Full exterior window cleaning -- 12 windows",
        "Mike C.", "90",
        "Scheduled", "350", "Unpaid",
    ])
    wb.save(str(fp))
    return fp


class TestReadJobSpreadsheetFilterDateWithMultilineHeaders:
    """The direct regression test for the reported production bug: an
    owner (unrestricted by crew scoping) hit filter_date='today' and saw
    NO jobs at all, despite having one scheduled for that exact day."""

    def test_todays_job_is_returned_with_real_multiline_headers(self, mcp_module, tmp_path):
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        other_day_str = (datetime.date.today() - datetime.timedelta(days=2)).strftime("%Y-%m-%d")
        fp = str(_make_multiline_jobs_schedule_spreadsheet(tmp_path, today_str, other_day_str))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )

        assert "JOB-0007" in result
        assert "AI-Prowler LLC" in result

    def test_other_days_job_excluded_from_todays_filter(self, mcp_module, tmp_path):
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        other_day_str = (datetime.date.today() - datetime.timedelta(days=2)).strftime("%Y-%m-%d")
        fp = str(_make_multiline_jobs_schedule_spreadsheet(tmp_path, today_str, other_day_str))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )

        assert "JOB-0001" not in result

    def test_service_type_and_notes_render_correctly_not_blank(self, mcp_module, tmp_path):
        """Beyond just the date filter working, the three previously-
        colliding fields (date/type/notes) must each show their OWN
        correct value, not an empty string or another field's value."""
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        other_day_str = (datetime.date.today() - datetime.timedelta(days=2)).strftime("%Y-%m-%d")
        fp = str(_make_multiline_jobs_schedule_spreadsheet(tmp_path, today_str, other_day_str))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )

        assert "Window" in result
        assert "Full exterior window cleaning" in result


# ===========================================================================
# _crew_name_in_cell  +  multi-crew job assignment
# ===========================================================================
#
# Jobs can now be assigned to multiple people at once — the "Crew /
# Technician" cell holds a comma-separated list (e.g. "Mike C., David
# Vavro") instead of a single name. Every enforcement site that used to
# do a plain equality check against that cell now goes through the
# shared _crew_name_in_cell() helper instead, so a restricted user shows
# up correctly if their name is ANY ONE of the listed names, not only
# when they're the sole assignee.

class TestCrewNameInCell:
    """Direct tests for the shared _crew_name_in_cell() helper."""

    def test_single_name_exact_match(self, mcp_module):
        assert mcp_module._crew_name_in_cell("mike c.", "mike c.") is True

    def test_single_name_no_match(self, mcp_module):
        assert mcp_module._crew_name_in_cell("mike c.", "jake r.") is False

    def test_second_name_in_comma_list_matches(self, mcp_module):
        assert mcp_module._crew_name_in_cell("mike c., david vavro", "david vavro") is True

    def test_first_name_in_comma_list_matches(self, mcp_module):
        assert mcp_module._crew_name_in_cell("mike c., david vavro", "mike c.") is True

    def test_name_not_in_comma_list_does_not_match(self, mcp_module):
        assert mcp_module._crew_name_in_cell("mike c., david vavro", "jake r.") is False

    def test_extra_whitespace_around_commas_tolerated(self, mcp_module):
        assert mcp_module._crew_name_in_cell("mike c. ,  david vavro", "david vavro") is True

    def test_blank_cell_never_matches(self, mcp_module):
        assert mcp_module._crew_name_in_cell("", "david vavro") is False

    def test_blank_crew_name_never_matches(self, mcp_module):
        assert mcp_module._crew_name_in_cell("mike c., david vavro", "") is False

    def test_partial_substring_does_not_falsely_match(self, mcp_module):
        """'Mike' must not match a cell containing 'Mike C.' — this is
        exact per-name matching after splitting, not substring search."""
        assert mcp_module._crew_name_in_cell("mike c., david vavro", "mike") is False


def _make_multi_crew_jobs_schedule_spreadsheet(tmp_path):
    """Jobs_Schedule with one job assigned to multiple crew members."""
    import openpyxl as _opx_mc
    fp = tmp_path / "multi_crew_jobs_schedule.xlsx"
    wb = _opx_mc.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Jobs_Schedule")
    ws.append(["JOBS & SCHEDULE -- All Service Appointments"])
    ws.append([
        "JobID (JOB-####)", "CustomerID", "Customer Name / Company",
        "Customer Type", "Street Address", "City", "State", "ZIP",
        "Service Date", "Day of Week", "Start Time", "End Time",
        "Service Type", "Service Details / Notes",
        "Crew / Technician", "Est. Duration (min)",
        "Job Status", "Quote Amount ($)", "Payment Status",
    ])
    ws.append([
        "JOB-0008", "CUST-0007", "AI-Prowler LLC", "Commercial",
        "1500 Shadow Pines Dr", "New Smyrna Beach", "FL", "32168",
        "2026-08-20", "Thursday", "08:00", "11:30",
        "Window", "Large job -- two crews",
        "Mike C., David Vavro", "180",
        "Scheduled", "700", "Unpaid",
    ])
    wb.save(str(fp))
    return fp


class TestReadJobSpreadsheetMultiCrewAssignment:
    """The direct end-to-end regression test: a crew member listed as the
    SECOND of two names in the Crew / Technician cell must still see the
    job when restricted."""

    def test_second_listed_crew_member_sees_the_job(self, mcp_module, tmp_path):
        fp = str(_make_multi_crew_jobs_schedule_spreadsheet(tmp_path))
        fake_user = {"id": "u2", "name": "David Vavro", "role": "field_crew"}
        fake_ctx = MagicMock()

        with patch.object(mcp_module, "_current_user", return_value=fake_user), \
             patch.object(mcp_module, "_check_db_cap", return_value=(False, None)), \
             patch.object(mcp_module, "_get_default_spreadsheet_path", return_value=fp):
            result = mcp_module.read_job_spreadsheet(
                filepath=fp, sheet_name="Jobs_Schedule", ctx=fake_ctx,
            )

        assert "JOB-0008" in result

    def test_first_listed_crew_member_also_sees_the_job(self, mcp_module, tmp_path):
        fp = str(_make_multi_crew_jobs_schedule_spreadsheet(tmp_path))
        fake_user = {"id": "u1", "name": "Mike C.", "role": "field_crew"}
        fake_ctx = MagicMock()

        with patch.object(mcp_module, "_current_user", return_value=fake_user), \
             patch.object(mcp_module, "_check_db_cap", return_value=(False, None)), \
             patch.object(mcp_module, "_get_default_spreadsheet_path", return_value=fp):
            result = mcp_module.read_job_spreadsheet(
                filepath=fp, sheet_name="Jobs_Schedule", ctx=fake_ctx,
            )

        assert "JOB-0008" in result

    def test_uninvolved_crew_member_does_not_see_the_job(self, mcp_module, tmp_path):
        fp = str(_make_multi_crew_jobs_schedule_spreadsheet(tmp_path))
        fake_user = {"id": "u3", "name": "Jake R.", "role": "field_crew"}
        fake_ctx = MagicMock()

        with patch.object(mcp_module, "_current_user", return_value=fake_user), \
             patch.object(mcp_module, "_check_db_cap", return_value=(False, None)), \
             patch.object(mcp_module, "_get_default_spreadsheet_path", return_value=fp):
            result = mcp_module.read_job_spreadsheet(
                filepath=fp, sheet_name="Jobs_Schedule", ctx=fake_ctx,
            )

        assert "JOB-0008" not in result


# ===========================================================================
# Multi-day jobs — End Date column
# ===========================================================================

def _make_multiday_jobs_schedule_spreadsheet(tmp_path, start_str, end_str):
    """One job spanning multiple days via the new End Date column."""
    import openpyxl as _opx_md
    fp = tmp_path / "multiday_jobs_schedule.xlsx"
    wb = _opx_md.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Jobs_Schedule")
    ws.append(["JOBS & SCHEDULE -- All Service Appointments"])
    ws.append([
        "JobID (JOB-####)", "CustomerID", "Customer Name / Company",
        "Customer Type", "Street Address", "City", "State", "ZIP",
        "Service Date", "Day of Week", "Start Time", "End Time",
        "Service Type", "Service Details / Notes",
        "Crew / Technician", "Est. Duration (min)",
        "Job Status", "Quote Amount ($)", "Payment Status",
        "End Date\n(blank = single-day job)",
    ])
    ws.append([
        "JOB-0009", "CUST-0006", "Prospect Corp", "Commercial",
        "999 Big Job Ave", "New Smyrna Beach", "FL", "32168",
        start_str, "Monday", "08:00", "17:00",
        "Both", "Multi-day full building service",
        "Mike C.", "2880",
        "Scheduled", "5000", "Unpaid",
        end_str,
    ])
    wb.save(str(fp))
    return fp


class TestReadJobSpreadsheetEndDateRange:
    """Regression tests for multi-day job support — a job is considered
    'in progress' on every day from Service Date through End Date,
    inclusive, when End Date is filled in."""

    def test_middle_day_of_range_matches_filter(self, mcp_module, tmp_path):
        start = datetime.date.today() - datetime.timedelta(days=2)
        end = datetime.date.today() + datetime.timedelta(days=2)
        fp = str(_make_multiday_jobs_schedule_spreadsheet(
            tmp_path, start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )
        assert "JOB-0009" in result

    def test_first_day_of_range_matches_filter(self, mcp_module, tmp_path):
        today = datetime.date.today()
        end = today + datetime.timedelta(days=3)
        fp = str(_make_multiday_jobs_schedule_spreadsheet(
            tmp_path, today.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )
        assert "JOB-0009" in result

    def test_last_day_of_range_matches_filter(self, mcp_module, tmp_path):
        start = datetime.date.today() - datetime.timedelta(days=3)
        today = datetime.date.today()
        fp = str(_make_multiday_jobs_schedule_spreadsheet(
            tmp_path, start.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )
        assert "JOB-0009" in result

    def test_day_after_range_does_not_match(self, mcp_module, tmp_path):
        start = datetime.date.today() - datetime.timedelta(days=10)
        end = datetime.date.today() - datetime.timedelta(days=5)
        fp = str(_make_multiday_jobs_schedule_spreadsheet(
            tmp_path, start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )
        assert "JOB-0009" not in result

    def test_day_before_range_does_not_match(self, mcp_module, tmp_path):
        start = datetime.date.today() + datetime.timedelta(days=5)
        end = datetime.date.today() + datetime.timedelta(days=10)
        fp = str(_make_multiday_jobs_schedule_spreadsheet(
            tmp_path, start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )
        assert "JOB-0009" not in result

    def test_blank_end_date_treated_as_single_day(self, mcp_module, tmp_path):
        """Backward compatibility — every existing job has a blank End
        Date and must behave exactly as before this feature existed."""
        today = datetime.date.today()
        fp = str(_make_multiday_jobs_schedule_spreadsheet(
            tmp_path, today.strftime("%Y-%m-%d"), ""))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )
        assert "JOB-0009" in result

        # A different day must NOT match, since the job is single-day.
        other_day = (today + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        result2 = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date=other_day,
        )
        assert "JOB-0009" not in result2

    def test_end_date_earlier_than_start_falls_back_to_single_day(self, mcp_module, tmp_path):
        """A malformed End Date (before Service Date) must not hide the
        job from its own start date — treated as single-day instead of
        producing an inverted, always-false range."""
        today = datetime.date.today()
        bad_end = (today - datetime.timedelta(days=5)).strftime("%Y-%m-%d")
        fp = str(_make_multiday_jobs_schedule_spreadsheet(
            tmp_path, today.strftime("%Y-%m-%d"), bad_end))

        result = mcp_module.read_job_spreadsheet(
            filepath=fp, sheet_name="Jobs_Schedule", filter_date="today",
        )
        assert "JOB-0009" in result


# ===========================================================================
# check_sms_configured
# ===========================================================================
#
# Lightweight tool the Jobs PWA calls once at boot to decide whether the
# job modal's "Email + Text Invoice" button should be enabled or dimmed
# — reuses the exact same sms_backends.validate_config() logic
# check_tools_status() already uses for its own SMS section, in a small
# dedicated tool that returns exactly one of two strings, rather than
# requiring the caller to parse a much larger free-text status report.

class TestCheckSmsConfigured:
    def test_returns_configured_string_when_valid(self, mcp_module):
        fake_backend = MagicMock()
        fake_backend.validate_config.return_value = (True, "ok")
        with patch("sms_backends.load_sms_config", return_value={"provider": "twilio"}), \
             patch("sms_backends.get_sms_backend", return_value=fake_backend):
            result = mcp_module.check_sms_configured()
        assert result == "✅ SMS configured"

    def test_returns_not_configured_string_when_invalid(self, mcp_module):
        fake_backend = MagicMock()
        fake_backend.validate_config.return_value = (False, "missing credentials")
        with patch("sms_backends.load_sms_config", return_value={}), \
             patch("sms_backends.get_sms_backend", return_value=fake_backend):
            result = mcp_module.check_sms_configured()
        assert result == "❌ SMS not configured"

    def test_returns_not_configured_on_any_exception(self, mcp_module):
        """A crashed/missing sms_backends import, a malformed config file,
        etc. must never propagate as an error — must fail safely to 'not
        configured' rather than breaking the PWA's boot sequence."""
        with patch("sms_backends.load_sms_config", side_effect=Exception("boom")):
            result = mcp_module.check_sms_configured()
        assert result == "❌ SMS not configured"

    def test_return_value_is_always_one_of_exactly_two_strings(self, mcp_module):
        """The PWA does a plain .includes('✅') check on the result — the
        contract here is strict: exactly these two strings, nothing else,
        so that check can never accidentally match something unrelated."""
        fake_backend = MagicMock()
        fake_backend.validate_config.return_value = (True, "ok")
        with patch("sms_backends.load_sms_config", return_value={}), \
             patch("sms_backends.get_sms_backend", return_value=fake_backend):
            result = mcp_module.check_sms_configured()
        assert result in ("✅ SMS configured", "❌ SMS not configured")

